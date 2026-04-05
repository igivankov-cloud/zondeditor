from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from .excel_import_detect import (
    MODE_BLOCKS_RIGHT,
    MODE_VERTICAL,
    ROLE_DEPTH,
    ROLE_FS,
    ROLE_IGNORE,
    infer_type_from_roles,
    try_parse_float,
)


@dataclass(slots=True)
class WorkbookSheet:
    name: str
    rows: list[list[Any]]


@dataclass(slots=True)
class WorkbookData:
    path: Path
    sheets: list[WorkbookSheet]


@dataclass(slots=True)
class ExcelImportConfig:
    mode: str
    header_row: int
    data_start_row: int
    column_roles: dict[int, str]
    repeat_first_block: bool = False
    sounding_names: list[str] = field(default_factory=list)


@dataclass(slots=True)
class ImportedRow:
    depth_m: float
    values: dict[str, float | None]


@dataclass(slots=True)
class ImportedSounding:
    source_name: str
    display_name: str
    rows: list[ImportedRow]


@dataclass(slots=True)
class ImportPreview:
    detected_type: int | None
    soundings: list[ImportedSounding]
    warnings: list[str]
    min_depth: float | None
    max_depth: float | None


class ExcelImportError(RuntimeError):
    pass


def _normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        txt = value.strip()
        return txt if txt else None
    return value


def _read_xlsx(path: Path) -> list[WorkbookSheet]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise ExcelImportError("Для чтения .xlsx установите зависимость openpyxl.") from exc

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    sheets: list[WorkbookSheet] = []
    for ws in wb.worksheets:
        rows: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append([_normalize_cell(v) for v in row])
        sheets.append(WorkbookSheet(name=ws.title, rows=rows))
    wb.close()
    return sheets


def _read_xls(path: Path) -> list[WorkbookSheet]:
    try:
        import xlrd  # type: ignore
    except Exception as exc:
        raise ExcelImportError("Для чтения .xls установите зависимость xlrd>=2.") from exc

    book = xlrd.open_workbook(str(path))
    sheets: list[WorkbookSheet] = []
    for sh in book.sheets():
        rows: list[list[Any]] = []
        for r in range(sh.nrows):
            rows.append([_normalize_cell(sh.cell_value(r, c)) for c in range(sh.ncols)])
        sheets.append(WorkbookSheet(name=sh.name, rows=rows))
    return sheets


def read_excel_workbook(path: str | Path) -> WorkbookData:
    p = Path(path)
    if not p.exists():
        raise ExcelImportError(f"Файл не найден: {p}")
    ext = p.suffix.lower()
    if ext == ".xlsx":
        sheets = _read_xlsx(p)
    elif ext == ".xls":
        sheets = _read_xls(p)
    else:
        raise ExcelImportError("Поддерживаются только .xls и .xlsx файлы.")
    if not sheets:
        raise ExcelImportError("В файле нет листов для импорта.")
    return WorkbookData(path=p, sheets=sheets)


def _row_is_header_like(row: list[Any], column_roles: dict[int, str]) -> bool:
    for col, role in column_roles.items():
        if role in (ROLE_IGNORE, ROLE_DEPTH):
            continue
        if col >= len(row):
            continue
        val = row[col]
        if isinstance(val, str) and val:
            if any(token in val.lower() for token in ("глуб", "depth", "qc", "fs", "лоб", "бок", "общ")):
                return True
    return False


def _first_text_for_name(rows: list[list[Any]], col_from: int, col_to: int, header_row_1based: int) -> str | None:
    search_rows = range(max(0, header_row_1based - 6), max(0, header_row_1based - 1))
    for ridx in reversed(list(search_rows)):
        row = rows[ridx] if ridx < len(rows) else []
        for c in range(col_from, col_to + 1):
            if c >= len(row):
                continue
            val = row[c]
            if isinstance(val, str) and val.strip():
                return val.strip()
    return None


def _parse_vertical(sheet: WorkbookSheet, config: ExcelImportConfig, fallback_name: str) -> tuple[list[ImportedSounding], list[str]]:
    depth_col = next((c for c, r in config.column_roles.items() if r == ROLE_DEPTH), None)
    if depth_col is None:
        raise ExcelImportError("Не найден столбец глубины.")

    rows_out: list[ImportedRow] = []
    warnings: list[str] = []
    for i in range(max(0, config.data_start_row - 1), len(sheet.rows)):
        row = sheet.rows[i]
        if _row_is_header_like(row, config.column_roles):
            continue
        depth = try_parse_float(row[depth_col] if depth_col < len(row) else None)
        if depth is None:
            continue
        vals: dict[str, float | None] = {}
        has_data = False
        for col, role in config.column_roles.items():
            if role in (ROLE_IGNORE, ROLE_DEPTH):
                continue
            val = try_parse_float(row[col] if col < len(row) else None)
            vals[role] = val
            if val is not None:
                has_data = True
        if not has_data:
            continue
        if any(v is None for v in vals.values()):
            warnings.append(f"Строка {i+1}: частично заполнена.")
        rows_out.append(ImportedRow(depth_m=depth, values=vals))

    if not rows_out:
        raise ExcelImportError("Не найдено строк данных для импорта.")

    src_name = _first_text_for_name(sheet.rows, 0, max(config.column_roles.keys(), default=0), config.header_row) or fallback_name
    sounding = ImportedSounding(source_name=src_name, display_name=src_name, rows=rows_out)
    return [sounding], warnings


def _block_signature(config: ExcelImportConfig, depth_col: int) -> tuple[int, dict[int, str]]:
    data_cols = sorted([c for c, role in config.column_roles.items() if role not in (ROLE_IGNORE, ROLE_DEPTH)])
    if not data_cols:
        raise ExcelImportError("В режиме «Блоки вправо» не задан первый блок.")
    block_start = min(data_cols)
    block_end = max(data_cols)
    width = block_end - block_start + 1
    rel = {c - block_start: config.column_roles[c] for c in range(block_start, block_end + 1) if config.column_roles.get(c, ROLE_IGNORE) not in (ROLE_IGNORE, ROLE_DEPTH)}
    if not rel:
        raise ExcelImportError("В режиме «Блоки вправо» не заданы роли столбцов первого блока.")
    return width, rel


def _parse_blocks_right(sheet: WorkbookSheet, config: ExcelImportConfig, fallback_name: str) -> tuple[list[ImportedSounding], list[str]]:
    depth_col = next((c for c, r in config.column_roles.items() if r == ROLE_DEPTH), None)
    if depth_col is None:
        raise ExcelImportError("Не найден столбец глубины.")

    width, rel_roles = _block_signature(config, depth_col)
    base_col = min(c for c, r in config.column_roles.items() if r not in (ROLE_IGNORE, ROLE_DEPTH))

    soundings: list[ImportedSounding] = []
    warnings: list[str] = []
    empty_blocks = 0
    block_idx = 0
    while True:
        block_origin = base_col + block_idx * width
        rows_out: list[ImportedRow] = []
        for i in range(max(0, config.data_start_row - 1), len(sheet.rows)):
            row = sheet.rows[i]
            if _row_is_header_like(row, config.column_roles):
                continue
            depth = try_parse_float(row[depth_col] if depth_col < len(row) else None)
            if depth is None:
                continue
            vals: dict[str, float | None] = {}
            has_data = False
            for rel_col, role in rel_roles.items():
                col = block_origin + rel_col
                val = try_parse_float(row[col] if col < len(row) else None)
                vals[role] = val
                if val is not None:
                    has_data = True
            if has_data:
                if any(v is None for v in vals.values()):
                    warnings.append(f"Блок {block_idx+1}, строка {i+1}: частично заполнена.")
                rows_out.append(ImportedRow(depth_m=depth, values=vals))

        if not rows_out:
            empty_blocks += 1
            if empty_blocks >= 2:
                break
            block_idx += 1
            continue

        empty_blocks = 0
        name = _first_text_for_name(sheet.rows, block_origin, block_origin + width - 1, config.header_row) or f"{fallback_name} {block_idx+1}"
        soundings.append(ImportedSounding(source_name=name, display_name=name, rows=rows_out))
        block_idx += 1
        if not config.repeat_first_block:
            break

    if not soundings:
        raise ExcelImportError("Не найдено валидных блоков данных справа.")
    return soundings, warnings


def make_unique_names(existing_names: set[str], desired: list[str]) -> list[str]:
    used = {x.strip() for x in existing_names if x and x.strip()}
    out: list[str] = []
    for raw in desired:
        base = (raw or "Зондировка").strip() or "Зондировка"
        candidate = base
        idx = 2
        while candidate in used:
            candidate = f"{base} ({idx})"
            idx += 1
        used.add(candidate)
        out.append(candidate)
    return out


def build_import_preview(sheet: WorkbookSheet, config: ExcelImportConfig, *, fallback_name: str) -> ImportPreview:
    if config.mode == MODE_BLOCKS_RIGHT:
        soundings, warnings = _parse_blocks_right(sheet, config, fallback_name)
    else:
        soundings, warnings = _parse_vertical(sheet, config, fallback_name)

    roles = set()
    for s in soundings:
        for row in s.rows:
            roles.update(row.values.keys())
    detected_type = infer_type_from_roles(roles)

    all_depths = [r.depth_m for s in soundings for r in s.rows]
    return ImportPreview(
        detected_type=detected_type,
        soundings=soundings,
        warnings=warnings,
        min_depth=min(all_depths) if all_depths else None,
        max_depth=max(all_depths) if all_depths else None,
    )
