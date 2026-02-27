# =====================================================================
# === FILE MAP BEGIN ===
# FILE MAP (обновляй при правках; указывай строки Lx–Ly)
# Правило:
#   - если в процессе правок обнаружил/затронул функцию — изучи и добавь её в FILE MAP.
#   - при сдвиге строк после изменений — обнови номера строк в FILE MAP.
# PATCHLOG (кратко; обновляй при крупных изменениях):
#   - v4.2.0: фиксированная шапка (hcanvas + canvas), синхронизированный горизонтальный скролл
#   - v4.2.6: GEO-template immutable + расширенный OBJ_geo_debug.txt (expected/actual/tests_current/prepared)
#   - v4.2.7: расширенный FILE MAP + кнопка «Карта» (просмотр FILE MAP в программе)
#   - v4.3.5: железный одинарный клик по qc/fs (сразу редактирование + выделение), навигация стрелками
#   - v4.3.6: закрытие активной ячейки при клике вне зондирования и при скроллинге
#   - v4.3.7: фикс кнопки «Корректировка» + убрана подсказка «Изм. начальную глубину»
#   - v4.4.0: хвост без разрывов — клик ниже конца только в следующую строку (без прокликивания дыр)
#   - v4.4.3: верх без разрывов — клик выше начала только в строку перед первой; вставка строки сверху
#   - v4.6.7: новая горизонтальная прокрутка (слайдер над подвалом), удалён старый hbar
# Окружение/стиль/лицензия/логи:
#   - _resource_path: L135–L139
#   - _apply_win11_style: L140–L169
#   - _programdata_dir: L170–L173
#   - _license_path: L174–L176
#   - _logs_dir: L177–L179
#   - _get_machine_guid: L180–L188
#   - _calc_machine_hash: L189–L192
#   - _write_license_file: L193–L202
#   - _check_license_or_exit: L203–L224
#   - _setup_shared_logger: L225–L236
#   - _log_event: L237–L241
#   - _open_logs_folder: L242–L249
# Утилиты данных/дат:
#   - _bcd: L250–L254
#   - _try_parse_dt: L255–L605
#   - _patch_dt_in_block: L269–L284
#   - _bcd_to_int: L543–L549
#   - _int_to_bcd: L550–L554
# Пересборка GEO:
#   - _rebuild_geo_from_template: L285–L434
# Иконки Win11 / глифы:
#   - _pick_icon_font: L435–L441
# Tooltip:
#   - ToolTip: L440–L540
# Диалог даты:
#   - _sanitize_int_0_300: L941–L951
#   - _format_date_ru: L952–L954
#   - CalendarDialog: L953–L1040
# Генерация/экспорт GXL:
#   - save_gxl_generated: L4464–L4884
# Главное приложение:
#   - GeoCanvasEditor: L1041–L4107
# GeoCanvasEditor — методы:
#   - __init__: L444–L1099
#   - _compute_depth_grid: L1100–L1169
#   - _snapshot: L1170–L1206
#   - _restore: L1207–L1260
#   - _push_undo: L1261–L1270
#   - undo: L1271–L1278
#   - redo: L1279–L1287
#   - _build_ui: L1288–L1526
#   - pick_file_and_load: L1527–L1549
#   - _parse_depth_step: L1550–L1577
#   - _update_status_loaded: L1578–L1595
#   - _normalize_test_lengths: L1596–L1631
#   - _apply_gxl_calibration_from_meta: L1632–L1652
#   - _calc_qc_fs_from_del: L1653–L1679
#   - _prompt_missing_geo_params: L1680–L1780
#   - _set_geo_inputs_enabled: L1781–L1798
#   - _depth_at: L1799–L1801
#   - load_and_render: L1802–L1973
#   - add_test: L1974–L2006
#   - convert_10_to_5: L2007–L2103
#   - _content_size: L2104–L2121
#   - _update_scrollregion: L2122–L2132
#   - _refresh_display_order: L2133–L2156
#   - _on_left_click: L2157–L2234
#   - _set_hover: L2253–L2261
#   - _hide_canvas_tip: L2262–L2276
#   - _schedule_canvas_tip: L2277–L2290
#   - _on_motion: L2291–L2313
#   - _delete_test: L2314–L2330
#   - _duplicate_test: L2331–L2404
#   - _cell_bbox: L2405–L2420
#   - _header_bbox: L2421–L2429
#   - _redraw: L2430–L2606
#   - _hit_test: L2607–L2669
#   - _on_double_click: L2670–L2699
#   - _on_right_click: L2738–L2759
#   - _ctx_delete_above: L2760–L2766
#   - _ctx_delete_below: L2767–L2773
#   - _ctx_delete_row: L2774–L2780
#   - _delete_by_display_row: L2781–L2837
#   - _delete_range_indices: L2838–L2899
#   - _edit_header: L2900–L3084
#   - _begin_edit: L3085–L3145
#   - _begin_edit_depth0: L3146–L3175
#   - _end_edit_depth0: L3176–L3227
#   - _end_edit: L3228–L3299
#   - _append_row: L3355–L3377
#   - _on_mousewheel: L3378–L3385
#   - _on_mousewheel_linux: L3386–L3392
#   - _choose_tail_k: L3393–L3400
#   - fix_by_algorithm: L3401–L3566
#   - _read_calc_params: L3567–L3584
#   - _safe_sheet_name: L3585–L3594
#   - export_excel: L3628–L3734
#   - _has_issues_for_fix_prompt: L3735–L3764
#   - export_credo_zip: L3765–L3868
#   - _center_child: L3869–L3883
#   - _place_calendar_near_header: L3884–L3903
#   - _ensure_object_code: L3904–L3948
#   - _extract_file_map_text: L3949–L3978
#   - show_file_map: L3979–L4025
#   - _on_close: L4026–L4038
#   - export_bundle: L4039–L4206
#   - _write_meta_txt: L4207–L4224
#   - _export_excel_silent: L4225–L4260
#   - _export_credo_silent: L4261–L4303
#   - save_file: L4304–L4344
#   - save_gxl: L4345–L4751# === FILE MAP END ===
# =====================================================================
import datetime
import traceback
import struct

# --- GEO rebuild helpers (preserve original headers/params) ---

from pathlib import Path
import os
import sys
import json
import hashlib
import platform
import getpass
import logging
from logging.handlers import RotatingFileHandler

# --- Win11 look + icon helpers (added) ---
APP_NAME = "ZondEditor"
APP_VER = "2.1"
APP_TITLE = f"{APP_NAME} — v{APP_VER}"

def _resource_path(rel: str) -> str:
    import os, sys
    base = getattr(sys, "_MEIPASS", os.path.abspath("."))
    return os.path.join(base, rel)

def _apply_win11_style(root):
    # Theme similar to Windows 11 (optional dependency)
    try:
        import sv_ttk
        sv_ttk.set_theme("light")
    except Exception:
        pass
    # Default font
    try:
        from tkinter import font
        f = font.nametofont("TkDefaultFont")
        f.configure(family="Segoe UI", size=10)
    except Exception:
        pass
    # Nice padding for ttk buttons
    try:
        import tkinter.ttk as ttk
        s = ttk.Style()
        s.configure("TButton", padding=(12, 8))
    except Exception:
        pass
    # AppUserModelID (helps Windows taskbar icon grouping)
    try:
        import ctypes
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("ZondEditor.SZ")
    except Exception:
        pass
APP_NAME = "ZondEditor"
_LICENSE_SALT = "ZOND-V1::static-sounding"

def _programdata_dir() -> Path:
    base = os.environ.get("PROGRAMDATA", r"C:\ProgramData")
    return Path(base) / APP_NAME

def _license_path() -> Path:
    return _programdata_dir() / "license.dat"

def _logs_dir() -> Path:
    return _programdata_dir() / "logs"

def _get_machine_guid() -> str:
    try:
        import winreg
        with winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\Cryptography") as k:
            guid, _ = winreg.QueryValueEx(k, "MachineGuid")
            return str(guid)
    except Exception:
        return platform.node()

def _calc_machine_hash() -> str:
    raw = (_get_machine_guid() + "|" + _LICENSE_SALT).encode("utf-8", "ignore")
    return hashlib.sha256(raw).hexdigest()

def _write_license_file() -> None:
    d = _programdata_dir()
    d.mkdir(parents=True, exist_ok=True)
    lp = _license_path()
    data = {
        "machine_hash": _calc_machine_hash(),
        "created_at": datetime.datetime.now().isoformat(timespec="seconds"),
    }
    lp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

def _check_license_or_exit(messagebox) -> None:
    lp = _license_path()
    if not lp.exists():
        messagebox.showerror(
            "Нет доступа",
            "Лицензионный файл не найден.\n"
            "Обратитесь к администратору (license.dat)."
        )
        raise SystemExit(1)
    try:
        data = json.loads(lp.read_text(encoding="utf-8"))
    except Exception:
        messagebox.showerror("Ошибка", "Лицензионный файл повреждён.")
        raise SystemExit(1)
    if data.get("machine_hash") != _calc_machine_hash():
        messagebox.showerror(
            "Нет доступа",
            "Лицензия не подходит для этого компьютера.\n"
            "Если программу скопировали на другой ПК — она не запустится."
        )
        raise SystemExit(1)

def _setup_shared_logger() -> logging.Logger:
    log_dir = _logs_dir()
    log_dir.mkdir(parents=True, exist_ok=True)
    log_path = log_dir / "usage.log"
    logger = logging.getLogger("usage")
    logger.setLevel(logging.INFO)
    if not logger.handlers:
        h = RotatingFileHandler(str(log_path), maxBytes=2_000_000, backupCount=10, encoding="utf-8")
        h.setFormatter(logging.Formatter("%(asctime)s\t%(message)s"))
        logger.addHandler(h)
    return logger

def _log_event(logger: logging.Logger, kind: str, **fields):
    user = getpass.getuser()
    parts = [f"{k}={v}" for k, v in fields.items()]
    logger.info(f"{kind}\tuser={user}\t" + "\t".join(parts))

def _open_logs_folder():
    try:
        d = _logs_dir()
        d.mkdir(parents=True, exist_ok=True)
        os.startfile(str(d))
    except Exception:
        pass

def _bcd(n: int) -> int:
    n = int(n)
    return ((n // 10) << 4) | (n % 10)


def _try_parse_dt(s: str):
    try:
        from datetime import datetime
        s = str(s).strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M"):
            try:
                return datetime.strptime(s, fmt)
            except Exception:
                pass
    except Exception:
        pass
    return None


def _canvas_view_bbox(cnv: tk.Canvas):
    """(vx0, vy0, vx1, vy1) of currently visible area in canvas coordinates."""
    try:
        vx0 = cnv.canvasx(0)
        vy0 = cnv.canvasy(0)
        w = max(1, int(cnv.winfo_width() or 1))
        h = max(1, int(cnv.winfo_height() or 1))
        return vx0, vy0, vx0 + w, vy0 + h
    except Exception:
        return 0.0, 0.0, 0.0, 0.0


def _patch_dt_in_block(raw: bytes, dt_offset: int, dt):
    if dt is None:
        return raw
    b = bytearray(raw)
    try:
        b[dt_offset + 0] = _bcd(dt.second)
        b[dt_offset + 1] = _bcd(dt.minute)
        b[dt_offset + 2] = _bcd(dt.hour)
        b[dt_offset + 3] = _bcd(dt.day)
        b[dt_offset + 4] = _bcd(dt.month)
        b[dt_offset + 5] = _bcd(dt.year % 100)
    except Exception:
        pass
    return bytes(b)


def _rebuild_geo_from_template(original_bytes: bytes, blocks_info: list, prepared_tests: list) -> bytes:
    """Rebuild GEO blocks preserving each block header, but allow variable data length.

    ВАЖНО (фикс "воскресшего первого зондирования"):
    - Раньше id патчился по смещению id_off из blocks_info. Для некоторых файлов (особенно первый блок)
      смещение может быть рассчитано неверно, из-за чего id не менялся и "воскресал" исходный номер.
    - Теперь делаем два шага: (1) пробуем патч по id_off, (2) обязательно делаем надёжный патч по маркеру
      FF FF <id> 1E 0A внутри блока.

    Также добавлен выбор шаблонного блока по orig_id (если доступен), чтобы уменьшить шанс смешения форматов.
    """
    if not blocks_info:
        return original_bytes

    # prepare raw blocks (+ попытка вытащить базовый id из маркера)
    blocks = []
    id_to_block = {}
    for bi in blocks_info:
        raw = original_bytes[bi.header_start:bi.data_end]
        base_id = None
        try:
            # найти маркер FF FF <id> 1E 0A в пределах блока
            for j in range(0, max(0, len(raw) - 5)):
                if raw[j] == 0xFF and raw[j + 1] == 0xFF and raw[j + 3] == 0x1E and raw[j + 4] == 0x0A:
                    base_id = raw[j + 2]
                    break
        except Exception:
            base_id = None

        item = {
            'raw': raw,
            'id_off': bi.id_pos - bi.header_start,
            'dt_off': bi.dt_pos - bi.header_start,
            'data_off': bi.data_start - bi.header_start,
            'data_len': bi.data_end - bi.data_start,
            'base_id': base_id,
        }
        blocks.append(item)
        if base_id is not None and base_id not in id_to_block:
            id_to_block[base_id] = item

    def _patch_id_in_block(buf: bytearray, tid: int) -> None:
        """Надёжно заменить <id> в маркере FF FF <id> 1E 0A (первая встреча)."""
        try:
            n = len(buf)
            for j in range(0, max(0, n - 5)):
                if buf[j] == 0xFF and buf[j + 1] == 0xFF and buf[j + 3] == 0x1E and buf[j + 4] == 0x0A:
                    buf[j + 2] = tid
                    return
        except Exception:
            return

    out_parts = []
    for i, t in enumerate(prepared_tests):
        # choose template block
        base = None
        try:
            oid = getattr(t, 'orig_id', None)
            if oid is not None:
                try:
                    oid_int = int(oid)
                except Exception:
                    oid_int = None
                if oid_int is not None:
                    base = id_to_block.get(oid_int)
        except Exception:
            base = None

        if base is None:
            base = blocks[i] if i < len(blocks) else blocks[-1]

        raw0 = base['raw']
        b = bytearray(raw0)

        # patch id (двухходовый: смещение + маркер)
        tid = int(getattr(t, 'tid', i + 1) or (i + 1))
        tid = max(1, min(255, tid))
        try:
            off = int(base.get('id_off', -1))
            if 0 <= off < len(b):
                b[off] = tid
        except Exception:
            pass
        _patch_id_in_block(b, tid)

        # patch datetime (BCD) — по смещению, как раньше
        dt = _try_parse_dt(getattr(t, 'dt', ''))
        b = bytearray(_patch_dt_in_block(bytes(b), int(base['dt_off']), dt))

        data_off = int(base['data_off'])
        data_len = int(base['data_len'])

        qc_in = list(getattr(t, 'qc', []) or [])
        fs_in = list(getattr(t, 'fs', []) or [])

        # ВАЖНО: НЕ ограничиваемся исходной длиной data_len.
        # Если пользователь «дотянул хвост» или при конвертации шага появились новые строки,
        # их нужно физически записать в GEO (блок станет длиннее).
        n_pairs = min(len(qc_in), len(fs_in))

        # rewrite: header + pairs(n_pairs) + tail(after original data region)
        head = bytes(b[:data_off])
        # tail берём ПОСЛЕ исходного data_len (служебные байты блока, если есть)
        tail = bytes(b[data_off + data_len:]) if (data_off + data_len) <= len(b) else b''

        data = bytearray()
        for k in range(n_pairs):
            try:
                qv = int(float(str(qc_in[k]).replace(',', '.'))) if str(qc_in[k]).strip() != '' else 0
            except Exception:
                qv = 0
            try:
                fv = int(float(str(fs_in[k]).replace(',', '.'))) if str(fs_in[k]).strip() != '' else 0
            except Exception:
                fv = 0
            qv = max(0, min(255, qv))
            fv = max(0, min(255, fv))
            data.append(qv)
            data.append(fv)

        out_parts.append(head + bytes(data) + tail)

    first = blocks_info[0].header_start
    last = max(bi.data_end for bi in blocks_info)
    return original_bytes[:first] + b''.join(out_parts) + original_bytes[last:]
APP_TITLE = "ZondEditor SZ v2.0"
APP_VERSION = "2.0"

import re
from dataclasses import dataclass
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import random
import datetime as _dt
import datetime
import copy
import xml.etree.ElementTree as ET

import openpyxl
from openpyxl.styles import PatternFill
import shutil
import zipfile
import tempfile
import re





# --- Fluent/Win11-like icons (Segoe Fluent Icons / Segoe MDL2 Assets) ---
ICON_FONT_FALLBACKS = ("Segoe Fluent Icons", "Segoe MDL2 Assets", "Segoe UI Symbol")
ICON_COPY = "\uE8C8"    # Copy
ICON_DELETE = "\uE74D"  # Delete/Trash
ICON_CALENDAR = "\uE787"  # Calendar

def _pick_icon_font(size: int = 12):
    # Tk will fall back silently if a family is missing; we keep the preferred order.
    return (ICON_FONT_FALLBACKS[0], size)

# Русские названия месяцев для диалога выбора даты (не зависит от системной locale)
_MONTHS_RU = ['Январь','Февраль','Март','Апрель','Май','Июнь','Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь']

class ToolTip:
    """Simple tooltip for Tk widgets."""
    def __init__(self, widget, text: str, delay_ms: int = 350):
        self.widget = widget
        self.text = text
        self.delay_ms = delay_ms
        self._after = None
        self._tip = None
        widget.bind("<Enter>", self._on_enter, add="+")
        widget.bind("<Leave>", self._on_leave, add="+")
        widget.bind("<ButtonPress>", self._on_leave, add="+")

    def _on_enter(self, _e=None):
        self._schedule()

    def _on_leave(self, _e=None):
        self._unschedule()
        self._hide()

    def _schedule(self):
        self._unschedule()
        self._after = self.widget.after(self.delay_ms, self._show)

    def _unschedule(self):
        if self._after is not None:
            try:
                self.widget.after_cancel(self._after)
            except Exception:
                pass
            self._after = None

    def _show(self):
        if self._tip or not self.text:
            return
        try:
            x = self.widget.winfo_rootx() + 12
            y = self.widget.winfo_rooty() + self.widget.winfo_height() + 8
        except Exception:
            return
        self._tip = tk.Toplevel(self.widget)
        self._tip.wm_overrideredirect(True)
        try:
            self._tip.attributes("-topmost", True)
        except Exception:
            pass
        lbl = tk.Label(
            self._tip, text=self.text, justify="left",
            background="#FFFFE0", relief="solid", borderwidth=1,
            font=("Segoe UI", 9)
        )
        lbl.pack(ipadx=6, ipady=3)
        self._tip.wm_geometry(f"+{x}+{y}")

    def _hide(self):
        if self._tip:
            try:
                self._tip.destroy()
            except Exception:
                pass
            self._tip = None

# ============================================================
# GEO Viewer + qc/fs редактор (Canvas, раскраска ячеек) → Excel
# Версия: v2.1
#
# Новое:
# - Удаление первой/последней строки: ПКМ → контекстное меню "Удалить строку"
#   (строка реально удаляется, можно удалять дальше; в Excel удалённые строки
#   не попадают).
# - Верхняя панель кнопок: Загрузить GEO, Сохранить GEO, Назад, Вперёд.
# - Сохранение обратно в .GEO: сохраняем структуру исходного файла и кодировку,
#   заменяем только:
#     * id опыта (1 байт в заголовке)
#     * дату/время (6 BCD-байт после маркера)
#     * пары (qc,fs) в блоке данных
#   ВНИМАНИЕ: сохранение поддерживает только опыты, существующие в исходном файле.
#   Если добавлены новые опыты через "+", они сохраняются только в Excel.
#
# Остальное:
# - Шаг в см: 5 / 10
# - Кнопка "10см → 5см" (работает когда текущий шаг 10 см)
# - Редактирование №/дата-время: двойной клик по шапке
# - Подсветка ячеек как в v2.3.1 (алгоритм "Исправить")
# ============================================================

HEADER_RE = re.compile(b"\xFF\xFF(?:\xFF\xFF)?(.)\x1E(\x0A|\x14)")

FILL_YELLOW = PatternFill("solid", fgColor="FFF7D6")  # light yellow
FILL_RED    = PatternFill("solid", fgColor="FFD6D6")  # light red
FILL_BLUE   = PatternFill("solid", fgColor="D6E8FF")  # light blue
FILL_PURPLE = PatternFill("solid", fgColor="E8D6FF")  # light purple (manual edit)

GUI_YELLOW = "#FFF7D6"
GUI_RED    = "#FFD6D6"
GUI_ORANGE = "#FFB347"  # problems (scan/attention) - более насыщенный оранжевый
GUI_BLUE   = "#D6E8FF"
GUI_PURPLE = "#E8D6FF"
GUI_GREEN  = "#D6FFD6"  # light green (algorithmic correction)

# Preview (autocheck) pale colors
GUI_YELLOW_P = "#FFFDF0"
GUI_RED_P    = "#FFECEC"
GUI_ORANGE_P = "#FFE6BF"
GUI_BLUE_P   = "#EEF6FF"
GUI_GRID   = "#D0D0D0"
GUI_HDR = "#D6E8FF"
GUI_DEPTH_BG = "#F2F2F2"


def _bcd_to_int(b: int) -> int:
    hi, lo = (b >> 4) & 0xF, b & 0xF
    if hi > 9 or lo > 9:
        return -1
    return hi * 10 + lo


def _int_to_bcd(n: int) -> int:
    n = max(0, min(99, int(n)))
    return ((n // 10) << 4) | (n % 10)


def _parse_datetime_bcd(data: bytes, pos: int) -> str | None:
    if pos + 6 > len(data):
        return None
    ss = _bcd_to_int(data[pos + 0])
    mm = _bcd_to_int(data[pos + 1])
    HH = _bcd_to_int(data[pos + 2])
    DD = _bcd_to_int(data[pos + 3])
    MO = _bcd_to_int(data[pos + 4])
    YY = _bcd_to_int(data[pos + 5])
    if -1 in (ss, mm, HH, DD, MO, YY):
        return None
    if not (0 <= ss <= 59 and 0 <= mm <= 59 and 0 <= HH <= 23 and 1 <= DD <= 31 and 1 <= MO <= 12):
        return None
    year = 2000 + YY
    return f"{year:04d}-{MO:02d}-{DD:02d} {HH:02d}:{mm:02d}:{ss:02d}"


def _try_parse_dt(text) -> _dt.datetime | None:
    """Пробует разобрать дату/время.
    Принимает str или datetime (иногда t.dt уже хранится как datetime после редактирования).
    """
    if text is None:
        return None
    if isinstance(text, _dt.datetime):
        return text
    if isinstance(text, _dt.date):
        return _dt.datetime(text.year, text.month, text.day, 0, 0, 0)

    s = str(text).strip()
    if not s:
        return None

    # supported: YYYY-MM-DD HH:MM:SS / YYYY-MM-DD HH:MM / DD.MM.YYYY HH:MM:SS / DD.MM.YYYY HH:MM
    fmts = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
    ]
    for f in fmts:
        try:
            return _dt.datetime.strptime(s, f)
        except Exception:
            pass

    # fallback regex
    m = re.match(r"(\d{4})-(\d{1,2})-(\d{1,2})\s+(\d{1,2}):(\d{2})(?::(\d{2}))?", s)
    if m:
        y, mo, d, hh, mm, ss = m.groups()
        return _dt.datetime(int(y), int(mo), int(d), int(hh), int(mm), int(ss or 0))

    m = re.match(r"(\d{1,2})\.(\d{1,2})\.(\d{4})\s+(\d{1,2}):(\d{2})(?::(\d{2}))?", s)
    if m:
        d, mo, y, hh, mm, ss = m.groups()
        return _dt.datetime(int(y), int(mo), int(d), int(hh), int(mm), int(ss or 0))

    return None


def _cm2_to_m2(a_cm2: float) -> float:
    return a_cm2 * 1e-4


def _parse_cell_int(txt: str) -> int | None:
    s = (txt or "").strip()
    if s == "":
        return None
    s2 = "".join(ch for ch in s if (ch.isdigit() or ch in ".-,"))
    if s2 == "":
        return None
    try:
        return int(float(s2.replace(",", ".")))
    except Exception:
        return None


def _parse_depth_float(txt: str) -> float | None:
    s = (txt or "").strip().replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except Exception:
        return None


def _max_zero_run(vals: list[int]) -> int:
    best = cur = 0
    for v in vals:
        if v == 0:
            cur += 1
            best = max(best, cur)
        else:
            cur = 0
    return best


def _interp_with_noise(a: int, b: int, t: float) -> int:
    v = a + (b - a) * t
    v = int(round(v))
    if random.random() < 0.8:
        v += random.choice([-3, -2, -1, 1, 2, 3])
    return max(1, min(250, v))  # не возвращаем 0


def _noise_around(v: int) -> int:
    v2 = v + random.choice([-2, -1, 1, 2]) if random.random() < 0.85 else v
    return max(1, min(250, v2))


@dataclass
class TestFlags:
    invalid: bool
    interp_cells: set[tuple[int, str]]   # (row, \'qc\'/\'fs\')
    force_cells: set[tuple[int, str]]    # (row, \'qc\'/\'fs\')
    user_cells: set[tuple[int, str]]     # (row, \'qc\'/\'fs\') manual edits (purple)
    algo_cells: set[tuple[int, str]]     # (row, 'qc'/'fs') changed by algorithm (green)
    force_tail_rows: set[int]            # rows (grid index) suggested to add below tail (blue row)


@dataclass
class GeoBlockInfo:
    order_index: int
    header_start: int
    header_end: int
    id_pos: int          # absolute pos of id byte
    dt_pos: int          # absolute pos of first datetime BCD byte (6 bytes)
    data_start: int      # absolute pos of first qc/fs byte
    data_end: int        # absolute pos of end of block (start of next header or EOF)
    marker_byte: int


@dataclass
class TestData:
    tid: int
    dt: str
    depth: list[str]   # meters as string
    qc: list[str]
    fs: list[str]
    incl: list[str] = None  # K4: inclinometer / U channel (may be all zeros)
    marker: str = ""
    header_pos: str = ""
    # binding to original GEO (for save-back)
    orig_id: int | None = None
    block: GeoBlockInfo | None = None



def _decode_xml_bytes(b: bytes) -> str:
    # GeoExplorer GXL is XML, often with CP1251 payload.
    for enc in ("utf-8", "cp1251", "windows-1251", "utf-16", "latin1"):
        try:
            return b.decode(enc)
        except Exception:
            continue
    return b.decode("utf-8", errors="ignore")




def _xml_escape(s: str) -> str:
    return (s or '').replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('"','&quot;').replace("'",'&apos;')
def parse_gxl_file(path: Path):
    """Parse GeoExplorer .gxl export (XML). Returns (tests, meta_rows)."""
    xml_text = _decode_xml_bytes(path.read_bytes())
    root = ET.fromstring(xml_text)

    obj = root.find(".//object")
    meta_rows = []

    if obj is not None:
        def tx(tag):
            v = obj.findtext(tag)
            if v is None:
                v = obj.findtext(".//" + tag)
            return (v or "").strip()

        meta_rows.append({"key": "source", "value": "gxl"})
        meta_rows.append({"key": "name", "value": tx("name") or tx("FullName")})
        meta_rows.append({"key": "NumArch", "value": tx("NumArch")})
        meta_rows.append({"key": "Cashman", "value": tx("Cashman")})
        meta_rows.append({"key": "Appendix", "value": tx("Appendix")})
        meta_rows.append({"key": "scale", "value": tx("scale")})
        meta_rows.append({"key": "scaleostria", "value": tx("scaleostria")})
        meta_rows.append({"key": "scalemufta", "value": tx("scalemufta")})

    tests = []
    if obj is None:
        return tests, meta_rows

    for t in obj.findall("test"):
        num = (t.findtext("numtest") or "").strip()
        if not num:
            continue
        try:
            tid = int(float(num.replace(",", ".")))
        except Exception:
            continue

        date_s = (t.findtext("date") or "").strip()
        time_s = (t.findtext("time") or "").strip()
        dt = (f"{date_s} {time_s}".strip()) if (date_s or time_s) else ""

        deepbegin = (t.findtext("deepbegin") or "").strip()
        stepzond = (t.findtext("stepzond") or "").strip()
        try:
            d0 = float(deepbegin.replace(",", "."))
        except Exception:
            d0 = 0.0
        try:
            step = float(stepzond.replace(",", "."))
        except Exception:
            step = 0.05

        dat = t.findtext("dat") or ""
        lines = [ln.strip() for ln in dat.splitlines() if ln.strip()]
        qc = []
        fs = []
        for ln in lines:
            parts = ln.split(";")
            q = parts[0].strip() if len(parts) > 0 else ""
            f = parts[1].strip() if len(parts) > 1 else ""
            qc.append(q)
            fs.append(f)

        depth = [f"{(d0 + step*i):.2f}" for i in range(len(qc))]
        tests.append(TestData(tid=tid, dt=dt, qc=qc, fs=fs, depth=depth, block=None))

    return tests, meta_rows


# --- K4 GEO parser (Geotest K4) ---
# Delegated to src/zondeditor/io/k4_reader.py to start modularization.
from src.zondeditor.io.k4_reader import (
    K4_SIG,
    detect_geo_kind as _detect_geo_kind_mod,
    parse_k4_geo_strict as _parse_k4_mod,
)

def detect_geo_kind(data: bytes) -> str:
    return _detect_geo_kind_mod(data)

def parse_k4_geo_strict(data: bytes) -> list[TestData]:
    # Keep original return type (list[TestData])
    return _parse_k4_mod(data, TestData)



def parse_geo_with_blocks(data: bytes) -> tuple[list[TestData], list[dict]]:
    """Parse geo and return tests + meta rows with block infos for save-back."""
    headers = []
    for m in HEADER_RE.finditer(data):
        hs = m.start()
        # detect id position depending on optional extra FF FF
        # pattern: FF FF [FF FF]? id 1E marker
        # if there is extra FF FF, id byte is at hs+4, else hs+2
        id_pos = hs + (4 if data[hs:hs+4] == b"\xFF\xFF\xFF\xFF" else 2)
        test_id = data[id_pos]
        marker_pos = id_pos + 2  # id + 1E
        marker = data[marker_pos]
        header_end = m.end()
        dt_pos = header_end  # after marker
        dt = _parse_datetime_bcd(data, dt_pos)
        headers.append((hs, header_end, id_pos, dt_pos, test_id, marker))

    if not headers:
        raise ValueError(
            "Не найдены заголовки опытов.\n"
            "Ожидались маркеры: FF FF <id> 1E 0A или FF FF <id> 1E 14."
        )

    tests_out: list[TestData] = []
    meta_rows: list[dict] = []

    for i, (hs, header_end, id_pos, dt_pos, test_id, marker) in enumerate(headers):
        second = data.find(b"\xFF\xFF", header_end)
        if second == -1:
            continue
        data_start = second + 2
        data_end = headers[i + 1][0] if i + 1 < len(headers) else len(data)
        block = data[data_start:data_end]
        if len(block) < 2:
            pairs = []
        else:
            if len(block) % 2 == 1:
                block = block[:-1]
            pairs = [(block[j], block[j + 1]) for j in range(0, len(block), 2)]

        bid = GeoBlockInfo(
            order_index=i,
            header_start=hs,
            header_end=header_end,
            id_pos=id_pos,
            dt_pos=dt_pos,
            data_start=data_start,
            data_end=data_end,
            marker_byte=marker,
        )

        dt_str = _parse_datetime_bcd(data, dt_pos) or ""
        t = TestData(
            tid=int(test_id),
            dt=dt_str,
            depth=[],
            qc=[str(int(p[0])) for p in pairs],
            fs=[str(int(p[1])) for p in pairs],
            marker=f"0x{marker:02X}",
            header_pos=str(hs),
            orig_id=int(test_id),
            block=bid,
        )
        tests_out.append(t)
        meta_rows.append({
            "test_id": int(test_id),
            "datetime": dt_str,
            "marker": f"0x{marker:02X}",
            "header_pos": hs,
            "points": len(pairs),
        })

    if not tests_out:
        raise ValueError("Опытов не извлечено. Возможно другой вариант кодирования данных в GEO.")
    return tests_out, meta_rows




# ---------------- UI helpers: validation + calendar ----------------

def _only_digits(s: str) -> bool:
    return s.isdigit() if s is not None else False

def _validate_tid_key(p: str) -> bool:
    # allow empty while typing
    if p == "":
        return True
    return p.isdigit() and len(p) <= 3

def _validate_hh_key(p: str) -> bool:
    if p == "":
        return True
    if not p.isdigit() or len(p) > 2:
        return False
    try:
        v = int(p)
    except Exception:
        return False
    return 0 <= v <= 23

def _validate_mm_key(p: str) -> bool:
    if p == "":
        return True
    if not p.isdigit() or len(p) > 2:
        return False
    try:
        v = int(p)
    except Exception:
        return False
    return 0 <= v <= 59

def _validate_int_0_300_key(p: str) -> bool:
    # qc/fs cells: allow blank, digits only, 0..300
    if p == "":
        return True
    if not p.isdigit() or len(p) > 3:
        return False
    try:
        v = int(p)
    except Exception:
        return False
    return 0 <= v <= 300

def _validate_nonneg_float_key(p: str) -> bool:
    """Allow non-negative float typing (digits, dot/comma). Empty is allowed."""
    if p is None:
        return True
    s = p.strip()
    if s == "":
        return True
    s2 = s.replace(',', '.')
    if s2.startswith('.'):
        s2 = '0' + s2
    if not re.fullmatch(r"\d+(?:\.\d*)?", s2):
        return False
    try:
        v = float(s2)
    except Exception:
        return False
    return v >= 0.0


def _is_depth_on_5cm_grid(v_m: float) -> bool:
    """True if depth is multiple of 0.05 m (i.e., cm ends with 0 or 5)."""
    cm = round(v_m * 100.0)
    if abs(v_m * 100.0 - cm) > 1e-6:
        return False
    return (cm % 5) == 0


def _validate_depth_0_4_key(p: str) -> bool:
    """Typing validator for depth0: 0..4 m, non-negative float."""
    if p is None:
        return True
    s = p.strip()
    if s == "":
        return True
    if not _validate_nonneg_float_key(s):
        return False
    try:
        v = float(s.replace(',', '.'))
    except Exception:
        return False
    return 0.0 <= v <= 4.0


def _sanitize_int_0_300(val: str) -> str:
    s = (val or "").strip()
    if s == "":
        return ""
    try:
        v = int(float(s.replace(",", ".")))
    except Exception:
        return ""
    v = max(0, min(300, v))
    return str(v)

def _format_date_ru(d: _dt.date) -> str:
    return f"{d.day:02d}.{d.month:02d}.{d.year:04d}"

class CalendarDialog(tk.Toplevel):
    """Minimal calendar dialog (no external deps). Returns selected date via self.selected.
    Ограничение: нельзя выбирать/листать даты из будущего (выше сегодняшней).
    """
    def __init__(self, parent, initial: _dt.date | None = None, title: str = "Выбор даты"):
        super().__init__(parent)
        self.title(title)
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()

        self.selected: _dt.date | None = None
        self._max_date = _dt.date.today()

        self._cur = initial or self._max_date
        if self._cur > self._max_date:
            self._cur = self._max_date

        self._view_year = self._cur.year
        self._view_month = self._cur.month

        self._hdr = ttk.Frame(self)
        self._hdr.pack(fill="x", padx=10, pady=(10, 0))

        self._btn_prev = ttk.Button(self._hdr, text="◀", width=3, command=self._prev_month)
        self._btn_prev.pack(side="left")
        self._lbl = ttk.Label(self._hdr, text="", width=18, anchor="center")
        self._lbl.pack(side="left", padx=6)
        self._btn_next = ttk.Button(self._hdr, text="▶", width=3, command=self._next_month)
        self._btn_next.pack(side="left")

        self._grid = ttk.Frame(self)
        self._grid.pack(padx=10, pady=10)

        self._build()

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=10, pady=(0, 10))
        ttk.Button(btns, text="Сегодня", command=self._pick_today).pack(side="left")
        ttk.Button(btns, text="Отмена", command=self._cancel).pack(side="right")

        self.bind("<Escape>", lambda _e: self._cancel())

    def _prev_month(self):
        m = self._view_month - 1
        y = self._view_year
        if m == 0:
            m = 12
            y -= 1
        self._view_year, self._view_month = y, m
        self._build()

    def _next_month(self):
        # не даём листать в будущие месяцы
        max_y, max_m = self._max_date.year, self._max_date.month
        if (self._view_year, self._view_month) >= (max_y, max_m):
            return
        m = self._view_month + 1
        y = self._view_year
        if m == 13:
            m = 1
            y += 1
        # ещё раз проверим
        if (y, m) > (max_y, max_m):
            return
        self._view_year, self._view_month = y, m
        self._build()

    def _pick_today(self):
        self.selected = self._max_date
        self.destroy()

    def _cancel(self):
        self.selected = None
        self.destroy()

    def _select(self, day: int):
        try:
            d = _dt.date(self._view_year, self._view_month, day)
            if d > self._max_date:
                return  # будущее запрещено
            self.selected = d
        except Exception:
            self.selected = None
        self.destroy()

    def _build(self):
        import calendar as _cal
        for w in self._grid.winfo_children():
            w.destroy()

        month_name = f"{_MONTHS_RU[self._view_month-1]} {self._view_year}"
        self._lbl.config(text=month_name)

        # disable next button if at max month
        max_y, max_m = self._max_date.year, self._max_date.month
        try:
            self._btn_next.config(state=("disabled" if (self._view_year, self._view_month) >= (max_y, max_m) else "normal"))
        except Exception:
            pass

        headers = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
        for c, h in enumerate(headers):
            ttk.Label(self._grid, text=h, width=3, anchor="center").grid(row=0, column=c, padx=1, pady=(0, 2))

        cal = _cal.Calendar(firstweekday=_cal.MONDAY)
        weeks = cal.monthdayscalendar(self._view_year, self._view_month)

        # выбранная дата (для подсветки)
        sel = self._cur
        if sel and (sel.year, sel.month) != (self._view_year, self._view_month):
            sel = None

        for r, week in enumerate(weeks, start=1):
            for c, day in enumerate(week):
                if day == 0:
                    ttk.Label(self._grid, text=" ", width=3).grid(row=r, column=c, padx=1, pady=1)
                    continue

                d = _dt.date(self._view_year, self._view_month, day)
                is_future = d > self._max_date
                is_sel = (sel is not None and d == sel)

                # tk.Button для управляемых цветов (ttk плохо красится на Win)
                bg = "#1e6bd6" if is_sel else None
                fg = "white" if is_sel else None

                b = tk.Button(
                    self._grid,
                    text=str(day),
                    width=3,
                    relief="ridge",
                    bd=1,
                    command=(lambda dd=day: self._select(dd)),
                    state=("disabled" if is_future else "normal"),
                )
                if is_sel:
                    b.config(background=bg, foreground=fg, activebackground=bg, activeforeground=fg)
                b.grid(row=r, column=c, padx=1, pady=1)

class GeoCanvasEditor(tk.Tk):
    def __init__(self):
        super().__init__()
        _apply_win11_style(self)
        self.title(APP_TITLE)
        try:
            self.iconbitmap(_resource_path('SZ_icon_adaptive.ico'))
        except Exception:
            pass
        # Центрируем главное окно на экране

        W0, H0 = 1480, 880

        try:

            sw = self.winfo_screenwidth()

            sh = self.winfo_screenheight()

            x = max((sw - W0) // 2, 0)

            y = max((sh - H0) // 2, 0)

            self.geometry(f"{W0}x{H0}+{x}+{y}")

        except Exception:

            self.geometry("1480x880")


        # --- offline machine-bound license + shared logs ---
        from tkinter import messagebox
        _check_license_or_exit(messagebox)
        self.usage_logger = _setup_shared_logger()
        self.bind_all("<F12>", lambda e: _open_logs_folder())

        self.geo_path: Path | None = None
        self.is_gxl = False
        self.geo_kind = 'K2'
        self.original_bytes: bytes | None = None

        # template block infos from original GEO (needed to rebuild GEO correctly after deletions)
        self._geo_template_blocks_info = []  # legacy: blocks info (do not mutate on edits)
        self._geo_template_blocks_info_full = []  # immutable template blocks info from original GEO
        self.meta_rows: list[dict] = []
        self.tests: list[TestData] = []
        self.flags: dict[int, TestFlags] = {}
        self.depth_start: float | None = None
        self.step_m: float | None = None
        self._depth_confirmed = False
        self._step_confirmed = False

        # Для GEO: индивидуальная начальная глубина по каждому опыту (tid -> h0)
        # Если не задано — используется self.depth_start.
        self.depth0_by_tid = {}

        self.undo_stack: list[dict] = []
        self.redo_stack: list[dict] = []

        self._dirty = False

        # Algorithm preview mode (autocheck on open): use pale colors, no data modification
        self._algo_preview_mode = False
        self.object_code = ""
        self.protocol("WM_DELETE_WINDOW", self._on_close)

        self.display_cols: list[int] = []  # indices of self.tests in left-to-right order

        self.row_h = 22
        self.hdr_h = 64
        self.col_gap = 12
        self.w_depth = 64
        self.w_val = 56
        self.pad_x = 8
        self.pad_y = 8

        self._editing = None  # (test_idx,row,field, entry)
        self._ctx_menu = None
        self._ctx_target = None  # (ti,row) for delete
        self._rc_preview = None  # (ti,row) transient red row highlight for context menu

        self._build_ui()
        # realtime footer status
        self._footer_force_live = True
        try:
            self.after(400, self._footer_live_tick)
        except Exception:
            pass

    # ---------------- history ----------------

    def _compute_depth_grid(self):
        """Build a common depth grid to align all tests visually.

        Returns: (grid_floats, step_m, row_maps, start_rows)
          - row_maps[ti]: dict(grid_row_index -> data_index_in_test)
          - start_rows[ti]: first grid row where the test has data (depth[0])
        """
        if not getattr(self, "tests", None):
            return [], None, {}, {}

        depth_lists = {}
        min_d = None
        max_d = None
        diffs = []
        for ti, t in enumerate(self.tests):
            dvals = []
            for ds in (getattr(t, "depth", []) or []):
                dv = _parse_depth_float(ds)
                if dv is None:
                    continue
                dvals.append(float(dv))
            if not dvals:
                continue
            depth_lists[ti] = dvals
            min_d = dvals[0] if (min_d is None or dvals[0] < min_d) else min_d
            max_d = dvals[-1] if (max_d is None or dvals[-1] > max_d) else max_d
            for a, b in zip(dvals, dvals[1:]):
                dd = b - a
                if dd > 1e-6:
                    diffs.append(dd)

        if min_d is None or max_d is None:
            return [], None, {}, {}

        step = None
        if diffs:
            step = min(diffs)
            for cand in (0.05, 0.1):
                if abs(step - cand) < 0.005:
                    step = cand
                    break
            step = round(step, 4)
        if step is None or step <= 0:
            step = float(getattr(self, "step_m", 0.05) or 0.05)

        grid = []
        x = round(float(min_d), 2)
        lim = float(max_d) + 1e-6
        i = 0
        while x <= lim and i < 20000:
            grid.append(round(x, 2))
            x = round(x + step, 4)
            i += 1

        row_maps = {}
        start_rows = {}
        grid_index = {round(v, 2): idx for idx, v in enumerate(grid)}
        for ti, dvals in depth_lists.items():
            mp = {}
            for di, dv in enumerate(dvals):
                key = round(dv, 2)
                gi = grid_index.get(key)
                if gi is not None:
                    mp[gi] = di
            row_maps[ti] = mp
            if mp:
                start_rows[ti] = min(mp.keys())

        return grid, step, row_maps, start_rows

    def _snapshot(self) -> dict:
        """Снимок состояния для Undo/Redo: данные + раскраска."""
        tests_snap: list[dict] = []
        for t in self.tests:
            tests_snap.append({
                "tid": t.tid,
                "dt": t.dt,
                "depth": list(t.depth),
                "qc": list(t.qc),
                "fs": list(t.fs),
                "marker": t.marker,
                "header_pos": t.header_pos,
                "orig_id": t.orig_id,
                "export_on": bool(getattr(t, "export_on", True)),
                "block": None if t.block is None else {
                    "order_index": t.block.order_index,
                    "header_start": t.block.header_start,
                    "header_end": t.block.header_end,
                    "id_pos": t.block.id_pos,
                    "dt_pos": t.block.dt_pos,
                    "data_start": t.block.data_start,
                    "data_end": t.block.data_end,
                    "marker_byte": t.block.marker_byte,
                },
            })
        flags_snap: dict[int, dict] = {}
        try:
            for tid, fl in (self.flags or {}).items():
                flags_snap[int(tid)] = {
                    "invalid": bool(getattr(fl, "invalid", False)),
                    "interp": sorted(list(getattr(fl, "interp_cells", set()))),
                    "force": sorted(list(getattr(fl, "force_cells", set()))),
                    "user": sorted(list(getattr(fl, "user_cells", set()))),
                    "algo": sorted(list(getattr(fl, "algo_cells", set()))),
                    "tail": sorted(list(getattr(fl, "force_tail_rows", set()))),
                }
        except Exception:
            flags_snap = {}
        return {"tests": tests_snap, "flags": flags_snap, "step_m": float(getattr(self, "step_m", 0.05) or 0.05)}

    def _restore(self, snap: dict):
        self.tests = []
        self._credo_force_export = False  # after user acknowledged issues/fix, export proceeds without re-check
        self.flags = {}
        # восстановить выбранный шаг сетки (м), иначе после Undo возможны «пропуски»
        try:
            self.step_m = float(snap.get("step_m", getattr(self, "step_m", 0.05) or 0.05) or 0.05)
        except Exception:
            pass

        for d in snap.get("tests", []):
            blk = None
            if d.get("block") is not None:
                b = d["block"]
                blk = GeoBlockInfo(
                    order_index=b["order_index"],
                    header_start=b["header_start"],
                    header_end=b["header_end"],
                    id_pos=b["id_pos"],
                    dt_pos=b["dt_pos"],
                    data_start=b["data_start"],
                    data_end=b["data_end"],
                    marker_byte=b["marker_byte"],
                )
            t = TestData(
                tid=int(d["tid"]),
                dt=d["dt"],
                depth=list(d["depth"]),
                qc=list(d["qc"]),
                fs=list(d["fs"]),
                marker=d.get("marker",""),
                header_pos=d.get("header_pos",""),
                orig_id=d.get("orig_id", None),
                block=blk,
            )
            self.tests.append(t)
            try:
                t.export_on = bool(d.get('export_on', True))
            except Exception:
                pass
            # создать флаги для восстановленного зондирования
            try:
                self.flags[t.tid] = TestFlags(False, set(), set(), set(), set(), set())
            except Exception:
                pass
        fsnap = snap.get("flags", {}) or {}
        for t in self.tests:
            s = fsnap.get(int(t.tid))
            if not s:
                continue
            fl = self.flags.get(t.tid)
            if not fl:
                continue
            try:
                fl.invalid = bool(s.get("invalid", False))
                fl.interp_cells = set(tuple(x) for x in s.get("interp", []))
                fl.force_cells = set(tuple(x) for x in s.get("force", []))
                fl.user_cells = set(tuple(x) for x in s.get("user", []))
                fl.algo_cells = set(tuple(x) for x in s.get("algo", []))
                fl.force_tail_rows = set(int(x) for x in s.get("tail", []))
            except Exception:
                pass

        self._end_edit(commit=False)
        self._redraw()

        # После успешной корректировки — синяя строка в подвале
        try:
            self._footer_force_live = False
            self.footer_cmd.config(foreground="#0b5ed7")
            self.footer_cmd.config(text="Статическое зондирование откорректировано.")
        except Exception:
            pass

    def _push_undo(self):
        if not self.tests:
            return
        self.undo_stack.append(self._snapshot())
        # limit
        if len(self.undo_stack) > 50:
            self.undo_stack.pop(0)
        self.redo_stack.clear()
        self._dirty = True




    def undo(self):
        # Если сейчас редактируется ячейка — завершаем редактирование и только потом делаем UNDO
        if getattr(self, '_editing', None):
            try:
                if len(self._editing) >= 4:
                    ti, row, field, e = self._editing[0], self._editing[1], self._editing[2], self._editing[3]
                else:
                    ti, row, field, e = self._editing
                if field == 'depth':
                    self._end_edit_depth0(ti, e, commit=True)
                else:
                    self._end_edit(commit=True)
            except Exception:
                pass
        if not self.undo_stack:
            return
        self.redo_stack.append(self._snapshot())
        snap = self.undo_stack.pop()
        self._restore(snap)
        # не затираем статус текстом Undo
        self._footer_force_live = True
        try:
            self._redraw()
        except Exception:
            pass
        # После Undo — вернуть красную строку (или серую, если проблем нет)
        try:
            self._set_footer_from_scan()
        except Exception:
            pass


    def redo(self):
        # Если сейчас редактируется ячейка — завершаем редактирование и только потом делаем REDO
        if getattr(self, '_editing', None):
            try:
                if len(self._editing) >= 4:
                    ti, row, field, e = self._editing[0], self._editing[1], self._editing[2], self._editing[3]
                else:
                    ti, row, field, e = self._editing
                if field == 'depth':
                    self._end_edit_depth0(ti, e, commit=True)
                else:
                    self._end_edit(commit=True)
            except Exception:
                pass
        if not self.redo_stack:
            return
        self.undo_stack.append(self._snapshot())
        snap = self.redo_stack.pop()
        self._restore(snap)
        # не затираем статус текстом Redo
        self._footer_force_live = True
        try:
            self._redraw()
        except Exception:
            pass
        # После Redo — показать актуальную строку по текущему состоянию
        try:
            self._set_footer_from_scan()
        except Exception:
            pass

    def _build_ui(self):
        # ========= RIBBON (Word-like) =========
        ribbon = ttk.Frame(self, padding=(10,8))
        ribbon.pack(side="top", fill="x")

        # Win11-ish ttk styles
        s = ttk.Style()
        try:
            s.configure("Ribbon.TButton", padding=(14,10))
            s.configure("RibbonAccent.TButton", padding=(14,10))
        except Exception:
            pass

        left = ttk.Frame(ribbon)
        left.pack(side="left", padx=10, pady=8)
        mid = ttk.Frame(ribbon)
        mid.pack(side="left", fill="x", expand=True, padx=(10, 0), pady=8)

        # Текущий файл (перенесено в левую часть ленты)
        self.file_var = tk.StringVar(value="(не выбран)")
        f_lbl = ttk.Label(mid, textvariable=self.file_var)
        f_lbl.pack(side="left", padx=(0, 10))
        ToolTip(f_lbl, "Текущий файл")
        def make_btn(parent, text, tip, cmd, style="Ribbon.TButton", w=4):
            b = ttk.Button(parent, text=text, command=cmd, style=style, width=w)
            ToolTip(b, tip)
            return b

        # File + history
        make_btn(left, "📂", "Загрузить GEO/GE0 или GXL", self.pick_file_and_load).grid(row=0, column=0, padx=4)
        make_btn(left, "📦", "Экспорт-архив (GEO+GXL+XLSX+CSV CREDO)", self.export_bundle, ).grid(row=0, column=1, padx=4)
        make_btn(left, "↶", "Назад (Undo)", self.undo, ).grid(row=0, column=2, padx=(16, 4))
        make_btn(left, "↷", "Вперёд (Redo)", self.redo, ).grid(row=0, column=3, padx=4)

        # Project map (FILE MAP)
        right = ttk.Frame(ribbon)
        right.pack(side="right", padx=10, pady=8)
        make_btn(right, "Карта", "Показать карту проекта (FILE MAP)", self.show_file_map, w=6).pack(side="right", padx=4)


        # Inputs (скрыты: параметры спрашиваем только при необходимости при открытии GEO)
        self.depth_var = tk.StringVar(value="")
        self.step_choice = tk.StringVar(value="")

        self._depth_label = ttk.Label(left, text="Глубина начала, м:")
        self._depth_label.grid(row=0, column=4, padx=(20, 4))
        self._depth_entry = ttk.Entry(left, textvariable=self.depth_var, width=12)
        self._depth_entry.grid(row=0, column=5, padx=4)
        ToolTip(self._depth_entry, "Изменить начальную глубину зондирования")
        ToolTip(self._depth_label, "Изменить начальную глубину зондирования")

        self._step_label = ttk.Label(left, text="Шаг зондирования, см:")
        self._step_label.grid(row=0, column=6, padx=(12, 4))
        self._step_box = ttk.Frame(left)
        self._step_box.grid(row=0, column=7, padx=4)
        self.rb5 = ttk.Radiobutton(self._step_box, text="5", variable=self.step_choice, value="5")
        self.rb10 = ttk.Radiobutton(self._step_box, text="10", variable=self.step_choice, value="10")
        self.rb5.pack(side="left")
        self.rb10.pack(side="left", padx=(8, 0))
        try:
            self.rb5.deselect(); self.rb10.deselect(); self.step_choice.set("")
        except Exception:
            pass
        ToolTip(self._step_box, "Шаг по глубине (см): 5 или 10")

        # Скрываем эти элементы — теперь они появляются только в окне ввода, если данных нет в GEO.
        try:
            self._depth_label.grid_remove()
            self._depth_entry.grid_remove()
            self._step_label.grid_remove()
            self._step_box.grid_remove()
        except Exception:
            pass


        # Action buttons (moved from bottom)
        btns = ttk.Frame(ribbon)
        btns.pack(side="left", padx=16, pady=8)
        self.btn_show = make_btn(btns, "👁", "Показать зондирования (читать GEO)", self.load_and_render, style="RibbonAccent.TButton")
        self.btn_show.grid(row=0, column=0, padx=4)
        try:
            self.btn_show.grid_remove()
        except Exception:
            pass
        make_btn(btns, "🛠", "Корректировка", self.fix_by_algorithm, ).grid(row=0, column=1, padx=4)
        make_btn(btns, "10→5", "Конвертировать шаг 10 см → 5 см", self.convert_10_to_5, w=5).grid(row=0, column=2, padx=4)  
        
        make_btn(btns, "⚙", "Параметры зондирований (GEO)", self.open_geo_params_dialog, w=3).grid(row=0, column=3, padx=4)
        make_btn(btns, "➕", "Добавить зондирование", self.add_test).grid(row=0, column=5, padx=4)

        # Right: calc params
        right = ttk.Frame(ribbon)
        right.pack(side="right", padx=10, pady=6)

        params = ttk.LabelFrame(right, text="Параметры пересчёта", padding=(10,6))
        params.pack(side="right")

        self.scale_var = tk.StringVar(value="250")
        self.fcone_var = tk.StringVar(value="30")
        self.fsleeve_var = tk.StringVar(value="10")
        self.acon_var = tk.StringVar(value="10")
        self.asl_var = tk.StringVar(value="350")

        def p_row(r, c, label, var, tip):
            ttk.Label(params, text=label).grid(row=r, column=c, sticky="w", padx=(8, 4), pady=2)
            ent = ttk.Entry(params, textvariable=var, width=10, validate="key", validatecommand=(self.register(_validate_nonneg_float_key), "%P"))
            ent.grid(row=r, column=c+1, sticky="w", padx=(0, 10), pady=2)
            ToolTip(ent, tip)
            def _sanitize(_ev=None, _v=var):
                s=_v.get().strip()
                if s=='':
                    return
                try:
                    v=float(s.replace(',', '.'))
                except Exception:
                    _v.set('0')
                    return
                if v<0:
                    v=0.0
                if abs(v-round(v))<1e-9:
                    _v.set(str(int(round(v))))
                else:
                    _v.set(str(v).replace('.', ','))
            ent.bind('<FocusOut>', _sanitize)

        p_row(0, 0, "Шкала:", self.scale_var, "Максимум делений прибора (обычно 250)")
        p_row(0, 2, "Aкон (см²):", self.acon_var, "Площадь лба конуса, см²")
        p_row(1, 0, "Fкон (кН):", self.fcone_var, "Макс. усилие по лбу конуса, кН")
        p_row(1, 2, "Aмуф (см²):", self.asl_var, "Площадь боковой поверхности муфты, см²")
        p_row(2, 0, "Fмуф (кН):", self.fsleeve_var, "Макс. усилие по муфте, кН")
        # ========= Main canvas (fixed header) =========
        mid = ttk.Frame(self)
        mid.pack(side="top", fill="both", expand=True)

        self.mid = mid  # host for table + hscroll (between table and footer)

        # Верхняя фиксированная шапка
        self.hcanvas = tk.Canvas(mid, background="white", highlightthickness=0, height=120)
        self.hcanvas.pack(side="top", fill="x")

        # Нижняя область с данными (скролл)
        body = ttk.Frame(mid)
        body.pack(side="top", fill="both", expand=True)

        self.vbar = ttk.Scrollbar(body, orient="vertical")
        self.vbar.pack(side="right", fill="y")

        self.canvas = tk.Canvas(
            body, background="white", highlightthickness=0,
            yscrollcommand=self.vbar.set
        )
        self.canvas.pack(side="left", fill="both", expand=True)

        def _xview_proxy(*args):
            # ЕДИНЫЙ ИСТОЧНИК X — только body canvas.
            # Шапку синхронизируем ПОСЛЕ того как Tk применит прокрутку (after_idle),
            # иначе на правом краю из-за округлений бывает дрейф.
            self._end_edit(commit=True)
            try:
                self.canvas.xview(*args)
            except Exception:
                return
            # sync header (и зажим правого края при перетаскивании ползунка/скролле):
            # как только последняя колонка ВИДНА ПОЛНОСТЬЮ — вправо больше не двигаем.
            def _sync():
                try:
                    w = float(getattr(self, "_scroll_w", 0) or 0)
                except Exception:
                    w = 0.0
                if w <= 1:
                    try:
                        w = float(self._content_size()[0])
                        self._scroll_w = w
                    except Exception:
                        w = 1.0
                try:
                    view_w = float(self.canvas.winfo_width())
                except Exception:
                    view_w = 0.0

                try:
                    frac = float(self.canvas.xview()[0])
                except Exception:
                    frac = 0.0

                # вычислим максимально допустимую позицию X, при которой последняя колонка видна полностью
                try:
                    n_tests = len(self.tests)
                except Exception:
                    n_tests = 0
                try:
                    col_w = float(self.w_depth + self.w_val*2 + (self.w_val if getattr(self, 'geo_kind', 'K2')=='K4' else 0))
                except Exception:
                    col_w = 0.0
                try:
                    gap = float(self.col_gap)
                except Exception:
                    gap = 0.0
                try:
                    pad = float(self.pad_x)
                except Exception:
                    pad = 0.0

                last_left_px = pad + (col_w + gap) * max(0, n_tests - 1)
                last_right_px = last_left_px + col_w
                max_px = max(0.0, w - max(1.0, view_w))
                allow_px = min(max_px, max(0.0, last_right_px - max(1.0, view_w)))

                cur_px = frac * w
                if cur_px > (allow_px + 0.5):
                    # «зажать» вправо
                    frac2 = 0.0 if w <= 1 else (allow_px / w)
                    try:
                        self.canvas.xview_moveto(frac2)
                    except Exception:
                        pass
                    try:
                        frac = float(self.canvas.xview()[0])
                    except Exception:
                        frac = frac2

                try:
                    self.hcanvas.xview_moveto(frac)
                except Exception:
                    pass
            try:
                self.after_idle(_sync)
            except Exception:
                _sync()

        def _on_xscroll_command(first, last):
            # first/last: доли [0..1] видимой области
            # Обновляем сам горизонтальный скролл (если уже создан)
            try:
                if hasattr(self, "hscroll"):
                    self.hscroll.set(first, last)
            except Exception:
                pass

            # Синхронизируем X для canvas (тело) и hcanvas (шапка) без дрожания/уезда.
            # xscrollcommand вызывается и для canvas, и для hcanvas — используем lock.
            if getattr(self, "_xsync_lock", False):
                return
            self._xsync_lock = True
            try:
                f = float(first)
                try:
                    c0 = float(self.canvas.xview()[0])
                except Exception:
                    c0 = f
                try:
                    h0 = float(self.hcanvas.xview()[0])
                except Exception:
                    h0 = f

                if abs(c0 - f) > 1e-4:
                    try:
                        self.canvas.xview_moveto(f)
                    except Exception:
                        pass
                if abs(h0 - f) > 1e-4:
                    try:
                        self.hcanvas.xview_moveto(f)
                    except Exception:
                        pass
            finally:
                self._xsync_lock = False

        # назначаем xscrollcommand сразу, а сам hscroll свяжем позже, когда создадим в footer
        self.canvas.configure(xscrollcommand=_on_xscroll_command)
        # Важно: xscrollcommand назначаем ТОЛЬКО на основной canvas.
        # Иначе на самом правом краю могут появляться расхождения фракций и «уезд» шапки.
        # Шапку двигаем синхронно через _xview_proxy / _on_xscroll_command.
        self._xview_proxy = _xview_proxy
        self._on_xscroll_command = _on_xscroll_command

        def _yview_proxy(*args):
            self._end_edit(commit=True)
            try:
                self.canvas.yview(*args)
            except Exception:
                pass
        self.vbar.config(command=_yview_proxy)
        # configure/redraw
        self.canvas.bind("<Configure>", lambda _e: self._update_scrollregion())
        self.hcanvas.bind("<Configure>", lambda _e: (self.hcanvas.configure(width=self.canvas.winfo_width()), self._update_scrollregion()))

        # scrolling and events: таблица
        self.canvas.bind("<MouseWheel>", self._on_mousewheel)
        self.canvas.bind("<Button-4>", lambda e: self._on_mousewheel_linux(-1))
        self.canvas.bind("<Button-5>", lambda e: self._on_mousewheel_linux(1))
        self.canvas.bind("<Double-1>", lambda _e: "break")  # disable dblclick edit
        self.canvas.bind("<Button-3>", self._on_right_click)
        self.canvas.bind("<Control-Button-1>", self._on_right_click)
        self.canvas.bind("<Button-1>", self._on_left_click)
        # глобальный клик вне canvases закрывает редактирование
        self.bind_all("<Button-1>", self._on_global_click, add="+")
        # Навигация стрелками по активным ячейкам (qc/fs)
        for _k in ("<Up>", "<Down>", "<Left>", "<Right>"):
            self.bind(_k, self._on_arrow_key)
        self.canvas.bind("<Motion>", self._on_motion)
        self.canvas.bind("<Leave>", lambda _e: self._set_hover(None))

        # события шапки (клики по иконкам/галочке)
        self.hcanvas.bind("<Button-1>", self._on_left_click)
        # супер-фишка: колесо мыши по шапке листает горизонтально
        self.hcanvas.bind("<MouseWheel>", self._on_mousewheel_x)
        self.hcanvas.bind("<Button-4>", lambda e: self._on_mousewheel_linux_x(-1))
        self.hcanvas.bind("<Button-5>", lambda e: self._on_mousewheel_linux_x(1))
        self.hcanvas.bind("<Motion>", self._on_motion)
        self.hcanvas.bind("<Leave>", lambda _e: self._set_hover(None))
        self._ctx_menu = tk.Menu(self, tearoff=0)
        self._ctx_menu.add_command(label="Удалить выше (вкл.)", command=self._ctx_delete_above)
        self._ctx_menu.add_command(label="Удалить ниже (вкл.)", command=self._ctx_delete_below)

        # ========= Footer: команды + легенда цветов =========
        self.footer = ttk.Frame(self, padding=(12, 4))
        self.footer.pack(side="bottom", fill="x")

        # ===== Горизонтальная прокрутка (классический Win/Win11 Scrollbar) =====
        # Требование (по скрину): полоска должна быть ВЫШЕ строки статуса (например: «Добавлена новая зондирование...»),
        # но НИЖЕ таблицы и ВЫШЕ нижнего подвала (легенда/индикаторы).
        #
        # Схема по pack (снизу вверх):
        #   footer (легенда)          — самый низ
        #   status (текстовый статус) — над footer
        #   hscroll_frame             — над status, у нижней кромки таблицы
        #
        # Синхронизация X:
        #   - Scrollbar двигает одновременно canvas (тело) и hcanvas (шапку) через единый прокси self._xview_proxy
        #   - xscrollcommand навешан на оба canvas
        self.hscroll_frame = ttk.Frame(self.mid, padding=(12, 0, 12, 0))
        self.hscroll = ttk.Scrollbar(self.hscroll_frame, orient="horizontal")
        self.hscroll.pack(fill="x")
        try:
            self.hscroll.configure(command=self._xview_proxy)
        except Exception:
            pass
        # супер-фишка: колесо мыши по горизонтальному скроллу листает горизонтально
        try:
            self.hscroll.bind("<MouseWheel>", self._on_mousewheel_x)
            self.hscroll.bind("<Button-4>", lambda e: self._on_mousewheel_linux_x(-1))
            self.hscroll.bind("<Button-5>", lambda e: self._on_mousewheel_linux_x(1))
        except Exception:
            pass
        # по умолчанию скрыт — показываем только когда ширина контента больше видимой области
        self._hscroll_hidden = True

        self.footer_cmd = ttk.Label(
            self.footer,
            text="",
            foreground="#666666",
        )
        self.footer_cmd.pack(side="left")

        leg = ttk.Frame(self.footer)
        leg.pack(side="right")

        def _leg_item(parent, color: str, text: str):
            box = tk.Label(parent, width=2, height=1, bg=color, relief="solid", bd=1)
            box.pack(side="left", padx=(8, 4), pady=2)
            ttk.Label(parent, text=text).pack(side="left")

        # ЛЕГЕНДА (строго по промту)
        _leg_item(leg, GUI_PURPLE, "исправлено")
        _leg_item(leg, GUI_YELLOW, "отсутствуют значения")
        _leg_item(leg, GUI_BLUE, "отсутствует отказ")
        _leg_item(leg, GUI_GREEN, "откорректировано")
        _leg_item(leg, GUI_RED, "некорректный опыт")

        self.status = ttk.Label(self, text="Готов.", padding=(12, 6))

        # статусная строка — над подвалом
        self.status.pack(side="bottom", fill="x", before=self.footer)

        # горизонтальная полоса прокрутки — СРАЗУ после таблицы (над статусом)
        # по умолчанию скрыта; при показе перепаковываем статус НИЖЕ полосы

        # hscroll живёт ВНУТРИ mid (между таблицей и нижними статус/подвал)
        self.hscroll_frame.pack(side="bottom", fill="x")
        self.hscroll_frame.pack_forget()
    def pick_file_and_load(self):
        path = filedialog.askopenfilename(
            title="Выберите файл GEO/GE0 или GXL",
            filetypes=[
                ("GeoExplorer GEO / GXL", "*.geo *.ge0 *.GEO *.GE0 *.gxl *.GXL"),
                ("Все файлы", "*.*"),
            ],
        )
        if not path:
            return
        self.geo_path = Path(path)
        self.file_var.set(str(self.geo_path))
        self.is_gxl = (self.geo_path.suffix.lower() == ".gxl")
        # путь текущего загруженного файла
        self.loaded_path = str(self.geo_path)
        # Современный поток: сразу загружаем и показываем.
        self.load_and_render()
        try:
            _log_event(self.usage_logger, "OPEN", file=str(self.geo_path))
        except Exception:
            pass
        self._ensure_object_code()

    def open_geo_params_dialog(self):
        """Открыть окно параметров GEO для текущего файла."""
        if not getattr(self, "tests", None):
            return
        # Снимок для Undo делаем только если пользователь нажмёт OK
        snap = None
        try:
            snap = self._snapshot()
        except Exception:
            snap = None

        ok = False
        try:
            ok = self._prompt_geo_build_params(self.tests, need_depth=True, need_step=True)
        except Exception as e:
            try:
                self._log(f"[geo_params] error: {e}")
            except Exception:
                pass
            ok = False

        if ok:
            try:
                if snap is not None:
                    self.undo_stack.append(snap)
                    if len(self.undo_stack) > 50:
                        self.undo_stack.pop(0)
                    self.redo_stack.clear()
                    self._dirty = True
            except Exception:
                pass
            try:
                self._redraw()
            except Exception:
                pass


    def _parse_depth_step(self):
        if not self.geo_path or not self.geo_path.exists():
            messagebox.showerror("Ошибка", "Сначала выберите файл .GEO/.GE0")
            return None

        dtxt = self.depth_var.get().strip()
        if not dtxt:
            messagebox.showwarning("Внимание", "Укажи начальную глубину (м).")
            return None
        try:
            depth_start = float(dtxt.replace(",", "."))
        except Exception:
            messagebox.showerror("Ошибка", "Некорректная начальная глубина. Пример: 1.5")
            return None

        stxt = self.step_choice.get().strip()
        if not stxt:
            messagebox.showwarning("Внимание", "Выбери шаг (5 или 10 см).")
            return None
        step_cm = int(stxt)
        if step_cm not in (5, 10):
            messagebox.showerror("Ошибка", "Шаг должен быть 5 или 10 см.")
            return None

        step_m = step_cm / 100.0
        return depth_start, step_m


    def _update_status_loaded(self, prefix: str):
        """Обновляет строку статуса подвала (1-я строка): количество опытов + параметры.
        Формат: 'Загружено опытов N шт. параметры: шкала делений 250, Fкон 30кН, Fмуф 10кН, шаг 10см'
        """
        try:
            scale = self.scale_var.get().strip() if hasattr(self, "scale_var") else ""
            fcone = self.fcone_var.get().strip() if hasattr(self, "fcone_var") else ""
            fsleeve = self.fsleeve_var.get().strip() if hasattr(self, "fsleeve_var") else ""
            step = self.step_cm_var.get().strip() if hasattr(self, "step_cm_var") else ""

            parts = []
            if scale:
                parts.append(f"шкала делений {scale}")
            if fcone:
                parts.append(f"Fкон {fcone}кН")
            if fsleeve:
                parts.append(f"Fмуф {fsleeve}кН")
            if step:
                parts.append(f"шаг {step}см")

            tail = (" параметры: " + ", ".join(parts)) if parts else ""
            # status (1-я строка) всегда чёрная — не красим её предупреждениями
            self.status.config(text=prefix + tail)
        except Exception:
            self.status.config(text=prefix)


    def _set_status_loaded(self, prefix: str):
        # alias for legacy calls
        return self._update_status_loaded(prefix)

    def _normalize_test_lengths(self, t):
        """
        Нормализует длины массивов внутри ОДНОГО опыта:
        - строки с пустой depth считаются удалёнными и выкидываются целиком
        - qc/fs при пустом значении -> "0"
        Гарантия: len(depth)==len(qc)==len(fs) для каждого опыта отдельно.
        """
        depth = list(getattr(t, "depth", []) or [])
        qc    = list(getattr(t, "qc", []) or [])
        fs    = list(getattr(t, "fs", []) or [])

        n = max(len(depth), len(qc), len(fs))
        new_d, new_qc, new_fs = [], [], []

        for i in range(n):
            d = depth[i] if i < len(depth) else ""
            d = (d or "").strip()
            if not d:
                continue  # удалённая/пустая строка

            q = qc[i] if i < len(qc) else "0"
            f = fs[i] if i < len(fs) else "0"

            q = "0" if (q is None or str(q).strip() == "") else str(q)
            f = "0" if (f is None or str(f).strip() == "") else str(f)

            new_d.append(str(d))
            new_qc.append(q)
            new_fs.append(f)

        t.depth = new_d
        t.qc = new_qc
        t.fs = new_fs
        return t


    def _apply_gxl_calibration_from_meta(self, meta_rows: list[dict]):
        """Если в meta_rows есть шкала/тарировки — подставить в поля пересчёта."""
        if not meta_rows:
            return
        kv = {}
        for row in meta_rows:
            try:
                k = str(row.get("key", "")).strip().lower()
                v = str(row.get("value", "")).strip()
                if k:
                    kv[k] = v
            except Exception:
                pass

        if kv.get("scale"):
            self.scale_var.set(kv["scale"])
        if kv.get("scaleostria"):
            self.fcone_var.set(kv["scaleostria"])
        if kv.get("scalemufta"):
            self.fsleeve_var.set(kv["scalemufta"])

    def _calc_qc_fs_from_del(self, qc_del: int, fs_del: int) -> tuple[float, float]:
        """Пересчёт делений в qc (МПа) и fs (кПа) как в GeoExplorer.
        Использует: шкала (дел.), Fконуса (кН), Fмуфты (кН).
        Приняты площади: конус 10 см², муфта 350 см² (типовая для GeoExplorer).
        """
        def _f(x: str, default: float) -> float:
            try:
                s = (x or '').strip().replace(',', '.')
                return float(s) if s else default
            except Exception:
                return default

        scale_div = int(round(_f(self.scale_var.get() if getattr(self, 'scale_var', None) else '250', 250.0)))
        if scale_div <= 0:
            scale_div = 250
        fcone_kn = _f(self.fcone_var.get() if getattr(self, 'fcone_var', None) else '30', 30.0)
        fsleeve_kn = _f(self.fsleeve_var.get() if getattr(self, 'fsleeve_var', None) else '10', 10.0)

        CONE_AREA_CM2 = 10.0
        SLEEVE_AREA_CM2 = 350.0

        # qc: (del/scale)*F(kN) / A(cm2) * 10 -> MPa  (1 kN/cm2 = 10 MPa)
        qc_mpa = (qc_del / scale_div) * fcone_kn * (10.0 / CONE_AREA_CM2)
        # fs: (del/scale)*F(kN) / A(cm2) * 10000 -> kPa (1 kN/cm2 = 10000 kPa)
        fs_kpa = (fs_del / scale_div) * fsleeve_kn * (10000.0 / SLEEVE_AREA_CM2)
        return qc_mpa, fs_kpa

    def _prompt_missing_geo_params(self, *, need_depth: bool, need_step: bool) -> bool:
        """Спросить у пользователя недостающие параметры для GEO: h0 (0..4 м) и/или шаг (5/10 см)."""
        dlg = tk.Toplevel(self)
        dlg.title("Параметры зондирования")
        dlg.transient(self)
        dlg.grab_set()
        dlg.resizable(False, False)

        frm = ttk.Frame(dlg, padding=12)
        frm.pack(fill="both", expand=True)

        try:
            frm.columnconfigure(1, weight=1)
            frm.columnconfigure(2, weight=1)
            frm.columnconfigure(3, weight=1)
            frm.columnconfigure(4, weight=1)
        except Exception:
            pass

        row = 0
        info = ttk.Label(frm, text="В файле GEO отсутствуют (или нулевые) параметры глубины/шага.\nЗаполни недостающие значения.", justify="left")
        info.grid(row=row, column=0, columnspan=3, sticky="w")
        row += 1

        depth_var = tk.StringVar(value="")
        step_var = tk.StringVar(value="10")

        if need_depth:
            ttk.Label(frm, text="Начальная глубина h0, м (0..4):").grid(row=row, column=0, sticky="w", pady=(10, 4))
            ent = ttk.Entry(frm, textvariable=depth_var, width=10)
            ent.grid(row=row, column=1, sticky="w", pady=(10, 4))
            row += 1
        else:
            ttk.Label(frm, text=f"Начальная глубина h0, м: {float(getattr(self, 'depth_start', 0.0) or 0.0):g}").grid(row=row, column=0, columnspan=3, sticky="w", pady=(10, 4))
            row += 1

        if need_step:
            ttk.Label(frm, text="Шаг зондирования, см:").grid(row=row, column=0, sticky="w", pady=(6, 4))
            rbfrm = ttk.Frame(frm)
            rbfrm.grid(row=row, column=1, sticky="w", pady=(6, 4))
            ttk.Radiobutton(rbfrm, text="5", value="5", variable=step_var).pack(side="left")
            ttk.Radiobutton(rbfrm, text="10", value="10", variable=step_var).pack(side="left", padx=(10, 0))
            row += 1
        else:
            step_cm = 10 if float(getattr(self, "step_m", 0.10) or 0.10) >= 0.075 else 5
            ttk.Label(frm, text=f"Шаг, см: задан ранее ({step_cm})").grid(row=row, column=0, columnspan=3, sticky="w", pady=(6, 4))
            row += 1

        msg_var = tk.StringVar(value="")
        ttk.Label(frm, textvariable=msg_var, foreground="#b00020").grid(row=row, column=0, columnspan=3, sticky="w", pady=(6, 0))
        row += 1

        btns = ttk.Frame(frm)
        btns.grid(row=row, column=0, columnspan=3, sticky="e", pady=(12, 0))

        result = {"ok": False}

        def on_ok():
            if need_depth:
                try:
                    h0 = float(depth_var.get().strip().replace(",", "."))
                except Exception:
                    msg_var.set("Некорректная глубина. Пример: 1.2")
                    return
                if not (0.0 <= h0 <= 4.0):
                    msg_var.set("h0 должна быть в диапазоне 0..4 м.")
                    return
                self.depth_start = h0
                self._depth_confirmed = True
                try:
                    self.depth_var.set(f"{h0:g}")
                except Exception:
                    pass

            if need_step:
                st = step_var.get().strip()
                if st not in ("5", "10"):
                    msg_var.set("Выберите шаг 5 или 10 см.")
                    return
                self.step_m = 0.05 if st == "5" else 0.10
                self._step_confirmed = True
                try:
                    self.step_choice.set(st)
                except Exception:
                    pass

            result["ok"] = True
            dlg.destroy()

        def on_cancel():
            dlg.destroy()

        ttk.Button(btns, text="Отмена", command=on_cancel).pack(side="right")
        ttk.Button(btns, text="OK", command=on_ok).pack(side="right", padx=(0, 10))

        try:
            # Enter = OK, Esc = Cancel; focus first field when possible
            if need_depth and 'ent' in locals():
                ent.focus_set()
                ent.bind('<Return>', lambda e: on_ok())
            dlg.bind('<Return>', lambda e: on_ok())
            dlg.bind('<Escape>', lambda e: on_cancel())
        except Exception:
            pass

        self.wait_window(dlg)
        return bool(result["ok"])


    def _prompt_geo_build_params(self, tests_list, *, need_depth: bool, need_step: bool) -> bool:
        """Окно после открытия GEO:
        - по центру рабочей области
        - шаг 10/5 см (по умолчанию 10)
        - поле 'Объект' сверху
        - общая начальная глубина + 'Применить ко всем'
        - список опытов: h0 + дата/время + кнопка календаря
        - Enter перескакивает по ячейкам h0
        - клик по неактивной ячейке снимает 'Применить ко всем' и активирует все поля
        """
        dlg = tk.Toplevel(self)
        dlg.title("Параметры GEO")
        dlg.transient(self)
        dlg.grab_set()
        dlg.resizable(False, False)

        frm = ttk.Frame(dlg, padding=12)
        frm.pack(fill="both", expand=True)

        ntests = len(tests_list or [])
        ttk.Label(frm, text=f"В файле GEO {ntests} опытов статического зондирования.", font=("Segoe UI", 10, "bold"))\
            .grid(row=0, column=0, columnspan=5, sticky="w", pady=(0, 8))

        # --- переменные ---
        # шаг (по умолчанию 10 см)
        _sm = float(getattr(self, "step_m", 0.10) or 0.10)
        _default_step = "5" if abs(_sm - 0.05) < 1e-6 else "10"
        step_var = tk.StringVar(value=_default_step)
        # общая начальная глубина
        if getattr(self, "depth0_by_tid", None):
            try:
                common_depth0 = float(min(self.depth0_by_tid.values()))
            except Exception:
                common_depth0 = float(getattr(self, "depth_start", 0.0) or 0.0)
        else:
            common_depth0 = float(getattr(self, "depth_start", 0.0) or 0.0)
        common_var = tk.StringVar(value=f"{common_depth0:g}")
        apply_all_var = tk.BooleanVar(value=(False if getattr(self, 'geo_kind', 'K2')=='K4' else True))

        # объект (встроено)
        obj_var = tk.StringVar(value=(getattr(self, "object_code", "") or ""))

        # сообщение об ошибке
        msg_var = tk.StringVar(value="")
        # msg_lbl будет создан ниже, перед кнопками

        # --- объект ---
        r = 1
        ttk.Label(frm, text="Объект:").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=2)
        obj_ent = ttk.Entry(frm, textvariable=obj_var, width=52)
        obj_ent.grid(row=r, column=1, columnspan=4, sticky="we", pady=2)
        r += 1

        # --- шаг ---
        if need_step:
            ttk.Label(frm, text="Шаг, см:").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=2)
            rb_frame = ttk.Frame(frm)
            rb_frame.grid(row=r, column=1, sticky="w", pady=2)
            ttk.Radiobutton(rb_frame, text="10", value="10", variable=step_var).pack(side="left", padx=(0, 10))
            ttk.Radiobutton(rb_frame, text="5", value="5", variable=step_var).pack(side="left")
            r += 1

        # --- общая глубина + apply all ---
        ttk.Label(frm, text="Начальная глубина, м:").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=2)
        common_ent = ttk.Entry(
            frm,
            textvariable=common_var,
            width=10,
            validate="key",
            validatecommand=(dlg.register(_validate_depth_0_4_key), "%P"),
        )
        common_ent.grid(row=r, column=1, sticky="w", pady=2)

        apply_all_chk = ttk.Checkbutton(frm, text="Применить ко всем", variable=apply_all_var)
        apply_all_chk.grid(row=r, column=2, columnspan=3, sticky="w", padx=(12, 0), pady=2)
        r += 1

        ttk.Separator(frm).grid(row=r, column=0, columnspan=5, sticky="ew", pady=(8, 8))
        r += 1

        # --- таблица опытов ---
        table_wrap = ttk.Frame(frm)
        table_wrap.grid(row=r, column=0, columnspan=5, sticky="nsew")
        r += 1

        MAX_VISIBLE = 14
        use_scroll = ntests > MAX_VISIBLE

        if use_scroll:
            canvas = tk.Canvas(table_wrap, height=MAX_VISIBLE * 26, highlightthickness=0)
            ysb = ttk.Scrollbar(table_wrap, orient="vertical", command=canvas.yview)
            inner = ttk.Frame(canvas)
            inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
            canvas.create_window((0, 0), window=inner, anchor="nw")
            canvas.configure(yscrollcommand=ysb.set)
            canvas.pack(side="left", fill="both", expand=True)
            ysb.pack(side="right", fill="y")
            table = inner
        else:
            table = table_wrap

        ttk.Label(table, text="Опыт", font=("Segoe UI", 9, "bold")).grid(row=0, column=0, sticky="w", padx=(0, 10))
        ttk.Label(table, text="Нач. глубина, м", font=("Segoe UI", 9, "bold")).grid(row=0, column=1, sticky="w")
        ttk.Label(table, text="Дата/время", font=("Segoe UI", 9, "bold")).grid(row=0, column=2, sticky="w", padx=(12, 0))

        row_vars = []   # (t, tid, h0_var, ent, dt_var, dt_lbl)

        def _parse_depth_str(s: str):
            try:
                ss = (s or "").strip().replace(",", ".")
                if ss == "":
                    return None
                return float(ss)
            except Exception:
                return None

        def _current_step_m():
            st = step_var.get().strip()
            return 0.05 if st == "5" else 0.10

        def _norm_dt(val):
            # из файла может прийти datetime, date или строка
            if isinstance(val, _dt.datetime):
                return val
            if isinstance(val, _dt.date):
                return _dt.datetime(val.year, val.month, val.day, 0, 0)
            try:
                if val is None:
                    return None
                return _try_parse_dt(str(val))
            except Exception:
                return None

        def _fmt_dt(dt_obj: _dt.datetime | None) -> str:
            if not dt_obj:
                return "--.--.---- --:--"
            return dt_obj.strftime("%d.%m.%Y %H:%M")

        # дефолт для общих/индивидуальных
        for i, t in enumerate(tests_list or [], start=1):
            tid = int(getattr(t, "tid", 0) or 0)

            ttk.Label(table, text=f"СЗ-{tid}").grid(row=i, column=0, sticky="w", padx=(0, 10), pady=2)

            try:
                init_v = float(self.depth0_by_tid.get(tid, common_depth0))
            except Exception:
                init_v = common_depth0

            h0_var = tk.StringVar(value=f"{init_v:g}")
            ent = ttk.Entry(
                table,
                textvariable=h0_var,
                width=10,
                validate="key",
                validatecommand=(dlg.register(_validate_depth_0_4_key), "%P"),
            )
            ent.grid(row=i, column=1, sticky="w", pady=2)
            try:
                ent._geo_param_entry = True
            except Exception:
                pass

            # дата/время (парсим из файла)
            dt0 = _norm_dt(getattr(t, "dt", None))
            dt_var = tk.StringVar(value=_fmt_dt(dt0))

            dt_lbl = ttk.Label(table, textvariable=dt_var, foreground="#666666", cursor="hand2")
            dt_lbl.grid(row=i, column=2, sticky="w", padx=(12, 0), pady=2)

            row_vars.append((t, tid, h0_var, ent, dt_var, dt_lbl))

        def _open_dt_calendar(row_tuple):
            t, tid, h0_var, ent, dt_var, dt_lbl = row_tuple
            # подсветка строки на время редактирования
            old = dt_var.get()
            dt_var.set(f"[{old}]")

            cur_dt = _norm_dt(getattr(t, "dt", None))
            cur_date = (cur_dt.date() if cur_dt else _dt.date.today())
            if cur_date > _dt.date.today():
                cur_date = _dt.date.today()

            cd = CalendarDialog(dlg, initial=cur_date, title="Выбор даты")
            self._center_child(cd)
            dlg.wait_window(cd)

            # вернуть подсветку/обновить
            sel = cd.selected
            if sel:
                # сохраняем время из файла, меняем только дату
                if cur_dt:
                    new_dt = _dt.datetime(sel.year, sel.month, sel.day, cur_dt.hour, cur_dt.minute, cur_dt.second)
                else:
                    new_dt = _dt.datetime(sel.year, sel.month, sel.day, 0, 0, 0)
                t.dt = new_dt.strftime("%Y-%m-%d %H:%M:%S")
                dt_var.set(_fmt_dt(new_dt))
            else:
                dt_var.set(old)

        # клик по дате открывает календарь
        for row in row_vars:
            row[-1].bind("<Button-1>", lambda e, r=row: _open_dt_calendar(r))
        recompute_busy = False
        def _recompute_apply_state():
            nonlocal recompute_busy
            if recompute_busy:
                return
            recompute_busy = True
            try:
                msg_var.set("")
                is_k4 = (getattr(self, 'geo_kind', 'K2') == 'K4')

                # K4: если apply_all выключен и глубины разные — показываем '(разные)' в общей ячейке
                if is_k4 and (not apply_all_var.get()):
                    try:
                        uniq_h0 = sorted({float(self.depth0_by_tid.get(int(tid), 0.0)) for (_t, tid, *_rest) in row_vars})
                    except Exception:
                        uniq_h0 = []
                    if len(uniq_h0) > 1:
                        if common_var.get().strip() != "(разные)":
                            common_var.set("(разные)")
                    elif len(uniq_h0) == 1:
                        v0 = uniq_h0[0]
                        if common_var.get().strip() != f"{v0:g}":
                            common_var.set(f"{v0:g}")

                cd = _parse_depth_str(common_var.get())
                if cd is None:
                    cd = 0.0

                if apply_all_var.get():
                    # применяем ко всем и блокируем индивидуальные поля
                    for (_t, _tid, h0_var, ent, _dt_var, _dt_lbl) in row_vars:
                        h0_var.set(f"{cd:g}")
                        try:
                            ent.config(state="disabled")
                        except Exception:
                            pass
                else:
                    # индивидуальные значения (для K4 — из файла)
                    for (_t, tid, h0_var, ent, _dt_var, _dt_lbl) in row_vars:
                        if is_k4:
                            try:
                                v = self.depth0_by_tid.get(int(tid), None)
                                if v is not None:
                                    h0_var.set(f"{float(v):g}")
                            except Exception:
                                pass
                        try:
                            ent.config(state="normal")
                        except Exception:
                            pass

                # управление общей ячейкой
                try:
                    if is_k4 and (not apply_all_var.get()):
                        common_ent.config(state="disabled")
                    else:
                        common_ent.config(state="normal")
                except Exception:
                    pass
            finally:
                recompute_busy = False


        def _on_common_change(*_):
            nonlocal recompute_busy
            if recompute_busy:
                return
            if apply_all_var.get():
                _recompute_apply_state()

        def _on_apply_toggle(*_):
            _recompute_apply_state()

        common_var.trace_add("write", _on_common_change)
        apply_all_var.trace_add("write", _on_apply_toggle)

        _recompute_apply_state()

        # если пользователь ввёл в строке значение != общего — снимаем apply_all
        def _make_row_trace(h0_var):
            def _on_row_change(*_):
                cd = _parse_depth_str(common_var.get())
                dv = _parse_depth_str(h0_var.get())
                if cd is None:
                    cd = 0.0
                if dv is None:
                    return
                if abs(dv - cd) > 1e-9 and apply_all_var.get():
                    apply_all_var.set(False)
                _recompute_apply_state()
            h0_var.trace_add("write", _on_row_change)

        for (t, tid, h0_var, ent, dt_var, dt_lbl) in row_vars:
            _make_row_trace(h0_var)

        # клик по неактивной ячейке: активировать все и снять галочку
        def _on_entry_click(event, ent_ref):
            try:
                st = str(ent_ref.cget("state"))
            except Exception:
                st = "normal"
            if st == "disabled":
                apply_all_var.set(False)
                dlg.after(0, lambda: (ent_ref.focus_set(), ent_ref.selection_range(0, "end")))
                return "break"
            return None

        for (t, tid, h0_var, ent, dt_var, dt_lbl) in row_vars:
            ent.bind("<Button-1>", lambda e, ee=ent: _on_entry_click(e, ee), add="+")
            ent.bind("<FocusIn>", lambda e, ee=ent: ee.selection_range(0, "end"), add="+")
        # (кнопки календаря убраны: редактирование по клику на дате)

        # Enter = переход к следующей ячейке h0
        def _focus_next(current_index: int):
            nxt = current_index + 1
            if nxt >= len(row_vars):
                on_ok()
                return "break"
            try:
                row_vars[nxt][3].focus_set()
                row_vars[nxt][3].selection_range(0, "end")
            except Exception:
                pass
            return "break"

        for idx, (t, tid, h0_var, ent, dt_var, dt_lbl) in enumerate(row_vars):
            ent.bind("<Return>", lambda e, k=idx: _focus_next(k))

        # применить начальные состояния
        _recompute_apply_state()

        # --- сообщение об ошибке + кнопки ---
        msg_lbl = ttk.Label(frm, textvariable=msg_var, foreground="#b00020")
        msg_lbl.grid(row=r, column=0, columnspan=5, sticky="w", pady=(8, 0))
        r += 1

        btns = ttk.Frame(frm)
        btns.grid(row=r, column=0, columnspan=5, sticky="e", pady=(12, 0))

        result = {"ok": False}

        def on_ok():
            # шаг
            if need_step:
                st = step_var.get().strip()
                if st not in ("5", "10"):
                    msg_var.set("Выберите шаг 5 или 10 см.")
                    return

            # общая глубина
            cd = _parse_depth_str(common_var.get())
            if cd is None:
                msg_var.set("Некорректная начальная глубина. Пример: 1.2")
                return
            if not (0.0 <= cd <= 4.0):
                msg_var.set("Начальная глубина должна быть в диапазоне 0..4 м.")
                return

            # объект
            self.object_code = (obj_var.get() or "").strip()

            # сохранить общие
            self.depth_start = cd
            self._depth_confirmed = True

            if need_step:
                self.step_m = 0.05 if step_var.get().strip() == "5" else 0.10
                self._step_confirmed = True

            # индивидуальные h0
            self.depth0_by_tid = {}
            for (t, tid, h0_var, ent, dt_var, dt_lbl) in row_vars:
                dv = _parse_depth_str(h0_var.get())
                if dv is None:
                    dv = cd
                if not (0.0 <= dv <= 4.0):
                    msg_var.set(f"СЗ-{tid}: начальная глубина должна быть 0..4 м.")
                    return
                self.depth0_by_tid[int(tid)] = float(dv)

            # обновить построение глубин здесь же (без перезагрузки)
            try:
                step = float(self.step_m or 0.10)
                for (t, tid, h0_var, ent, dt_var, dt_lbl) in row_vars:
                    d0 = float(self.depth0_by_tid.get(int(tid), float(self.depth_start or 0.0)))
                    if getattr(t, "qc", None) is not None:
                        t.depth = [f"{(d0 + i * step):g}" for i in range(len(t.qc))]
            except Exception:
                pass

            # обновим поля панели, если есть
            try:
                self.depth_var.set(f"{cd:g}")
            except Exception:
                pass
            try:
                if need_step:
                    self.step_choice.set(step_var.get().strip())
            except Exception:
                pass

            result["ok"] = True
            dlg.destroy()

        def on_cancel():
            dlg.destroy()

        ttk.Button(btns, text="Отмена", command=on_cancel).pack(side="right")
        ttk.Button(btns, text="OK", command=on_ok).pack(side="right", padx=(0, 10))

        try:
            dlg.bind("<Escape>", lambda e: on_cancel())
            # Enter = OK, но если курсор в ячейках начальной глубины — оставляем их логику (переход по Enter)
            def _on_dlg_return(e):
                w = None
                try:
                    w = dlg.focus_get()
                except Exception:
                    w = None
                if getattr(w, "_geo_param_entry", False):
                    return
                on_ok()
                return "break"
            dlg.bind("<Return>", _on_dlg_return)

            common_ent.focus_set()
            common_ent.selection_range(0, "end")
        except Exception:
            pass

        try:
            dlg.update_idletasks()
            self._center_child(dlg)
        except Exception:
            pass

        self.wait_window(dlg)
        return bool(result["ok"])
    def _set_geo_inputs_enabled(self, enabled: bool):
        # H0 and step controls are required only for GEO/GE0, not for GXL.
        state = "normal" if enabled else "disabled"
        for attr in ("h0_entry", "rb5", "rb10"):
            w = getattr(self, attr, None)
            if w is not None:
                try:
                    w.config(state=state)
                except Exception:
                    pass
        for attr in ("h0_label", "step_label"):
            w = getattr(self, attr, None)
            if w is not None:
                try:
                    w.config(fg="#333333" if enabled else "#888888")
                except Exception:
                    pass

    def _depth_at(self, idx: int) -> float:
        return round(float(self.depth_start) + idx * float(self.step_m), 4)

    def load_and_render(self):

            if not self.geo_path:

                messagebox.showwarning("Нет файла", "Сначала выбери файл.")

                return


            if getattr(self, "is_gxl", False) or self.geo_path.suffix.lower() == ".gxl":

                try:

                    tests_list, meta_rows = parse_gxl_file(self.geo_path)
                    self.loaded_path = str(self.geo_path)
                    self.is_gxl = True
                    self._geo_template_blocks_info = []
                    self._geo_template_blocks_info_full = []
                    self.original_bytes = None

                except Exception as e:

                    messagebox.showerror("Ошибка чтения GXL", str(e))

                    return


                self.meta_rows = meta_rows

                self.flags.clear()

                self.tests.clear()


                if tests_list:

                    self.depth_start = _parse_depth_float(tests_list[0].depth[0]) or 0.0

                    if len(tests_list[0].depth) >= 2:

                        d0 = _parse_depth_float(tests_list[0].depth[0])

                        d1 = _parse_depth_float(tests_list[0].depth[1])

                        self.step_m = (d1 - d0) if (d0 is not None and d1 is not None) else 0.05

                    else:

                        self.step_m = 0.05

                else:

                    self.depth_start = 0.0

                    self.step_m = 0.05


                for t in tests_list:

                    self.tests.append(t)

                    self.flags[t.tid] = TestFlags(False, set(), set(), set(), set(), set())


                self._end_edit(commit=False)

                self._redraw()

                self.undo_stack.clear()

                self.redo_stack.clear()

                self._apply_gxl_calibration_from_meta(meta_rows)
                self._update_status_loaded(prefix=f"GXL: загружено опытов {len(self.tests)}")

                self._auto_scan_after_load()

                return
            # GEO/GE0: читаем и разбираем без требований к параметрам
            try:
                data = self.geo_path.read_bytes()
                self.loaded_path = str(self.geo_path)
                self.is_gxl = False
                self.geo_kind = detect_geo_kind(data)
                self.original_bytes = data
                tests_list, meta_rows = (parse_k4_geo_strict(data), []) if (detect_geo_kind(data)=="K4") else parse_geo_with_blocks(data)
                # store template blocks (do not depend on current edited/deleted tests)
                self._geo_template_blocks_info = [t.block for t in tests_list if getattr(t, 'block', None)]
                self._geo_template_blocks_info_full = list(self._geo_template_blocks_info)

            except Exception as e:
                messagebox.showerror("Ошибка чтения", str(e))
                return


            # GEO/GE0: параметры глубины/шага
            # K2: в GEO/GE0 обычно нет надёжных параметров — считаем неизвестными и спрашиваем у пользователя.
            # K4: start/step уже зашиты в шапке опыта (marker) и depths уже посчитаны парсером.
            if getattr(self, "geo_kind", "K2") == "K4" and tests_list:
                # Попробуем восстановить шаг/начальную глубину для внутренних нужд UI (без диалогов)
                try:
                    d0 = _parse_depth_float(tests_list[0].depth[0]) or 0.0
                    d1 = _parse_depth_float(tests_list[0].depth[1]) if len(tests_list[0].depth) > 1 else None
                    self.depth_start = float(d0)
                    self.step_m = float(d1 - d0) if (d1 is not None and d0 is not None) else 0.05
                except Exception:
                    self.depth_start = 0.0
                    self.step_m = 0.05
                self._depth_confirmed = True
                self._step_confirmed = True
            else:
                self.depth_start = 0.0
                self.step_m = None
                self._depth_confirmed = False
                self._step_confirmed = False

            # Параметры глубины/шага в K2 GEO могут отсутствовать.
            # Если глубина отсутствует или равна 0 — просим указать в диапазоне 0..4 м.
            need_depth = (getattr(self, "depth_start", None) is None) or ((not getattr(self, "_depth_confirmed", False)) and (float(getattr(self, "depth_start", 0.0) or 0.0) == 0.0))
            # Если шаг отсутствует — предлагаем 5 или 10 см.
            need_step = (getattr(self, "step_m", None) is None) or (not getattr(self, "_step_confirmed", False))

            # K4: start/step уже зашиты в самом файле (marker), и глубины мы уже посчитали в parse_k4_geo_strict().
            # Поэтому диалог "Параметры GEO" не показываем.
            if getattr(self, "geo_kind", "K2") == "K4":
                need_depth = False
                need_step = False
                self._depth_confirmed = True
                self._step_confirmed = True

            if need_depth or need_step:
                ok = self._prompt_geo_build_params(tests_list, need_depth=need_depth, need_step=need_step)
                if not ok:
                    return

            self.meta_rows = meta_rows
            self.flags.clear()
            self.tests.clear()

            step = float(self.step_m or 0.05)
            for t in tests_list:
                # K4: depth уже рассчитана парсером (start/step из шапки опыта).
                if getattr(self, "geo_kind", "K2") != "K4":
                    tid = int(getattr(t, "tid", 0) or 0)
                    d0 = float(self.depth0_by_tid.get(tid, float(self.depth_start or 0.0)))
                    t.depth = [f"{(d0 + i * step):g}" for i in range(len(t.qc))]
                # K4: сохраним начальную глубину и дату/время для окна «Параметры GEO»
                try:
                    if getattr(self, "geo_kind", "K2") == "K4" and getattr(t, "depth", None):
                        self.depth0_by_tid[int(t.tid)] = float(_parse_depth_float(t.depth[0]) or 0.0)
                        if getattr(t, "dt", None):
                            self.dt_by_tid[int(t.tid)] = str(t.dt)
                except Exception:
                    pass
                self.tests.append(t)
                self.flags[t.tid] = TestFlags(False, set(), set(), set(), set(), set())

            self._end_edit(commit=False)
            self._redraw()
            self.undo_stack.clear()
            self.redo_stack.clear()

            self._update_status_loaded(prefix=f"GEO: загружено опытов {len(self.tests)}")

            self._auto_scan_after_load()
            return


            try:

                data = self.geo_path.read_bytes()

                self.original_bytes = data

                tests_list, meta_rows = parse_geo_with_blocks(data)
                # store template blocks (do not depend on current edited/deleted tests)
                self._geo_template_blocks_info = [t.block for t in tests_list if getattr(t, 'block', None)]
                self._geo_template_blocks_info_full = list(self._geo_template_blocks_info)


            except Exception as e:

                messagebox.showerror("Ошибка чтения", str(e))

                return


            self.meta_rows = meta_rows

            self.flags.clear()

            self.tests.clear()


            step = float(self.step_m or 0.05)
            for t in tests_list:
                tid = int(getattr(t, "tid", 0) or 0)
                d0 = float(self.depth0_by_tid.get(tid, float(self.depth_start or 0.0)))
                t.depth = [f"{(d0 + i * step):g}" for i in range(len(t.qc))]

                self.tests.append(t)

                self.flags[t.tid] = TestFlags(False, set(), set(), set(), set(), set())


            self._end_edit(commit=False)

            self._redraw()

            self.undo_stack.clear()

            self.redo_stack.clear()

            self._update_status_loaded(f"Загружено опытов {len(self.tests)} шт.")


    def _scan_by_algorithm(self):
        """Скан-проверка: подсветить, но не менять значения (qc/fs).
        Возвращает сводку (dict) и обновляет self.flags для подсветки.
        """
        summary = {
            "tests_total": 0,
            "tests_invalid": 0,
            "cells_interp": 0,
            "cells_force": 0,
            "cells_missing": 0,
        }
        if not self.tests:
            return summary

        self._algo_preview_mode = True
        summary["tests_total"] = len(self.tests)

        for t in self.tests:
            tid = t.tid
            prev = self.flags.get(tid) or TestFlags(False, set(), set(), set(), set(), set())
            user_cells = set(getattr(prev, "user_cells", set()) or set())
            # сохраняем ранее подсвеченные (чтобы повторный скан не затирал)
            interp_cells = set(getattr(prev, "interp_cells", set()) or set())
            force_cells = set(getattr(prev, "force_cells", set()) or set())
            force_tail_rows = set(getattr(prev, "force_tail_rows", set()) or set())

            qc = [(_parse_cell_int(v) or 0) for v in t.qc]
            fs = [(_parse_cell_int(v) or 0) for v in t.fs]

            # Считаем отсутствующие значения (нули/пусто) для нижней строки, даже если опыт некорректный.
            try:
                for i0 in range(min(len(qc), len(fs))):
                    if qc[i0] == 0 and (i0, "qc") not in user_cells:
                        summary["cells_missing"] += 1
                    if fs[i0] == 0 and (i0, "fs") not in user_cells:
                        summary["cells_missing"] += 1
            except Exception:
                pass

            invalid = (_max_zero_run(qc) > 5) or (_max_zero_run(fs) > 5)

            if invalid:
                self.flags[tid] = TestFlags(True, interp_cells, force_cells, user_cells, set(), set())
                summary["tests_invalid"] += 1
                continue

            # отметим короткие серии нулей (<=5) как кандидаты на интерполяцию
            def mark_short_zero_runs(arr, kind):
                n = len(arr)
                i = 0
                while i < n:
                    if arr[i] != 0:
                        i += 1
                        continue
                    j = i
                    while j < n and arr[j] == 0:
                        j += 1
                    gap = j - i
                    if 1 <= gap <= 5:
                        for k in range(gap):
                            cell = (i + k, kind)
                            # не перетираем ручные
                            if cell not in user_cells:
                                if cell not in interp_cells:
                                    interp_cells.add(cell)
                                    summary["cells_interp"] += 1
                    i = j

            mark_short_zero_runs(qc, "qc")
            mark_short_zero_runs(fs, "fs")

            # кандидаты на "дописать хвост": подсветить СИНЕЙ строкой ниже последней глубины
            # правило: если ни один столбец (qc и fs) не достиг 250
            try:
                qc_max = max(qc) if qc else 0
                fs_max = max(fs) if fs else 0
                # Подсветка СИНИМ (без конфликтов со скрытием строк):
                # если ОБА параметра (qc и fs) не дошли до 250, подсвечиваем ПОСЛЕДНИЕ 2 ЯЧЕЙКИ (qc+fs)
                # последней существующей строки опыта.
                if qc and fs and (qc_max < 250 and fs_max < 250):
                    last_row = max(0, len(t.depth) - 1)
                    # force_cells подсвечивает конкретные ячейки (row, 'qc'/'fs')
                    if (last_row, "qc") not in user_cells:
                        force_cells.add((last_row, "qc"))
                    if (last_row, "fs") not in user_cells:
                        force_cells.add((last_row, "fs"))
                    # cells_force считаем как "опытов без отказа" (1 раз на опыт)
                    summary["cells_force"] += 1
                    force_tail_rows = set(getattr(prev, "force_tail_rows", set()) or set())
                else:
                    force_tail_rows = set(getattr(prev, "force_tail_rows", set()) or set())
            except Exception:
                force_tail_rows = set(getattr(prev, "force_tail_rows", set()) or set())


            # сохраняем зелёную подсветку откорректированных ячеек и хвостовые строки (Undo/Redo + скан)
            prev_algo_cells = set(getattr(prev, 'algo_cells', set()) or set())
            prev_force_tail_rows = set(getattr(prev, 'force_tail_rows', set()) or set())
            self.flags[tid] = TestFlags(False, interp_cells, force_cells, user_cells, prev_algo_cells, force_tail_rows or prev_force_tail_rows)

        self._redraw()
        return summary


    def _set_footer_from_scan(self):
        """Поставить НИЖНЮЮ строку (footer_cmd) по текущей автопроверке.
        Важно: всегда перезаписывает цвет (красный/серый), чтобы не оставалось синего после отката.
        """
        try:
            info = self._scan_by_algorithm()
        except Exception:
            info = {}
        try:
            inv = int(info.get("tests_invalid", 0) or 0)
            miss = int(info.get("cells_missing", 0) or 0)
            no_ref = int(info.get("cells_force", 0) or 0)

            parts = []
            if inv:
                parts.append(f"Некорректный опыт {inv}")
            if miss:
                parts.append(f"отсутствуют значения {miss}")
            if no_ref:
                parts.append(f"отсутствует отказ {no_ref}")

            msg = ", ".join(parts)
            # Сначала цвет, потом текст — так надёжнее для ttk
            try:
                self.footer_cmd.config(foreground=("#8B0000" if msg else "#666666"))
            except Exception:
                pass
            self.footer_cmd.config(text=msg)
        except Exception:
            pass

    def _compute_footer_realtime(self):
        """Пересчитать нижнюю строку (в реальном времени) по ТЕКУЩИМ данным.
        Правила:
          - 'Некорректный опыт X' — количество опытов с invalid=True (или по критерию >5 нулей подряд).
          - 'отсутствуют значения Y' — количество нулевых ячеек qc/fs ТОЛЬКО по корректным опытам.
          - 'отсутствует отказ Z' — количество корректных опытов, где qc_max<250 И fs_max<250.
        """
        try:
            tests = list(getattr(self, "tests", []) or [])
            if not tests:
                return {"inv": 0, "miss": 0, "no_ref": 0}

            inv = 0
            miss = 0
            no_ref = 0

            for t in tests:
                tid = getattr(t, "tid", None)
                # Если опыт отключён галочкой (не экспортировать) — исключаем его из пересчёта.
                if not bool(getattr(t, "export_on", True)):
                    continue
                qc = [(_parse_cell_int(v) or 0) for v in (getattr(t, "qc", []) or [])]
                fs = [(_parse_cell_int(v) or 0) for v in (getattr(t, "fs", []) or [])]

                # invalid: считаем по критерию ВСЕГДА (и учитываем сохранённый флаг),
                # иначе добавленные/новые опыты с флагом invalid=False и нулями не попадут в статистику.
                fl = (getattr(self, "flags", {}) or {}).get(tid)
                invalid_flag = bool(getattr(fl, "invalid", False)) if fl is not None else False
                try:
                    invalid_calc = (_max_zero_run(qc) > 5) or (_max_zero_run(fs) > 5)
                except Exception:
                    invalid_calc = False
                invalid = bool(invalid_flag or invalid_calc)


                if invalid:
                    inv += 1
                    continue  # нули некорректного опыта не считаем в 'отсутствуют значения' и 'отсутствует отказ'

                # missing zeros (only valid tests)
                user_cells = set(getattr(fl, "user_cells", set()) or set()) if fl is not None else set()
                n = min(len(qc), len(fs))
                for i0 in range(n):
                    if qc[i0] == 0 and (i0, "qc") not in user_cells:
                        miss += 1
                    if fs[i0] == 0 and (i0, "fs") not in user_cells:
                        miss += 1

                try:
                    qc_max = max(qc) if qc else 0
                    fs_max = max(fs) if fs else 0
                    if qc and fs and (qc_max < 250 and fs_max < 250):
                        no_ref += 1
                except Exception:
                    pass

            return {"inv": inv, "miss": miss, "no_ref": no_ref}
        except Exception:
            return {"inv": 0, "miss": 0, "no_ref": 0}

    def _update_footer_realtime(self):
        """Обновить нижнюю строку (красная/серая) по текущему состоянию."""
        try:
            res = self._compute_footer_realtime()
            inv = int(res.get("inv", 0) or 0)
            miss = int(res.get("miss", 0) or 0)
            no_ref = int(res.get("no_ref", 0) or 0)

            parts = []
            if inv:
                parts.append(f"Некорректный опыт {inv}")
            if miss:
                parts.append(f"отсутствуют значения {miss}")
            if no_ref:
                parts.append(f"отсутствует отказ {no_ref}")

            msg = ", ".join(parts)
            # Если всё ОК (включая учёт отключённых опытов) — показываем синюю надпись
            if not msg:
                try:
                    self.footer_cmd.config(foreground="#0b5ed7")
                except Exception:
                    pass
                self.footer_cmd.config(text="Статическое зондирование откорректировано.")
                return
            try:
                self.footer_cmd.config(foreground="#8B0000")
            except Exception:
                pass
            self.footer_cmd.config(text=msg)
        except Exception:
            pass

    def _footer_live_tick(self):
        """Таймер: держит нижнюю строку актуальной при удалениях/ручных правках."""
        try:
            # Не перебиваем синее сообщение 'откорректировано' сразу после корректировки:
            # если пользователь ничего не менял, оно останется до следующего действия.
            if getattr(self, "_footer_force_live", True):
                self._update_footer_realtime()
        except Exception:
            pass
        try:
            self.after(350, self._footer_live_tick)
        except Exception:
            pass


    def _auto_scan_after_load(self):
        """Автопроверка при открытии: подсветить (бледно), без изменений, без всплывающих окон.
        Пишет сводку в подвал (status).
        """
        try:
            info = self._scan_by_algorithm()
            bad = (info.get("tests_invalid", 0) + info.get("cells_interp", 0) + info.get("cells_force", 0))
            if bad <= 0:
                self._algo_preview_mode = False
                self._redraw()
                self.footer_cmd.config(text="")
                return

            # оставляем предпросмотр (бледная подсветка)
            self._algo_preview_mode = True
            self._redraw()

            # Сформировать нижнюю строку по заданному формату:
            # 'Некорректный опыт 1, отсутствуют значения 14, отсутствует отказ 2'
            inv = int(info.get("tests_invalid", 0) or 0)
            miss = int(info.get("cells_missing", 0) or 0)
            no_ref = int(info.get("cells_force", 0) or 0)

            parts = []
            if inv:
                parts.append(f"Некорректный опыт {inv}")
            if miss:
                parts.append(f"отсутствуют значения {miss}")
            if no_ref:
                parts.append(f"отсутствует отказ {no_ref}")

            msg = ", ".join(parts)
            self.footer_cmd.config(text=msg)
            try:
                self.footer_cmd.config(foreground=("#8B0000" if msg else "#666666"))
            except Exception:
                pass
        except Exception:
            self._algo_preview_mode = False

    def add_test(self):
        """
        Добавление нового зондирования через диалог:
        - Номер (по порядку)
        - Дата/время (по умолчанию: сегодня, +10 минут, секунды случайные; в поле секунды не показываем)
        - Начальная глубина (по умолчанию: наиболее частое стартовое значение среди опытов)
        - Конечная глубина (по умолчанию: последнее значение глубины последнего опыта)
        OK / Enter — подтвердить, Esc — отмена.
        """
        if self.depth_start is None or self.step_m is None:
            messagebox.showwarning("Внимание", "Сначала загрузите зондирования (чтобы была задана глубина/шаг).")
            return
        if not self.tests:
            messagebox.showwarning("Внимание", "Сначала нажмите «Показать зондирования».")
            return

        def _f(v):
            try:
                return float(str(v).replace(",", ".").strip())
            except Exception:
                return None

        def _mode_start_depth():
            # Берём первое валидное значение глубины из каждого опыта и выбираем самое частое
            counts = {}
            for t in self.tests:
                d = None
                for x in (getattr(t, "depth", []) or []):
                    if str(x).strip() == "":
                        continue
                    d = _f(x)
                    if d is not None:
                        break
                if d is None:
                    # fallback на depth0_by_tid, если есть
                    d = _f(self.depth0_by_tid.get(getattr(t, "tid", None), None)) if getattr(self, "depth0_by_tid", None) else None
                if d is None:
                    continue
                # округлим до миллиметра, чтобы не плодить почти-одинаковые ключи
                k = round(d, 3)
                counts[k] = counts.get(k, 0) + 1
            if not counts:
                return float(self.depth_start or 0.0)
            # max по частоте, при равенстве — минимальная глубина
            best = sorted(counts.items(), key=lambda kv: (-kv[1], kv[0]))[0][0]
            return float(best)

        def _last_end_depth():
            # Последнее валидное значение глубины из последнего опыта
            t = self.tests[-1]
            arr = getattr(t, "depth", []) or []
            for x in reversed(arr):
                if str(x).strip() == "":
                    continue
                d = _f(x)
                if d is not None:
                    return float(d)
            # fallback: старт + шаг * (len-1)
            step = float(self.step_m or 0.05)
            d0 = _mode_start_depth()
            n = max(1, len(getattr(t, "depth", []) or []))
            return float(d0 + step * (n - 1))

        # defaults
        new_id = (max((t.tid for t in self.tests), default=0) + 1)
        d0_default = _mode_start_depth()
        d1_default = _last_end_depth()
        if d1_default < d0_default:
            d1_default = d0_default

        now_dt = _dt.datetime.now().replace(microsecond=0)
        dt_default_dt = (now_dt + _dt.timedelta(minutes=10)).replace(second=random.randint(0, 59))
        dt_default_str = dt_default_dt.strftime("%Y-%m-%d %H:%M")  # в поле без секунд

        # ---- dialog ----
        dlg = tk.Toplevel(self)
        dlg.title("Добавить зондирование")
        dlg.transient(self)
        dlg.grab_set()

        frm = ttk.Frame(dlg, padding=12)
        frm.pack(fill="both", expand=True)

        def _row(r, label, var):
            ttk.Label(frm, text=label).grid(row=r, column=0, sticky="w", pady=4)
            e = ttk.Entry(frm, textvariable=var, width=24)
            e.grid(row=r, column=1, sticky="we", pady=4)
            return e

        frm.columnconfigure(1, weight=1)

        
        v_id = tk.StringVar(value=str(new_id))

        # Раздельно: дата (с календарём) + время (HH:MM). Секунды НЕ показываем, но сохраняем (случайные).
        v_date = tk.StringVar(value=dt_default_dt.strftime("%Y-%m-%d"))
        v_time = tk.StringVar(value=dt_default_dt.strftime("%H:%M"))

        v_d0 = tk.StringVar(value=f"{d0_default:g}")
        v_d1 = tk.StringVar(value=f"{d1_default:g}")

        e_id = _row(0, "Номер:", v_id)


        def _open_calendar(_evt=None):
            # Используем уже реализованный CalendarDialog (с подсветкой выбранной даты и запретом будущих дат)
            try:
                s = (v_date.get() or "").strip()
                cur_date = None
                for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y"):
                    try:
                        cur_date = _dt.datetime.strptime(s, fmt).date()
                        break
                    except Exception:
                        pass
                if cur_date is None:
                    cur_date = _dt.date.today()
                if cur_date > _dt.date.today():
                    cur_date = _dt.date.today()

                cd = CalendarDialog(dlg, initial=cur_date, title="Выбор даты")
                self._center_child(cd)
                dlg.wait_window(cd)
                sel = getattr(cd, "selected", None)
                if sel:
                    v_date.set(sel.strftime("%Y-%m-%d"))
            except Exception:
                pass

        # --- Date row with calendar ---
        ttk.Label(frm, text="Дата:").grid(row=1, column=0, sticky="w", pady=4)
        date_row = ttk.Frame(frm)
        date_row.grid(row=1, column=1, sticky="we", pady=4)
        date_row.columnconfigure(0, weight=1)
        e_date = ttk.Entry(date_row, textvariable=v_date, width=14)
        e_date.grid(row=0, column=0, sticky="we")
        btn_cal = ttk.Button(date_row, text="📅", width=3)
        btn_cal.grid(row=0, column=1, padx=(6, 0))
        btn_cal.config(command=_open_calendar)

        ttk.Label(frm, text="Время (HH:MM):").grid(row=2, column=0, sticky="w", pady=4)
        e_time = ttk.Entry(frm, textvariable=v_time, width=10)
        e_time.grid(row=2, column=1, sticky="w", pady=4)

        e_d0 = _row(3, "Начальная глубина, м:", v_d0)
        e_d1 = _row(4, "Конечная глубина, м:", v_d1)

        btns = ttk.Frame(frm)
        btns.grid(row=5, column=0, columnspan=2, sticky="e", pady=(10, 0))

        result = {"ok": False}

        def _parse_date_time(date_s: str, time_s: str):
            date_s = (date_s or '').strip()
            time_s = (time_s or '').strip()
            # допускаем ввод даты как YYYY-MM-DD или DD.MM.YYYY
            d = None
            for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d.%m.%y'):
                try:
                    d = _dt.datetime.strptime(date_s, fmt).date()
                    break
                except Exception:
                    pass
            if d is None:
                return None
            # время HH:MM
            t = None
            for fmt in ('%H:%M',):
                try:
                    t = _dt.datetime.strptime(time_s, fmt).time()
                    break
                except Exception:
                    pass
            if t is None:
                return None
            return _dt.datetime.combine(d, t)

        def _ok(_evt=None):
            tid = None
            try:
                tid = int(v_id.get().strip())
            except Exception:
                messagebox.showwarning("Ошибка", "Номер должен быть целым числом.", parent=dlg)
                return
            if tid <= 0:
                messagebox.showwarning("Ошибка", "Номер должен быть > 0.", parent=dlg)
                return

            # запрет коллизий по tid
            existing_ids = {t.tid for t in self.tests}
            if tid in existing_ids:
                messagebox.showwarning("Ошибка", f"Зондирование №{tid} уже существует.", parent=dlg)
                return

            d0 = _f(v_d0.get())
            d1 = _f(v_d1.get())
            if d0 is None or d1 is None:
                messagebox.showwarning("Ошибка", "Глубины должны быть числами.", parent=dlg)
                return
            if d1 < d0:
                d0, d1 = d1, d0  # автоматически поменяем местами

            dt_user = _parse_date_time(v_date.get(), v_time.get())
            if dt_user is None:
                messagebox.showwarning("Ошибка", "Введите дату и время. Дата: YYYY-MM-DD (или DD.MM.YYYY), время: HH:MM.", parent=dlg)
                return
            # секунды всегда генерируем случайно (в поле не показываем)
            dt_user = dt_user.replace(second=random.randint(0, 59), microsecond=0)

            result.update(ok=True, tid=tid, d0=float(d0), d1=float(d1), dt=dt_user)
            dlg.destroy()

        def _cancel(_evt=None):
            dlg.destroy()


        ttk.Button(btns, text="Отмена", command=_cancel).pack(side="right", padx=(6, 0))
        ttk.Button(btns, text="OK", command=_ok).pack(side="right")

        dlg.bind("<Return>", _ok)
        dlg.bind("<Escape>", _cancel)

        dlg.update_idletasks()
        try:
            self._center_child(dlg)
        except Exception:
            pass

        dlg.minsize(dlg.winfo_width(), dlg.winfo_height())
        e_id.focus_set()
        e_id.selection_range(0, "end")

        self.wait_window(dlg)
        if not result.get("ok"):
            return

        # ---- apply ----
        self._push_undo()

        tid = int(result["tid"])
        d0 = float(result["d0"])
        d1 = float(result["d1"])
        dt_val = result["dt"]

        # сохраняем стартовую глубину для опыта
        self.depth0_by_tid[int(tid)] = d0

        step = float(self.step_m or 0.05)
        n = int(round((d1 - d0) / step)) + 1
        n = max(1, n)

        depth = [f"{(d0 + i * step):g}" for i in range(n)]
        qc = ["0"] * n
        fs = ["0"] * n

        now = dt_val.strftime("%Y-%m-%d %H:%M:%S")

        # вставляем по времени (у тебя везде логика: показывать по времени)
        new_test = TestData(tid=tid, dt=now, depth=depth, qc=qc, fs=fs, orig_id=None, block=None)

        def _parse_dt_any(s):
            try:
                return _dt.datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
            except Exception:
                try:
                    return _dt.datetime.strptime(s, "%Y-%m-%d %H:%M")
                except Exception:
                    return None

        insert_at = len(self.tests)
        new_dt_cmp = _parse_dt_any(now) or _dt.datetime.max
        for i, t in enumerate(self.tests):
            td = _parse_dt_any(getattr(t, "dt", "") or "")
            if td is None:
                continue
            if td > new_dt_cmp:
                insert_at = i
                break

        self.tests.insert(insert_at, new_test)
        self.flags[tid] = TestFlags(False, set(), set(), set(), set(), set())

        self._end_edit(commit=False)
        self._redraw()

        # Механика автопрокрутки как при копировании:
        # если добавленное зондирование не помещается — прокручиваем по X к нему,
        # при этом используем _xview_proxy, чтобы шапка и тело оставались синхронны.
        try:
            self.update_idletasks()
        except Exception:
            pass

        def _scroll_to_new():
            try:
                self._ensure_cell_visible(insert_at, 0, 'depth', pad=12)
            except Exception:
                try:
                    self._xview_proxy("moveto", 1.0)
                except Exception:
                    pass

        try:
            self.after_idle(_scroll_to_new)
        except Exception:
            _scroll_to_new()
        self.status.config(text=f"Добавлена новая зондирование {tid}. (В GEO-сохранение не попадёт — только Excel)")

        try:
            self._set_footer_from_scan()
        except Exception:
            pass

    # ---------------- conversion 10cm -> 5cm ----------------
    def convert_10_to_5(self):
        # 10см -> 5см: вставляем промежуточную строку ТОЛЬКО между двумя валидными соседними строками.
        # Пустые/удаленные строки не трогаем и не заполняем.
        if not self.tests:
            messagebox.showwarning("Нет данных", "Сначала загрузи и покажи зондирования.")
            return
        self._push_undo()



        for t in self.tests:
            old_flags = self.flags.get(t.tid) or TestFlags(False, set(), set(), set(), set(), set())

            n = min(len(t.depth), len(t.qc), len(t.fs))
            if n < 2:
                continue

            def blank(i):
                return (str(t.depth[i]).strip()=="" and str(t.qc[i]).strip()=="" and str(t.fs[i]).strip()=="")

            def valid(i):
                if blank(i):
                    return False
                d = _parse_depth_float(t.depth[i])
                if d is None:
                    return False
                q = _parse_cell_int(t.qc[i])
                f = _parse_cell_int(t.fs[i])
                return (q is not None or f is not None)

            new_depth, new_qc, new_fs = [], [], []
            map_old_to_new: dict[int,int] = {}
            created_rows: list[int] = []

            i = 0
            while i < n:
                map_old_to_new[i] = len(new_depth)
                new_depth.append(t.depth[i])
                new_qc.append(t.qc[i])
                new_fs.append(t.fs[i])

                if i+1 < n and valid(i) and valid(i+1):
                    d0 = _parse_depth_float(t.depth[i])
                    d1 = _parse_depth_float(t.depth[i+1])
                    if d0 is not None and d1 is not None and 0.09 <= (d1-d0) <= 0.11:
                        dm = d0 + 0.05
                        q0 = _parse_cell_int(t.qc[i]) or 0
                        q1 = _parse_cell_int(t.qc[i+1]) or 0
                        f0 = _parse_cell_int(t.fs[i]) or 0
                        f1 = _parse_cell_int(t.fs[i+1]) or 0
                        qm = int(round((q0+q1)/2))
                        fm = int(round((f0+f1)/2))

                        created_rows.append(len(new_depth))
                        new_depth.append(f"{dm:.2f}")
                        new_qc.append(str(qm))
                        new_fs.append(str(fm))
                i += 1

            t.depth, t.qc, t.fs = new_depth, new_qc, new_fs

            # перенос подсветки (не теряем фиолетовый и др. цвета)
            new_interp: set[tuple[int,str]] = set()
            new_force: set[tuple[int,str]] = set()
            new_user: set[tuple[int,str]] = set()
            new_algo: set[tuple[int,str]] = set()
            new_tail: set[int] = set()

            for (r, fld) in (old_flags.interp_cells or set()):
                if r in map_old_to_new:
                    new_interp.add((map_old_to_new[r], fld))
            for (r, fld) in (old_flags.force_cells or set()):
                if r in map_old_to_new:
                    new_force.add((map_old_to_new[r], fld))
            for (r, fld) in (old_flags.user_cells or set()):
                if r in map_old_to_new:
                    new_user.add((map_old_to_new[r], fld))
            for (r, fld) in (old_flags.algo_cells or set()):
                if r in map_old_to_new:
                    new_algo.add((map_old_to_new[r], fld))
            for r in (old_flags.force_tail_rows or set()):
                if r in map_old_to_new:
                    new_tail.add(map_old_to_new[r])

            # новые созданные строки (вставки при 10→5) помечаем ЗЕЛЁНЫМ как «откорректировано»
            for rr in created_rows:
                new_algo.add((rr, "qc"))
                new_algo.add((rr, "fs"))

            # если после 10→5 появился критерий некорректности (>5 нулей подряд) — считаем опыт некорректным (красным)
            try:
                qv = [(_parse_cell_int(v) or 0) for v in (t.qc or [])]
                fv = [(_parse_cell_int(v) or 0) for v in (t.fs or [])]
                invalid_now = bool(old_flags.invalid) or (_max_zero_run(qv) > 5) or (_max_zero_run(fv) > 5)
            except Exception:
                invalid_now = bool(old_flags.invalid)

            self.flags[t.tid] = TestFlags(bool(invalid_now), new_interp, new_force, new_user, new_algo, new_tail)

        # после конвертации считаем шаг 5 см
        try:
            self.step_m = 0.05
            self.step_choice.set("5")
        except Exception:
            pass

        self._redraw()
        self.status.config(text="Конвертация 10→5 выполнена. Новые строки помечены зелёным.")

    # ---------------- drawing helpers ----------------
    def _content_size(self):
        # Размеры контента для scrollregion.
        # Важно: таблица (цифры) теперь в отдельном canvas и скроллится по Y без шапки.
        max_rows = 0
        try:
            if getattr(self, "_grid", None):
                max_rows = len(self._grid)
            else:
                max_rows = max((len(t.qc) for t in self.tests), default=0)
        except Exception:
            max_rows = max((len(t.qc) for t in self.tests), default=0)

        col_w = self.w_depth + self.w_val*2 + (self.w_val if getattr(self, "geo_kind", "K2")=="K4" else 0)
        self._last_col_w = col_w
        total_w = self.pad_x * 2 + (col_w * len(self.tests)) + (self.col_gap * max(0, len(self.tests) - 1))
        body_h = max_rows * self.row_h
        header_h = int(self.pad_y + self.hdr_h)  # фиксированная область
        return total_w, body_h, header_h

    def _update_scrollregion(self):
        # сохраняем пиксельный сдвиг по X, чтобы после изменения ширины scrollregion
        # (например, после добавления зондирования) шапка и тело не расходились.
        try:
            old_w = float(getattr(self, "_scroll_w", 0) or 0)
        except Exception:
            old_w = 0.0
        try:
            old_frac = float(self.canvas.xview()[0])
        except Exception:
            old_frac = 0.0
        if old_w <= 0:
            old_w = 1.0
        old_px = old_frac * old_w

        w, body_h, header_h = self._content_size()
        w_content = w

        # вычисляем "правый зазор" (даёт свободное место справа, чтобы последняя шапка не прилипала к краю)
        try:
            vw = int(self.canvas.winfo_width() or 1)
        except Exception:
            vw = 1
        gap = int(getattr(self, "_last_col_w", 0) or 0)
        if gap < 24:
            gap = 24

        need_h = (w_content > max(vw, 1))
        if not need_h:
            gap = 0

        w_total = w_content + gap
        # SAFETY: небольшой запас по ширине, чтобы горизонтальный скролл доходил до конца
        try:
            if getattr(self, "geo_kind", "K2") == "K4":
                w_total += int(self.w_val)  # +1 колонка запаса
        except Exception:
            pass

        # scroll по Y только для таблицы
        self.canvas.configure(scrollregion=(0, 0, w_total, body_h))
        # шапка: только X-сдвиг, Y фиксирован
        try:
            self.hcanvas.configure(scrollregion=(0, 0, w_total, header_h))
            self.hcanvas.configure(height=header_h)
        except Exception:
            pass


        # восстановить X-сдвиг в пикселях
        try:
            self._scroll_w = float(w_total or 1)
        except Exception:
            self._scroll_w = float(w_total or 1)
        try:
            new_frac = 0.0 if (w_total <= 1) else (old_px / float(w_total))
            if new_frac < 0.0:
                new_frac = 0.0
            if new_frac > 1.0:
                new_frac = 1.0
            # двигаем через moveto, чтобы шапка и тело совпали и на правом краю
            self.canvas.xview_moveto(new_frac)
            try:
                self.hcanvas.xview_moveto(new_frac)
            except Exception:
                pass
        except Exception:
            pass

        # Горизонтальная прокрутка: показываем только если колонки не помещаются в видимую область
        if not need_h:
            try:
                self.canvas.xview_moveto(0)
                self.hcanvas.xview_moveto(0)
            except Exception:
                pass
            # скрыть скроллбар
            try:
                if not getattr(self, "_hscroll_hidden", True):
                    self.hscroll_frame.pack_forget()
            except Exception:
                pass
            self._hscroll_hidden = True
            try:
                # ttk.Scrollbar.set ожидает (first, last)
                self.hscroll.set(0.0, 1.0)
            except Exception:
                pass
        else:
            # показать горизонтальную полосу сразу после таблицы (над статусом)
            if getattr(self, "_hscroll_hidden", True):
                try:
                    self.hscroll_frame.pack(side="bottom", fill="x")
                except Exception:
                    try:
                        self.hscroll_frame.pack(side="bottom", fill="x")
                    except Exception:
                        pass
                # статус должен быть НИЖЕ полосы — перепаковываем
                try:
                    self.status.pack_forget()
                    self.status.pack(side="bottom", fill="x", before=self.footer)
                except Exception:
                    pass
                self._hscroll_hidden = False
    def _refresh_display_order(self):
        """Order tests for rendering.

        Rule: always show tests left-to-right by **time** (chronological).
        GeoExplorer files may contain tests recorded/added out of numeric order.

        Fallback: if dt is missing/unparseable, push such tests to the end,
        keeping a stable tie-break by test id and original index.
        """

        def _tid_key(tid):
            try:
                return int(str(tid).strip())
            except Exception:
                return 10**9

        def _key(i: int):
            t = self.tests[i]
            dt = _try_parse_dt(getattr(t, "dt", "") or "")
            dt_key = dt if dt is not None else _dt.datetime.max
            return (dt_key, _tid_key(getattr(t, "tid", "")), i)

        self.display_cols = sorted(range(len(self.tests)), key=_key)

    def _on_left_click(self, event):
        self._evt_widget = event.widget
        hit = self._hit_test(event.x, event.y)
        if not hit:
            # клик вне ячеек/шапки → закрываем активное редактирование
            self._end_edit(commit=True)
            self._hide_canvas_tip()
            return
        kind, ti, row, field = hit

        # Любой клик по UI (иконки/пустые/глубина) сначала закрывает активную ячейку
        # (кроме случая, когда мы тут же откроем новое редактирование).
        try:
            if getattr(self, '_editing', None):
                # не закрываем, если клик по текущему Entry
                ed = self._editing[3] if len(self._editing) >= 4 else None
                if ed is None or event.widget is not ed:
                    self._end_edit(commit=True)
        except Exception:
            pass

        # --- Header controls (icons / checkbox) ---
        if kind == "edit":
            self._edit_header(ti)
            return
        if kind == "dup":
            self._duplicate_test(ti)
            return
        if kind == "trash":
            self._delete_test(ti)
            return
        if kind == "export":
            try:
                self._push_undo()
                t = self.tests[ti]
                t.export_on = not bool(getattr(t, "export_on", True))
            except Exception:
                pass
            self._hide_canvas_tip()
            self._redraw()
            return

        # --- Single-click cell edit (ironclad) ---
        if kind == "cell" and ti is not None and row is not None:
            mp = (getattr(self, "_grid_row_maps", {}) or {}).get(ti, {})
            start_r = (getattr(self, "_grid_start_rows", {}) or {}).get(ti, 0)

            # Depth: single click on the first depth cell opens "start depth" editor
            if field == "depth":
                if row == start_r:
                    self._begin_edit_depth0(ti, display_row=row)
                return

            # qc/fs cells
            data_row = mp.get(row, None)

            if data_row is None:
                # Вставка строки СВЕРХУ/СНИЗУ (без разрывов).
                #   - Снизу: клик только в СЛЕДУЮЩУЮ строку после последней существующей.
                #   - Сверху: клик только в СТРОКУ ПЕРЕД первой существующей.
                if field in ("qc", "fs"):
                    t = self.tests[ti]

                    # --- TOP: разрешаем дописывать "верх" по принципу "низа" ---
                    top_disp = start_r - 1
                    if row == top_disp:
                        # не даём создавать "дырки" сверху: если первая строка уже пустая — заполни её сначала
                        if len(t.qc) > 0:
                            try:
                                q0 = t.qc[0]
                                f0 = t.fs[0]
                            except Exception:
                                q0 = None
                                f0 = None
                            if (q0 in (None, "")) and (f0 in (None, "")):
                                self._set_status("Сначала заполните следующую строку")
                                return
                        # вычисляем шаг по глубинам (если есть), иначе дефолт
                        step = 0.05
                        try:
                            if len(t.depth) >= 2:
                                step = float(str(t.depth[1]).replace(",", ".")) - float(str(t.depth[0]).replace(",", "."))
                                if step <= 0:
                                    step = 0.05
                            elif getattr(self, "current_step", None):
                                step = float(self.current_step)
                        except Exception:
                            step = 0.05
                        # новая глубина = первая - step
                        d_new = 0.0
                        try:
                            if len(t.depth) >= 1:
                                d_new = float(str(t.depth[0]).replace(",", ".")) - step
                            else:
                                d_new = 0.0
                        except Exception:
                            d_new = 0.0
                        t.depth.insert(0, f"{d_new:.2f}")
                        t.qc.insert(0, "")
                        t.fs.insert(0, "")
                        # сдвигаем подсветку (флаги) вниз на 1 строку
                        try:
                            fl = self.flags.get(getattr(t, "tid", None))
                            if fl:
                                # сдвиг всех cell-наборов на +1 (вставка строки в начале)
                                def _bump(cells):
                                    out=set()
                                    for (r, knd) in (cells or set()):
                                        try:
                                            rr=int(r)
                                        except Exception:
                                            continue
                                        out.add((rr+1, knd))
                                    return out
                                fl.user_cells = _bump(getattr(fl, 'user_cells', set()))
                                fl.interp_cells = _bump(getattr(fl, 'interp_cells', set()))
                                fl.force_cells = _bump(getattr(fl, 'force_cells', set()))
                        except Exception:
                            pass
                        self._redraw()
                        self._begin_edit(ti, 0, field, display_row=row)
                        return

                    # --- BOTTOM (tail) ---
                    next_disp = start_r + len(t.qc)
                    if row != next_disp:
                        return
                    # если последняя строка существует, но она полностью пустая — не даём добавлять следующую
                    if len(t.qc) > 0:
                        try:
                            q_last = t.qc[-1]
                            f_last = t.fs[-1]
                        except Exception:
                            q_last = None
                            f_last = None
                        if (q_last in (None, "") ) and (f_last in (None, "") ):
                            self._set_status("Сначала заполните предыдущую строку")
                            return
                    new_idx = len(t.qc)
                    self._append_row(ti)
                    self._begin_edit(ti, new_idx, field, display_row=row)
                return


            # Normal in-range cell → start edit immediately
            if field in ("qc", "fs"):
                self._begin_edit(ti, data_row, field, display_row=row)
            return

        # otherwise: click ends edit (commit)
        self._end_edit(commit=True)


    def _on_global_click(self, event):
        """Закрывает активную ячейку при клике вне зондирования/ячейки."""
        try:
            if not self._editing:
                return
            # если клик по текущему Entry — не закрываем
            ed = None
            if len(self._editing) >= 4:
                ed = self._editing[3]
            if ed is not None and event.widget is ed:
                return
            # если клик по canvas/hcanvas — пусть _on_left_click решит (мы просто закрываем заранее)
            # но чтобы не мешать клику по ячейке, закрываем только когда клик НЕ по canvas/hcanvas
            if event.widget in (getattr(self, "canvas", None), getattr(self, "hcanvas", None)):
                return
            self._end_edit(commit=True)
        except Exception:
            pass
    def _set_hover(self, hover):
        # hover: ("dup"/"trash", test_index) or None
        if getattr(self, "_hover", None) == hover:
            return
        self._hover = hover
        self._hide_canvas_tip()
        # redraw only when state changed
        self._redraw()

    def _hide_canvas_tip(self):
        if getattr(self, "_hover_after", None):
            try:
                self.after_cancel(self._hover_after)
            except Exception:
                pass
            self._hover_after = None
        tip = getattr(self, "_hover_tip", None)
        if tip is not None:
            try:
                tip.destroy()
            except Exception:
                pass
            self._hover_tip = None

    def _schedule_canvas_tip(self, text: str, x_root: int, y_root: int, delay_ms: int = 2000):
        self._hide_canvas_tip()
        def _show():
            # small native-ish tooltip
            tw = tk.Toplevel(self)
            tw.overrideredirect(True)
            tw.attributes("-topmost", True)
            lbl = ttk.Label(tw, text=text, padding=(8, 4))
            lbl.pack()
            tw.update_idletasks()
            tw.geometry(f"+{x_root + 12}+{y_root + 14}")
            self._hover_tip = tw
        self._hover_after = self.after(delay_ms, _show)

    def _on_motion(self, event):
        self._evt_widget = event.widget
        hit = self._hit_test(event.x, event.y)
        if not hit:
            self._set_hover(None)
            return
        kind, ti, row, field = hit
        if kind in ("edit", "dup", "trash"):
            self._set_hover((kind, ti))
            tip_text = "Редактировать" if kind == "edit" else ("Копировать" if kind == "dup" else "Удалить")
            self._schedule_canvas_tip(tip_text, event.x_root, event.y_root, delay_ms=1000)
        elif kind == "export":
            self._set_hover((kind, ti))
            try:
                ex_on = bool(getattr(self.tests[ti], "export_on", True))
            except Exception:
                ex_on = True
            tip_text = "Исключить из экспорта" if ex_on else "Экспортировать"
            self._schedule_canvas_tip(tip_text, event.x_root, event.y_root, delay_ms=1000)
        else:
            self._set_hover(None)


    def _delete_test(self, ti: int):
        if ti < 0 or ti >= len(self.tests):
            return
        t = self.tests[ti]
        self._push_undo()
        self._end_edit(commit=True)
        del self.tests[ti]
        # NOTE: НЕ МУТИРУЕМ шаблонные блоки GEO при удалении/копировании.
        #       Иначе при экспорте GEO могут «воскресать» удалённые зондирования.
        #       Шаблон хранится в self._geo_template_blocks_info_full (неизменяемый).
        pass
        # rebuild flags
        self.flags = {tt.tid: self.flags.get(tt.tid, TestFlags(False, set(), set(), set(), set(), set())) for tt in self.tests}
        self._redraw()
        self.status.config(text=f"Опыт {t.tid} удалён.")


    def _duplicate_test(self, ti: int):
        """Duplicate a test (copy) and insert right after it.
        New tid = max(tid)+1.         New tid = max(tid)+1. New datetime = max(existing dt)+10 minutes.
        """
        if ti < 0 or ti >= len(self.tests):
            return
        if not self.tests:
            return

        self._push_undo()

        src = self.tests[ti]

        # new id sequential
        existing_ids = {t.tid for t in self.tests}
        new_id = max(existing_ids) + 1 if existing_ids else 1

        # datetime: source +10 min (try parse), else max+10
        def _parse_dt(s):
            if isinstance(s, _dt.datetime):
                return s
            s = (str(s) if s is not None else "").strip()
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M"):
                try:
                    return _dt.datetime.strptime(s, fmt)
                except Exception:
                    pass
            return None

        # datetime: max(existing dt) + 10 min (so repeated copies get unique time); fallback now()
        dts = []
        for t in self.tests:
            d = _parse_dt(getattr(t, "dt", "") or "")
            if d is not None:
                dts.append(d)
        base_dt = max(dts) if dts else (_parse_dt(getattr(src, "dt", "") or "") or _dt.datetime.now())
        new_dt_dt = (base_dt + _dt.timedelta(minutes=10)).replace(microsecond=0)
        # Добавляем случайные секунды, чтобы времена у копий были уникальнее
        new_dt_dt = new_dt_dt.replace(second=random.randint(0, 59))
        new_dt = new_dt_dt.strftime("%Y-%m-%d %H:%M:%S")

# deep copy arrays
        depth = list(getattr(src, "depth", []) or [])
        qc = list(getattr(src, "qc", []) or [])
        fs = list(getattr(src, "fs", []) or [])

        new_test = TestData(tid=new_id, dt=new_dt, depth=depth, qc=qc, fs=fs, orig_id=None, block=None)

        # insert right after source
        insert_at = min(len(self.tests), ti + 1)
        self.tests.insert(insert_at, new_test)

        # keep GEO template blocks aligned: duplicate source block template
        try:
            if getattr(self, "_geo_template_blocks_info", None) and ti < len(self._geo_template_blocks_info):
                bi = self._geo_template_blocks_info[ti]
                # shallow copy is enough (dataclass-like with numeric fields)
                import copy as _copy
                self._geo_template_blocks_info.insert(insert_at, _copy.copy(bi))
        except Exception:
            pass

        # copy flags (deep copy sets) so visual edits/удаления сохраняются в копии
        fl = self.flags.get(getattr(src, "tid", None), TestFlags(False, set(), set(), set(), set(), set()))
        try:
            self.flags[new_id] = TestFlags(bool(getattr(fl, 'invalid', False)), set(getattr(fl, 'interp_cells', set()) or set()), set(getattr(fl, 'force_cells', set()) or set()), set(getattr(fl, 'user_cells', set()) or set()), set(getattr(fl, 'algo_cells', set()) or set()), set(getattr(fl, 'force_tail_rows', set()) or set()))
        except Exception:
            try:
                self.flags[new_id] = TestFlags(False, set(), set(), set(), set(), set())
            except Exception:
                pass

        self._end_edit(commit=False)
        self._redraw()

        # если опыт не помещается на экран — прокручиваем по X так, чтобы он попал в видимую область
        try:
            self.update_idletasks()
        except Exception:
            pass

        def _scroll_to_new():
            try:
                self._ensure_cell_visible(insert_at, 0, 'depth', pad=12)
            except Exception:
                try:
                    self._xview_proxy("moveto", 1.0)
                except Exception:
                    pass

        try:
            self.after_idle(_scroll_to_new)
        except Exception:
            _scroll_to_new()

        try:
            self.status.config(text=f"Опыт {getattr(src,'tid','?')} продублирован → {new_id} (+10 мин).")
        except Exception:
            pass


        try:
            self._set_footer_from_scan()
        except Exception:
            pass

    def _cell_bbox(self, col: int, row: int, field: str):
        x0 = self.pad_x
        col_w = self.w_depth + self.w_val*2 + (self.w_val if getattr(self, "geo_kind", "K2")=="K4" else 0)
        x0 += col * (col_w + self.col_gap)
        # Таблица (цифры) рисуется в отдельном canvas и скроллится по Y,
        # поэтому старт по Y = 0 (без hdr_h).
        y0 = row * self.row_h

        if field == "depth":
            return x0, y0, x0 + self.w_depth, y0 + self.row_h
        if field == "qc":
            return x0 + self.w_depth, y0, x0 + self.w_depth + self.w_val, y0 + self.row_h
        if field == "fs":
            return x0 + self.w_depth + self.w_val, y0, x0 + self.w_depth + self.w_val + self.w_val, y0 + self.row_h
        if field == "incl":
            return x0 + self.w_depth + self.w_val*2, y0, x0 + self.w_depth + self.w_val*3, y0 + self.row_h
        raise ValueError("bad field")

    def _ensure_cell_visible(self, col: int, row: int, field: str, pad: int = 6):
        """Автопрокрутка: при навигации стрелками/Enter держим редактируемую ячейку в видимой зоне."""
        try:
            bx0, by0, bx1, by1 = self._cell_bbox(col, row, field)
        except Exception:
            return

        cnv = self.canvas
        bbox_all = cnv.bbox("all")
        if not bbox_all:
            return
        ax0, ay0, ax1, ay1 = bbox_all
        aw = max(1.0, float(ax1 - ax0))
        ah = max(1.0, float(ay1 - ay0))

        vx0, vy0, vx1, vy1 = _canvas_view_bbox(cnv)
        vw = max(1.0, float(vx1 - vx0))
        vh = max(1.0, float(vy1 - vy0))

        # Горизонталь
        target_x = None
        if bx0 < vx0 + pad:
            target_x = bx0 - pad
        elif bx1 > vx1 - pad:
            target_x = bx1 - vw + pad
        if target_x is not None:
            frac = (target_x - ax0) / aw
            frac = 0.0 if frac < 0.0 else (1.0 if frac > 1.0 else frac)
            try:
                self._xview_proxy("moveto", frac)
            except Exception:
                pass

        # Вертикаль
        target_y = None
        if by0 < vy0 + pad:
            target_y = by0 - pad
        elif by1 > vy1 - pad:
            target_y = by1 - vh + pad
        if target_y is not None:
            frac = (target_y - ay0) / ah
            frac = 0.0 if frac < 0.0 else (1.0 if frac > 1.0 else frac)
            try:
                cnv.yview_moveto(frac)
            except Exception:
                pass

    def _header_bbox(self, col: int):
        col_w = self.w_depth + self.w_val*2 + (self.w_val if getattr(self, "geo_kind", "K2")=="K4" else 0)
        x0 = self.pad_x + col * (col_w + self.col_gap)
        y0 = self.pad_y
        x1 = x0 + col_w
        y1 = y0 + self.hdr_h
        return x0, y0, x1, y1


    def _redraw(self):
        # два холста: hcanvas (фиксированная шапка) + canvas (данные)
        try:
            self.canvas.delete("all")
        except Exception:
            pass
        try:
            self.hcanvas.delete("all")
        except Exception:
            pass

        if not self.tests:
            self._update_scrollregion()
            return

        grid, grid_step, row_maps, start_rows = self._compute_depth_grid()
        if not grid:
            max_rows = (len(getattr(self, '_grid', []) or []) or max(len(t.qc) for t in self.tests))
            grid = [None] * max_rows
            row_maps = {ti: {r: r for r in range(len(self.tests[ti].qc))} for ti in range(len(self.tests))}
            start_rows = {ti: 0 for ti in range(len(self.tests))}

        max_rows = len(grid)
        self._grid = grid
        self._grid_step = grid_step
        self._grid_row_maps = row_maps
        self._grid_start_rows = start_rows

        self._refresh_display_order()

        # фиксируем высоту шапки под текущие параметры
        try:
            self.hcanvas.configure(height=int(self.pad_y + self.hdr_h))
        except Exception:
            pass

        for col, ti in enumerate(self.display_cols):
            t = self.tests[ti]
            x0, y0, x1, y1 = self._header_bbox(col)

            # checked = will be exported
            ex_on = bool(getattr(t, "export_on", True))
            hdr_fill = GUI_HDR if ex_on else "#f2f2f2"
            hdr_text = "#111" if ex_on else "#8a8a8a"
            hdr_icon = "#444" if ex_on else "#8a8a8a"

            # --- ШАПКА (hcanvas) ---
            self.hcanvas.create_rectangle(x0, y0, x1, y1, fill=hdr_fill, outline=GUI_GRID)

            dt_val = getattr(t, "dt", "") or ""
            # t.dt может быть строкой из GEO или уже datetime (после редактирования)
            if isinstance(dt_val, datetime.datetime):
                dt_line = dt_val.strftime("%d.%m.%Y %H:%M:%S")
            elif isinstance(dt_val, datetime.date):
                dt_line = dt_val.strftime("%d.%m.%Y 00:00:00")
            else:
                dt_line = str(dt_val).strip()

            # display without seconds (HH:MM)
            dt_line = re.sub(r"(\d{2}:\d{2}):\d{2}\b", r"\1", dt_line)

            # --- export checkbox (Win11 style) ---
            top_pad = 8
            row_center_y = y0 + top_pad + 6  # aligns checkbox and title vertically
            cb_s = 14
            cb_x0 = x0 + 6
            cb_y0 = int(row_center_y - cb_s/2)

            # рамка чекбокса (без hover-подсветки фоном)
            self.hcanvas.create_rectangle(cb_x0, cb_y0, cb_x0 + cb_s, cb_y0 + cb_s,
                                          fill="white", outline="#b9b9b9")
            if ex_on:
                self.hcanvas.create_line(cb_x0 + 3, cb_y0 + 7, cb_x0 + 6, cb_y0 + 10,
                                         cb_x0 + 11, cb_y0 + 4,
                                         fill="#2563eb", width=2, capstyle="round", joinstyle="round")

            # Title and datetime
            title_x = cb_x0 + cb_s + 8
            self.hcanvas.create_text(title_x, row_center_y, anchor="w",
                                     text=f"Опыт {t.tid}", font=("Segoe UI", 9, "bold"), fill=hdr_text)
            if dt_line:
                self.hcanvas.create_text(title_x, row_center_y + 18, anchor="w",
                                         text=dt_line, font=("Segoe UI", 9), fill=hdr_text)

            # header actions (Win11-like icons + hover)
            ico_y = y0 + 14
            ico_font = _pick_icon_font(12)

            edit_x, dup_x, trash_x = (x1 - 66), (x1 - 40), (x1 - 14)
            box_w, box_h = 22, 20

            # hover background (только для иконок, не для галочки)
            if getattr(self, "_hover", None) == ("edit", ti):
                self.hcanvas.create_rectangle(edit_x - box_w/2, ico_y - box_h/2, edit_x + box_w/2, ico_y + box_h/2,
                                              fill="#e9e9e9", outline="")
            if getattr(self, "_hover", None) == ("dup", ti):
                self.hcanvas.create_rectangle(dup_x - box_w/2, ico_y - box_h/2, dup_x + box_w/2, ico_y + box_h/2,
                                              fill="#e9e9e9", outline="")
            if getattr(self, "_hover", None) == ("trash", ti):
                self.hcanvas.create_rectangle(trash_x - box_w/2, ico_y - box_h/2, trash_x + box_w/2, ico_y + box_h/2,
                                              fill="#e9e9e9", outline="")

            self.hcanvas.create_text(edit_x, ico_y, text=ICON_CALENDAR, font=ico_font, fill=hdr_icon, anchor="center")
            self.hcanvas.create_text(dup_x, ico_y, text=ICON_COPY, font=ico_font, fill=hdr_icon, anchor="center")
            self.hcanvas.create_text(trash_x, ico_y, text=ICON_DELETE, font=ico_font, fill=hdr_icon, anchor="center")

            # колонка заголовков (H/qc/fs) — в шапке и фиксирована
            sh_y = y0 + self.hdr_h - top_pad
            self.hcanvas.create_text(x0 + self.w_depth / 2, sh_y, text="H, м", font=("Segoe UI", 9), fill=hdr_text)
            self.hcanvas.create_text(x0 + self.w_depth + self.w_val / 2, sh_y, text="qc", font=("Segoe UI", 9), fill=hdr_text)
            self.hcanvas.create_text(x0 + self.w_depth + self.w_val + self.w_val / 2, sh_y, text="fs", font=("Segoe UI", 9), fill=hdr_text)
            if getattr(self, "geo_kind", "K2") == "K4":
                self.hcanvas.create_text(x0 + self.w_depth + self.w_val*2 + self.w_val/2, sh_y, text="U", font=("Segoe UI", 9), fill=hdr_text)

            # --- ТАБЛИЦА (canvas) ---
            fl = self.flags.get(t.tid, TestFlags(False, set(), set(), set(), set(), set()))
            mp = self._grid_row_maps.get(ti, {})
            start_r = self._grid_start_rows.get(ti, 0)

            for r in range(max_rows):
                if grid[r] is None:
                    depth_txt = t.depth[r] if (r < len(getattr(t, "depth", []) or [])) else ""
                else:
                    depth_txt = f"{grid[r]:.2f}"

                data_i = mp.get(r, None)
                has_row = (data_i is not None) and (data_i < len(getattr(t, "qc", []) or []))
                qc_txt = str(t.qc[data_i]) if has_row else ""
                fs_txt = str(t.fs[data_i]) if has_row else ""
                incl_txt = ""
                if getattr(self, "geo_kind", "K2") == "K4":
                    incl_list = getattr(t, "incl", None)
                    if has_row and incl_list is not None and data_i < len(incl_list):
                        incl_txt = str(incl_list[data_i])

                is_blank_row = (qc_txt.strip()=="" and fs_txt.strip()=="" and (incl_txt.strip()=="" if getattr(self, "geo_kind", "K2")=="K4" else True))

                if not has_row:
                    depth_txt = ""

                # Если строка данных пустая (оба значения пустые) — скрываем глубину напротив,
                # но во время редактирования показываем глубину (чтобы было понятно, куда вводим).
                _is_editing_this = False
                try:
                    ed = getattr(self, '_editing', None)
                    if ed and len(ed) >= 3:
                        ed_ti, ed_row, _ed_field = ed[0], ed[1], ed[2]
                        if ed_ti == ti and data_i is not None and ed_row == data_i:
                            _is_editing_this = True
                except Exception:
                    _is_editing_this = False
                if has_row and is_blank_row and not _is_editing_this:
                    depth_txt = ""


                if r == start_r and has_row:
                    depth_fill = "white"   # editable cell
                else:
                    depth_fill = (GUI_DEPTH_BG if has_row else "white")

                if not depth_txt:
                    depth_fill = "white"

                def fill_for(kind: str):
                    # Сначала — специальные подсветки, которые могут относиться к "пустым" строкам (хвост).
                    if data_i in getattr(fl, 'force_tail_rows', set()) and kind in ('depth','qc','fs','incl'):
                        return (GUI_BLUE_P if getattr(self, '_algo_preview_mode', False) else GUI_BLUE)

                    # Далее — обычная логика по существующим/пустым строкам
                    if not has_row or is_blank_row:
                        return "white"
                    if fl.invalid:
                        return GUI_RED

                    if (data_i, kind) in getattr(fl, 'user_cells', set()):
                        return GUI_PURPLE
                    if (data_i, kind) in getattr(fl, 'algo_cells', set()):
                        return GUI_GREEN
                    if (data_i, kind) in fl.force_cells:
                        return (GUI_BLUE_P if getattr(self, '_algo_preview_mode', False) else GUI_BLUE)
                    if (data_i, kind) in fl.interp_cells:
                        return (GUI_ORANGE_P if getattr(self, '_algo_preview_mode', False) else GUI_ORANGE)

                    # Подсветка нулей (qc/fs) — как "некорректно"
                    # (в т.ч. для только что добавленных зондирований, где значения по умолчанию 0)
                    if kind == 'qc' and str(qc_txt).strip() == '0':
                        return GUI_RED
                    if kind == 'fs' and str(fs_txt).strip() == '0':
                        return GUI_RED

                    return "white" 

                cells = [
                    ("depth", depth_txt, depth_fill),
                    ("qc", qc_txt, fill_for("qc")),
                    ("fs", fs_txt, fill_for("fs")),
                ]
                if getattr(self, "geo_kind", "K2") == "K4":
                    cells.append(("incl", incl_txt, fill_for("incl")))

                for field, txt, fill in cells:
                    # preview highlight for context-menu deletion
                    try:
                        if getattr(self, "_rc_preview", None) == (ti, r):
                            fill = GUI_RED
                    except Exception:
                        pass
                    bx0, by0, bx1, by1 = self._cell_bbox(col, r, field)
                    self.canvas.create_rectangle(bx0, by0, bx1, by1, fill=fill, outline=GUI_GRID)
                    if field == "depth":
                        tx, anchor, color = bx1 - 4, "e", "#555"
                    else:
                        tx, anchor, color = bx1 - 4, "e", "#000"
                    self.canvas.create_text(tx, (by0 + by1) / 2, text=txt, anchor=anchor, fill=color, font=("Segoe UI", 9))

        self._update_scrollregion()
    # ---------------- hit test & editing ----------------

    def _hit_test(self, x, y):
        # Определяем, по какому холсту пришло событие (шапка или таблица)
        w = getattr(self, "_evt_widget", None) or self.canvas

        if not self.tests:
            return None

        col_w = self.w_depth + self.w_val*2 + (self.w_val if getattr(self, "geo_kind", "K2")=="K4" else 0)

        if w is getattr(self, "hcanvas", None):
            cx = self.hcanvas.canvasx(x)
            cy = self.hcanvas.canvasy(y)
            y0 = self.pad_y  # верхний отступ внутри шапки

            self._refresh_display_order()
            for col, ti in enumerate(self.display_cols):
                x0 = self.pad_x + col * (col_w + self.col_gap)
                x1 = x0 + col_w
                if x0 <= cx <= x1 and (y0 <= cy <= y0 + self.hdr_h):
                    # export checkbox (left)
                    if (x0 + 6) <= cx <= (x0 + 20) and (y0 + 8) <= cy <= (y0 + 22):
                        return ("export", ti, None, None)
                    # icons
                    if (x1 - 78) <= cx <= (x1 - 54) and y0 <= cy <= (y0 + 24):
                        return ("edit", ti, None, None)
                    if (x1 - 52) <= cx <= (x1 - 28) and y0 <= cy <= (y0 + 24):
                        return ("dup", ti, None, None)
                    if (x1 - 26) <= cx <= (x1 - 2) and y0 <= cy <= (y0 + 24):
                        return ("trash", ti, None, None)
                    return ("header", ti, None, None)
            return None

        # --- таблица (числа) ---
        cx = self.canvas.canvasx(x)
        cy = self.canvas.canvasy(y)

        # row/col by coordinates
        if cy < 0:
            return None

        row = int(cy // self.row_h)
        if row < 0:
            return None

        self._refresh_display_order()
        for col, ti in enumerate(self.display_cols):
            x0 = self.pad_x + col * (col_w + self.col_gap)
            x1 = x0 + col_w
            if x0 <= cx <= x1:
                # which field
                # depth/qc/fs split
                relx = cx - x0
                if relx < self.w_depth:
                    field = "depth"
                elif relx < (self.w_depth + self.w_val):
                    field = "qc"
                else:
                    field = "fs"
                return ("cell", ti, row, field)

        return None


    def _on_double_click(self, event):
        self._evt_widget = event.widget
        hit = self._hit_test(event.x, event.y)
        if not hit:
            # клик вне ячеек/шапки → закрываем активное редактирование
            self._end_edit(commit=True)
            self._hide_canvas_tip()
            return
        kind, ti, row, field = hit
        if kind == "header":
            return

        mp = (getattr(self, "_grid_row_maps", {}) or {}).get(ti, {})
        start_r = (getattr(self, "_grid_start_rows", {}) or {}).get(ti, 0)

        if field == "depth":
            if row == start_r:
                self._begin_edit_depth0(ti, display_row=row)
            return

        data_row = mp.get(row, None)
        if data_row is None:
            # Клик по пустой строке ниже конца зондирования: добавляем строку и даём ввод
            if field in ("qc", "fs"):
                new_idx = len(self.tests[ti].qc)
                self._append_row(ti)
                self._begin_edit(ti, new_idx, field, display_row=row)
            return
        self._begin_edit(ti, data_row, field, display_row=row)

    def _on_arrow_key(self, event):
        # Стрелки: открываем соседнюю ячейку и выделяем всё значение
        if not getattr(self, "_editing", None):
            return "break"
        ti, row, field = self._editing[0], self._editing[1], self._editing[2]
        if field not in ("qc", "fs"):
            return "break"
        dx = 0; dy = 0
        if event.keysym == "Up":
            dy = -1
        elif event.keysym == "Down":
            dy = 1
        elif event.keysym == "Left":
            dx = -1
        elif event.keysym == "Right":
            dx = 1
        # commit current edit, then move
        self._end_edit(commit=True)
        t = self.tests[ti]
        new_field = field
        if dx != 0:
            # без «зацикливания»: вправо из fs и влево из qc упираемся в стенку
            if dx > 0 and field == "qc":
                new_field = "fs"
            elif dx < 0 and field == "fs":
                new_field = "qc"
            else:
                new_field = field
        new_row = row + dy
        if new_row < 0:
            new_row = 0
        if new_row >= len(t.qc):
            # упираемся в низ/верх (добавление строк только Enter или кликом)
            new_row = len(t.qc) - 1 if t.qc else 0
        # open editor at new cell (display row find)
        self._begin_edit(ti, new_row, new_field, display_row=None)
        return "break"

    def _on_right_click(self, event):
        self._evt_widget = event.widget
        # Правый клик где угодно закрывает активную ячейку
        try:
            self._end_edit(commit=True)
        except Exception:
            pass

        # Контекстное меню: удалить выше / удалить ниже (включая выбранную строку)
        try:
            hit = self._hit_test(event.x, event.y)
        except Exception:
            hit = None
        if not hit:
            return
        kind, ti, row, field = hit
        if kind != "cell" or ti is None or row is None:
            return

        # Не показываем меню на пустых ячейках
        try:
            t = self.tests[ti]
            mp = (getattr(self, "_grid_row_maps", {}) or {}).get(ti, {}) or {}
            data_i = mp.get(row, None)
            has_row = (data_i is not None) and (data_i < len(getattr(t, "qc", []) or []))
            if not has_row:
                return
            vq = t.qc[data_i] if has_row else None
            vf = t.fs[data_i] if has_row else None
            if vq is None and vf is None:
                return
        except Exception:
            return

        # Подсветка строки красным при вызове меню
        self._rc_preview = (ti, row)
        try:
            self._redraw()
        except Exception:
            pass

        self._ctx_target = (ti, row)
        try:
            self._ctx_menu.tk_popup(event.x_root, event.y_root)
        finally:
            try:
                self._ctx_menu.grab_release()
            except Exception:
                pass
            self._rc_preview = None
            try:
                self._redraw()
            except Exception:
                pass

    def _ctx_delete_above(self):
        """Удалить строки выше выбранной по глубине отображения."""
        if not getattr(self, '_ctx_target', None):
            return
        ti, row = self._ctx_target
        self._delete_by_display_row(ti, row, mode='above')

    def _ctx_delete_below(self):
        """Удалить строки ниже выбранной (вкл.) по глубине отображения."""
        if not getattr(self, '_ctx_target', None):
            return
        ti, row = self._ctx_target
        self._delete_by_display_row(ti, row, mode='below')

    def _ctx_delete_row(self):
        """Удалить одну строку по глубине отображения."""
        if not getattr(self, '_ctx_target', None):
            return
        ti, row = self._ctx_target
        self._delete_by_display_row(ti, row, mode='row')

    def _delete_by_display_row(self, ti, display_row: int, mode: str):
        if ti < 0 or ti >= len(self.tests):
            return
        t = self.tests[ti]

        def pdepth(v):
            try:
                s = str(v).strip().replace(',', '.')
                if s == '':
                    return None
                return float(s)
            except Exception:
                return None

        # глубина по сетке, если есть
        target_depth = None
        if getattr(self, '_grid', None) and 0 <= display_row < len(self._grid):
            target_depth = self._grid[display_row]
        else:
            if 0 <= display_row < len(getattr(t, 'depth', []) or []):
                target_depth = pdepth(t.depth[display_row])

        if target_depth is None:
            return

        n = max(len(getattr(t, 'depth', []) or []), len(getattr(t, 'qc', []) or []), len(getattr(t, 'fs', []) or []))
        while len(t.depth) < n: t.depth.append('')
        while len(t.qc) < n: t.qc.append('')
        while len(t.fs) < n: t.fs.append('')

        # список (depth, idx)
        pairs = []
        for i in range(n):
            d = pdepth(t.depth[i])
            if d is None:
                continue
            pairs.append((d, i))
        if not pairs:
            return

        nearest_d, nearest_i = min(pairs, key=lambda di: abs(di[0] - target_depth))

        if mode == 'row':
            r0 = r1 = nearest_i
        elif mode == 'above':
            inds = [i for (d, i) in pairs if d <= nearest_d + 1e-9]
            if not inds:
                return
            r0, r1 = min(inds), max(inds)
        else:  # below
            inds = [i for (d, i) in pairs if d >= nearest_d - 1e-9]
            if not inds:
                return
            r0, r1 = min(inds), max(inds)

        self._delete_range_indices(ti, r0, r1)

    def _delete_range_indices(self, ti, r0, r1):
        if ti < 0 or ti >= len(self.tests):
            return
        t = self.tests[ti]
        n = max(len(t.depth), len(t.qc), len(t.fs))
        if n <= 0:
            return
        r0 = max(0, int(r0))
        r1 = min(n - 1, int(r1))
        if r1 < r0:
            return
        while len(t.depth) < n: t.depth.append('')
        while len(t.qc) < n: t.qc.append('')
        while len(t.fs) < n: t.fs.append('')


        self._push_undo()

        # Сдвиг флагов подсветки (жёлтый/синий/фиолетовый) при удалении строк,
        # чтобы подсветка не «съезжала» относительно данных.
        k = (r1 - r0 + 1)
        try:
            fl = self.flags.get(t.tid)
            if fl:
                def _shift_cells(cells: set[tuple[int, str]]) -> set[tuple[int, str]]:
                    out = set()
                    for (r, kind) in (cells or set()):
                        try:
                            rr = int(r)
                        except Exception:
                            continue
                        if rr < r0:
                            out.add((rr, kind))
                        elif rr > r1:
                            out.add((rr - k, kind))
                        # rr in [r0..r1] -> удалено
                    return out

                fl.interp_cells = _shift_cells(getattr(fl, "interp_cells", set()))
                fl.force_cells  = _shift_cells(getattr(fl, "force_cells", set()))
                fl.user_cells   = _shift_cells(getattr(fl, "user_cells", set()))
                self.flags[t.tid] = fl
        except Exception:
            pass

        del t.depth[r0:r1+1]
        del t.qc[r0:r1+1]
        del t.fs[r0:r1+1]

        try:
            self._build_grid()
        except Exception:
            pass
        self._redraw()
        # если опыт не помещается на экран — прокручиваем по X так, чтобы он попал в видимую область
        try:
            self._ensure_cell_visible(insert_at, 0, 'depth', pad=12)
        except Exception:
            try:
                self.canvas.xview_moveto(1.0)
            except Exception:
                pass

        try:
            self.status.set(f"Удалено строк: {r1 - r0 + 1} (опыт {ti+1})")
        except Exception:
            pass



            
    def _edit_header(self, ti: int):
        t = self.tests[ti]
        win = tk.Toplevel(self)
        win.title("Параметры зондирования")
        win.resizable(False, False)
        win.transient(self)
        win.grab_set()
        # ВАЖНО: центрирование делаем ПОСЛЕ построения виджетов.
        # Иначе на Windows (особенно при масштабировании 125–175%) окно
        # центрируется по «пустому» reqsize и может оказаться меньше, чем нужно.

        


        style = ttk.Style(win)
        try:
            style.configure("Hdr.TButton", padding=(8, 1))
        except Exception:
            pass

        PADX = 12
        PADY = 6

        # ---- № ----
        ttk.Label(win, text="№ зондирования").grid(row=0, column=0, sticky="w", padx=PADX, pady=(PADY, 2))
        tid_var = tk.StringVar(value=str(t.tid))
        tid_entry = ttk.Entry(
            win,
            textvariable=tid_var,
            width=10,
            validate="key",
            validatecommand=(win.register(_validate_tid_key), "%P"),
        )
        tid_entry.grid(row=0, column=1, sticky="w", padx=(0, PADX), pady=(PADY, 2), columnspan=2)

        # ---- Date + Time ----
        ttk.Label(win, text="Дата").grid(row=1, column=0, sticky="w", padx=PADX, pady=2)

        parsed = _try_parse_dt(t.dt or "")
        if parsed is None:
            try:
                parsed = _dt.datetime.strptime((t.dt or "").strip(), "%d.%m.%Y")
            except Exception:
                parsed = None

        d0 = (parsed.date() if parsed else _dt.date.today())
        hh0 = (parsed.hour if parsed else 0)
        mm0 = (parsed.minute if parsed else 0)

        date_var = tk.StringVar(value=_format_date_ru(d0))
        date_entry = ttk.Entry(win, textvariable=date_var, width=12, state="readonly")
        date_entry.grid(row=1, column=1, sticky="w", padx=(0, 6), pady=2)

        def pick_date():
            try:
                cur = _dt.datetime.strptime(date_var.get().strip(), "%d.%m.%Y").date()
            except Exception:
                cur = _dt.date.today()
            dlg = CalendarDialog(win, initial=cur)
            self._place_calendar_near_header(dlg, ti)
            win.wait_window(dlg)
            if dlg.selected:
                date_var.set(_format_date_ru(dlg.selected))

        cal_btn = ttk.Button(win, text="📅", style="Hdr.TButton", command=pick_date)
        cal_btn.grid(row=1, column=2, sticky="w", padx=(0, PADX), pady=2)

        ttk.Label(win, text="Время").grid(row=2, column=0, sticky="w", padx=PADX, pady=2)

        time_frame = ttk.Frame(win)
        time_frame.grid(row=2, column=1, columnspan=2, sticky="w", padx=(0, PADX), pady=2)

        hh_var = tk.StringVar(value=f"{hh0:02d}")
        mm_var = tk.StringVar(value=f"{mm0:02d}")

        hh_entry = ttk.Entry(
            time_frame,
            textvariable=hh_var,
            width=3,
            justify="center",
            validate="key",
            validatecommand=(win.register(_validate_hh_key), "%P"),
        )
        hh_entry.pack(side="left")
        ttk.Label(time_frame, text=":").pack(side="left", padx=2)
        mm_entry = ttk.Entry(
            time_frame,
            textvariable=mm_var,
            width=3,
            justify="center",
            validate="key",
            validatecommand=(win.register(_validate_mm_key), "%P"),
        )
        mm_entry.pack(side="left")


        def apply():
            new_tid_txt = tid_var.get().strip()
            if not new_tid_txt:
                messagebox.showwarning("Внимание", "Номер зондирования не может быть пустым.", parent=win)
                return
            try:
                new_tid = int(new_tid_txt)
            except Exception:
                messagebox.showwarning("Внимание", "Номер зондирования должен быть числом.", parent=win)
                return
            if not (1 <= new_tid <= 999):
                messagebox.showwarning("Внимание", "Номер зондирования: от 1 до 999.", parent=win)
                return
            if any((i != ti and int(tt.tid) == new_tid) for i, tt in enumerate(self.tests)):
                messagebox.showwarning("Внимание", f"Номер {new_tid} уже существует.", parent=win)
                return

            try:
                d = _dt.datetime.strptime(date_var.get().strip(), "%d.%m.%Y").date()
            except Exception:
                messagebox.showwarning("Внимание", "Некорректная дата.", parent=win)
                return

            hh_txt = (hh_var.get() or "").strip() or "0"
            mm_txt = (mm_var.get() or "").strip() or "0"
            try:
                hh = int(hh_txt); mm = int(mm_txt)
            except Exception:
                messagebox.showwarning("Внимание", "Некорректное время.", parent=win)
                return
            if not (0 <= hh <= 23 and 0 <= mm <= 59):
                messagebox.showwarning("Внимание", "Время должно быть в диапазоне 00:00–23:59.", parent=win)
                return

            dt_text = f"{d.day:02d}.{d.month:02d}.{d.year:04d} {hh:02d}:{mm:02d}"

            self._push_undo()
            old_tid = t.tid
            t.tid = new_tid
            dt_obj = _try_parse_dt(dt_text)
            t.dt = dt_obj.strftime("%Y-%m-%d %H:%M:%S") if dt_obj else dt_text

            if old_tid in self.flags:
                self.flags[new_tid] = self.flags.pop(old_tid)
            else:
                self.flags[new_tid] = TestFlags(False, set(), set(), set(), set(), set())

            self._redraw()
            win.destroy()

        # Enter = сохранить
        for _w in (tid_entry, hh_entry, mm_entry):
            try:
                _w.bind('<Return>', lambda _e: apply())
            except Exception:
                pass
        try:
            win.bind('<Return>', lambda _e: apply())
        except Exception:
            pass

        
        # ---- Buttons (centered) ----
        btns = ttk.Frame(win)
        btns.grid(row=3, column=0, columnspan=3, sticky="ew", padx=12, pady=(8, 12))
        btns.columnconfigure(0, weight=1)
        btns.columnconfigure(2, weight=1)

        inner_btns = ttk.Frame(btns)
        inner_btns.grid(row=0, column=1)

        ttk.Button(inner_btns, text="Сохранить", style="Hdr.TButton", command=apply).pack(side="left", padx=(0, 8))
        ttk.Button(inner_btns, text="Отмена", style="Hdr.TButton", command=win.destroy).pack(side="left")

        # Enter = save (в том числе NumPad Enter)
        win.bind("<Return>", lambda _e: apply())
        win.bind("<KP_Enter>", lambda _e: apply())

        # финальная геометрия + центрирование (после того, как Tk посчитает reqsize)
        try:
            win.update_idletasks()
        except Exception:
            pass
        try:
            self._center_child(win)
        except Exception:
            pass



    def _begin_edit(self, ti: int, row: int, field: str, display_row: int | None = None):
        """Edit qc/fs cell. row is data index, display_row is grid index."""
        self._end_edit(commit=True)
        t = self.tests[ti]
        # Не даём вводить значения "после конца" зондирования.
        if row < 0 or row >= len(t.qc):
            return

        if display_row is None:
            try:
                mp = (getattr(self, "_grid_row_maps", {}) or {}).get(ti, {})
                for gr, di in mp.items():
                    if di == row:
                        display_row = gr
                        break
            except Exception:
                pass
        if display_row is None:
            display_row = row

        self._refresh_display_order()
        col = self.display_cols.index(ti)

        # Автопрокрутка (стрелки/Enter): держим ячейку в видимой зоне
        self._ensure_cell_visible(col, display_row, field)

        bx0, by0, bx1, by1 = self._cell_bbox(col, display_row, field)
        vx0 = bx0 - self.canvas.canvasx(0)
        vy0 = by0 - self.canvas.canvasy(0)

        current = t.qc[row] if field == "qc" else t.fs[row]
        e = tk.Entry(self.canvas, validate="key", validatecommand=(self.register(_validate_int_0_300_key), "%P"))
        e.insert(0, current)
        try:
            e.configure(bg="white")
        except Exception:
            pass
        e.select_range(0, tk.END)
        e.place(x=vx0 + 1, y=vy0 + 1, width=(bx1 - bx0) - 2, height=(by1 - by0) - 2)
        e.focus_set()

        def commit_and_next():
            self._end_edit(commit=True)

            # Enter: вниз. Если дошли до конца — добавляем новую строку и продолжаем ввод.
            next_row = row + 1
            if next_row < len(t.qc):
                self._begin_edit(ti, next_row, field, (display_row or row) + 1)
            else:
                # добавляем новую строку в хвост и начинаем редактирование
                self._append_row(ti)
                try:
                    self._begin_edit(ti, next_row, field, (display_row or row) + 1)
                except Exception:
                    pass

        e.bind("<Return>", lambda _ev: commit_and_next())
        for _k in ("<Up>","<Down>","<Left>","<Right>"):
            e.bind(_k, self._on_arrow_key)
        e.bind("<Escape>", lambda _ev: self._end_edit(commit=False))
        e.bind("<FocusOut>", lambda _ev: self._end_edit(commit=True))

        self._editing = (ti, row, field, e, display_row)

    def _begin_edit_depth0(self, ti: int, display_row: int = 0):
        """Редактирование первой глубины (depth[0]) с автопересчётом всей колонки depth."""
        self._end_edit(commit=True)
        t = self.tests[ti]
        if not getattr(t, "depth", None):
            return

        self._refresh_display_order()
        col = self.display_cols.index(ti)

        # Автопрокрутка (стрелки/Enter)
        self._ensure_cell_visible(col, display_row, "depth")

        bx0, by0, bx1, by1 = self._cell_bbox(col, display_row, "depth")
        vx0 = bx0 - self.canvas.canvasx(0)
        vy0 = by0 - self.canvas.canvasy(0)

        current = str(t.depth[0]).strip()
        e = tk.Entry(self.canvas, validate="key", validatecommand=(self.register(_validate_depth_0_4_key), "%P"))
        e.insert(0, current)
        e.select_range(0, tk.END)
        e.place(x=vx0 + 1, y=vy0 + 1, width=(bx1 - bx0) - 2, height=(by1 - by0) - 2)
        e.focus_set()

        def commit():
            self._push_undo()
            self._end_edit_depth0(ti, e, commit=True)

        e.bind("<Return>", lambda _ev: commit())
        e.bind("<Escape>", lambda _ev: self._end_edit_depth0(ti, e, commit=False))
        e.bind("<FocusOut>", lambda _ev: self._end_edit_depth0(ti, e, commit=True))

        self._editing = (ti, 0, "depth", e, display_row)

    def _end_edit_depth0(self, ti: int, e, commit: bool):
        try:
            val = e.get().strip()
        except Exception:
            val = ""
        try:
            e.place_forget()
            e.destroy()
        except Exception:
            pass
        self._editing = None
        if not commit:
            self._redraw()
            return

        t = self.tests[ti]
        old0 = _parse_depth_float(t.depth[0]) if getattr(t, "depth", None) else None
        new0 = _parse_depth_float(val)
        # ограничения: 0..4 м и шаг по 5 см (0.05 м)
        if new0 is not None:
            if new0 < 0.0 or new0 > 4.0:
                messagebox.showerror("Ошибка", "Начальная глубина должна быть в диапазоне 0…4 м.")
                self._redraw()
                return
            # Ограничение по сетке глубин:
            # - при шаге 10 см (0.10 м) разрешаем только кратность 0.10
            # - иначе (шаг 5 см) кратность 0.05
            step_m = float(self.step_m or 0.05)
            step_cm = int(round(step_m * 100.0))
            if step_cm <= 0:
                step_cm = 5
            cm = int(round(new0 * 100.0))
            if cm % step_cm != 0:
                if step_cm == 10:
                    messagebox.showerror("Ошибка", "При шаге 10 см глубина должна быть кратна 0.10 м: …0.00, …0.10, …0.20 и т.д.")
                else:
                    messagebox.showerror("Ошибка", "Глубина должна быть кратна 0.05 м (5 см): …0.00, …0.05, …0.10 и т.д.")
                self._redraw()
                return
        if new0 is None:
            self._redraw()
            return
        if old0 is None:
            old0 = new0

        delta = new0 - old0
        if abs(delta) < 1e-9:
            t.depth[0] = f"{new0:.2f}"
            self._redraw()
            return

        new_depth = []
        for ds in (getattr(t, "depth", []) or []):
            d = _parse_depth_float(ds)
            if d is None:
                new_depth.append(ds)
            else:
                new_depth.append(f"{(d + delta):.2f}")
        t.depth = new_depth

        # Не сбрасываем подсветку/флаги при сдвиге глубины: qc/fs не менялись
        # (иначе пропадает фиолетовая отметка ручных правок и др. подсветки)
        self._redraw()
    def _end_edit(self, commit: bool):
        if not self._editing:
            return
        if len(self._editing) == 4:
            ti, row, field, e = self._editing
        else:
            ti, row, field, e, _disp = self._editing
        try:
            val = e.get().strip()
        except Exception:
            val = ""
        try:
            e.place_forget()
            e.destroy()
        except Exception:
            pass
        self._editing = None

        if field == "depth":
            self._redraw()
            return

        if commit and self.tests:
            t = self.tests[ti]
            if row < len(t.qc):
                # keep previous coloring info, but mark this cell as manually edited (purple)
                fl = self.flags.get(t.tid) or TestFlags(False, set(), set(), set(), set(), set())
                old = t.qc[row] if field == 'qc' else t.fs[row]
                newv = _sanitize_int_0_300(val)
                # Undo: фиксируем снимок ДО изменения данных/раскраски
                if commit:
                    try:
                        # если реально меняем значение или удаляем строку
                        if (str(old).strip() != str(newv).strip()):
                            self._push_undo()
                    except Exception:
                        self._push_undo()
                # Запрет: в середине зондирования нельзя ставить 0 или оставлять пусто.
                # Пустое значение разрешено только на краях (первая/последняя строка) — тогда удаляем строку целиком.
                last_filled_before = self._last_filled_row(t)

                # edge-delete when clearing first/last filled row
                if newv.strip() == "":
                    if row == 0 or row == last_filled_before:
                        # удалить строку данных и глубину
                        fl = self.flags.get(t.tid) or TestFlags(False, set(), set(), set(), set(), set())
                        self._delete_data_row_in_test(t, fl, row)
                        self.flags[t.tid] = fl
                        self._redraw()
                        return
                    else:
                        self.status.config(text="Нельзя оставлять пустые значения в середине зондирования.")
                        self._redraw()
                        return

                if newv.strip() == "0" and (0 < row < last_filled_before):
                    self.status.config(text="Нельзя записывать 0 в середине зондирования.")
                    self._redraw()
                    return
                if field == 'qc':
                    t.qc[row] = newv
                else:
                    t.fs[row] = newv
                try:
                    if str(old).strip() != str(newv).strip():
                        fl.user_cells.add((row, field))
                        try:
                            fl.algo_cells.discard((row, field))
                        except Exception:
                            pass
                except Exception:
                    pass
                fl.invalid = False
                self.flags[t.tid] = fl
            self._redraw()

    def _last_filled_row(self, t: TestData) -> int:
        """Последняя строка с данными (qc или fs не пустые)."""
        try:
            n = min(len(getattr(t, 'qc', []) or []), len(getattr(t, 'fs', []) or []))
        except Exception:
            return -1
        for i in range(n - 1, -1, -1):
            if str(t.qc[i]).strip() != "" or str(t.fs[i]).strip() != "":
                return i
        return -1


    def _delete_data_row_in_test(self, t: TestData, fl: TestFlags, row: int):
        """Удаляет строку row из depth/qc/fs и корректирует раскраски (interp/force/user)."""
        try:
            if 0 <= row < len(t.depth):
                t.depth.pop(row)
        except Exception:
            pass
        try:
            if 0 <= row < len(t.qc):
                t.qc.pop(row)
        except Exception:
            pass
        try:
            if 0 <= row < len(t.fs):
                t.fs.pop(row)
        except Exception:
            pass

        def shift_cells(cells: set[tuple[int, str]]):
            out = set()
            for r, kind in list(cells or set()):
                if r == row:
                    continue
                if r > row:
                    out.add((r - 1, kind))
                else:
                    out.add((r, kind))
            return out

        try:
            fl.interp_cells = shift_cells(getattr(fl, "interp_cells", set()))
        except Exception:
            fl.interp_cells = set()
        try:
            fl.force_cells = shift_cells(getattr(fl, "force_cells", set()))
        except Exception:
            fl.force_cells = set()
        try:
            fl.user_cells = shift_cells(getattr(fl, "user_cells", set()))
        except Exception:
            fl.user_cells = set()


    def _append_row(self, ti: int):
        if self.depth_start is None or self.step_m is None:
            return
        t = self.tests[ti]
        last_d = None
        for j in range(len(t.depth)-1, -1, -1):
            d = _parse_depth_float(t.depth[j])
            if d is not None:
                last_d = d
                break
        if last_d is None:
            idx = len(t.depth)
            tid = int(getattr(t, "tid", 0) or 0)
            d0 = float(self.depth0_by_tid.get(tid, float(self.depth_start or 0.0)))
            step = float(self.step_m or 0.05)
            t.depth.append(f"{(d0 + idx * step):g}")
        else:
            t.depth.append(f"{(last_d + self.step_m):g}")
        t.qc.append("")
        t.fs.append("")
        # не сбрасываем подсветку при добавлении строки; только гарантируем наличие флагов
        if t.tid not in self.flags:
            self.flags[t.tid] = TestFlags(False, set(), set(), set(), set(), set())
        self._redraw()

    # ---------------- scrolling ----------------
    def _on_mousewheel(self, event):
        # скролл закрывает активную ячейку
        self._end_edit(commit=True)
        delta = int(-1 * (event.delta / 120)) if event.delta else 0
        if delta != 0:
            self.canvas.yview_scroll(delta, "units")
        return "break"

    def _on_mousewheel_linux(self, direction):
        # скролл закрывает активную ячейку
        self._end_edit(commit=True)
        self.canvas.yview_scroll(direction, "units")
        return "break"

    def _on_mousewheel_x(self, event):
        """Горизонтальная прокрутка колесом шагом 1 колонка (когда курсор над шапкой или горизонтальным скроллом)."""
        self._end_edit(commit=True)
        try:
            delta = int(-1 * (event.delta / 120)) if getattr(event, "delta", 0) else 0
        except Exception:
            delta = 0
        if not delta:
            return "break"
        self._scroll_x_by_one_column(delta)
        return "break"

    def _on_mousewheel_linux_x(self, direction):
        """Linux: Button-4/5 для горизонтальной прокрутки шагом 1 колонка."""
        self._end_edit(commit=True)
        try:
            direction = int(direction)
        except Exception:
            direction = 0
        if not direction:
            return "break"
        self._scroll_x_by_one_column(direction)
        return "break"

    def _scroll_x_by_one_column(self, direction: int):
        """Сдвиг по X на одну колонку зондирования (шаг = ширина блока шапки)."""
        try:
            direction = 1 if direction > 0 else -1
        except Exception:
            direction = 1
        # ширина одной колонки (Depth + qc + fs) + зазор между колонками
        col_block = int(self.w_depth + self.w_val + self.w_val + self.col_gap)
        try:
            w = float(getattr(self, "_scroll_w", 0) or 0)
        except Exception:
            w = 0.0
        if w <= 1:
            # на всякий случай обновим ширину
            try:
                w = float(self._content_size()[0])
                self._scroll_w = w
            except Exception:
                w = 1.0

        try:
            x0_frac = float(self.canvas.xview()[0])
        except Exception:
            x0_frac = 0.0
        x0_px = x0_frac * w
        target_px = max(0.0, x0_px + direction * col_block)

        # ограничим по правому краю
        try:
            view_w = float(self.canvas.winfo_width())
        except Exception:
            view_w = 0.0

        # Если последний (правый) опыт уже появился в видимой области,
        # блокируем дальнейший шаг вправо. Это убирает «последний лишний скролл»,
        # который и вызывает смещение шапки на самом правом краю.
        try:
            n_tests = len(self.tests)
        except Exception:
            n_tests = 0
        try:
            col_w = float(self.w_depth + self.w_val*2 + (self.w_val if getattr(self, 'geo_kind', 'K2')=='K4' else 0))
        except Exception:
            col_w = 0.0
        try:
            gap = float(self.col_gap)
        except Exception:
            gap = 0.0
        try:
            pad = float(self.pad_x)
        except Exception:
            pad = 0.0
        # левая/правая границы последней колонки (без учета правого паддинга)
        last_left_px = pad + (col_w + gap) * max(0, n_tests - 1)
        last_right_px = last_left_px + col_w
        view_right_px = x0_px + max(1.0, view_w)
        # Блокируем шаг вправо ТОЛЬКО когда последняя колонка ВИДНА ПОЛНОСТЬЮ.
        if direction > 0 and view_right_px >= (last_right_px - 0.5):
            return

        max_px = max(0.0, w - max(1.0, view_w))
        if target_px > max_px:
            target_px = max_px

        frac = 0.0 if w <= 1 else (target_px / w)
        try:
            self._xview_proxy("moveto", frac)
        except Exception:
            pass


    # ---------------- fix algorithm (from v2.3.1) ----------------
    def _choose_tail_k(self, last_val: int) -> int:
        d = abs(250 - last_val)
        if d <= 10:
            return 1
        if d <= 35:
            return 2
        return 3

    def fix_by_algorithm(self):
        if not self.tests:
            return

        self._push_undo()
        random.seed(42)

        for t in self.tests:
            tid = t.tid
            prev_flags = self.flags.get(tid) or TestFlags(False, set(), set(), set(), set(), set())
            _prev_user_cells = set(getattr(prev_flags, 'user_cells', set()) or set())
            # Снимок значений до автокорректировки (для зелёной подсветки)
            _orig_qc = list(getattr(t, 'qc', []) or [])
            _orig_fs = list(getattr(t, 'fs', []) or [])
            _orig_depth = list(getattr(t, 'depth', []) or [])
            algo_cells: set[tuple[int, str]] = set()
            n = len(t.qc)
            if n == 0:
                # сохраняем пользовательские правки (на случай пустого зондирования)
                self.flags[tid] = TestFlags(False, set(), set(), _prev_user_cells, algo_cells, set())
                continue

            qc = [(_parse_cell_int(v) or 0) for v in t.qc]
            fs = [(_parse_cell_int(v) or 0) for v in t.fs]

            invalid = (_max_zero_run(qc) > 5) or (_max_zero_run(fs) > 5)
            interp_cells: set[tuple[int, str]] = set(getattr(prev_flags, 'interp_cells', set()) or set())
            force_cells: set[tuple[int, str]] = set(getattr(prev_flags, 'force_cells', set()) or set())

            if invalid:
                self.flags[tid] = TestFlags(True, interp_cells, force_cells, _prev_user_cells, algo_cells, set())
                continue

            def interp_in_place(arr: list[int], kind: str):
                i = 0
                while i < n:
                    if arr[i] != 0:
                        i += 1
                        continue
                    j = i
                    while j < n and arr[j] == 0:
                        j += 1
                    gap_len = j - i
                    if gap_len <= 5:
                        left = i - 1
                        right = j
                        if left >= 0 and right < n and arr[left] != 0 and arr[right] != 0:
                            a = arr[left]; b = arr[right]
                            for k in range(gap_len):
                                tt = (k + 1) / (gap_len + 1)
                                if (i + k, kind) not in _prev_user_cells and (i + k, kind) not in interp_cells and (i + k, kind) not in force_cells:
                                    arr[i + k] = _interp_with_noise(a, b, tt)
                                interp_cells.add((i + k, kind))
                        elif left >= 0 and arr[left] != 0:
                            a = arr[left]
                            for k in range(gap_len):
                                arr[i + k] = _noise_around(a)
                                interp_cells.add((i + k, kind))
                        elif right < n and arr[right] != 0:
                            b = arr[right]
                            for k in range(gap_len):
                                arr[i + k] = _noise_around(b)
                                interp_cells.add((i + k, kind))
                    i = j

            interp_in_place(qc, "qc")
            interp_in_place(fs, "fs")

            # ensure no zeros
            for arr, kind in ((qc, "qc"), (fs, "fs")):
                for i in range(n):
                    if arr[i] != 0:
                        continue
                    left = i - 1
                    while left >= 0 and arr[left] == 0:
                        left -= 1
                    right = i + 1
                    while right < n and arr[right] == 0:
                        right += 1
                    if left >= 0 and right < n:
                        arr[i] = _interp_with_noise(arr[left], arr[right], 0.5)
                    elif left >= 0:
                        arr[i] = _noise_around(arr[left])
                    elif right < n:
                        arr[i] = _noise_around(arr[right])
                    else:
                        arr[i] = 1
                    interp_cells.add((i, kind))

            # --- finish to 250 (choose closer) ---
            # Важно: если в опыте уже был «отказ» (значения доходили/превышали 250),
            # НИЧЕГО не дописываем вниз. Исправляем только нули/пробелы.
            refusal = False
            try:
                mx = max((qc + fs) or [0])
                refusal = (mx >= 250)
            except Exception:
                refusal = False

            if not refusal:
                # добавляем 1–3 строки ниже, не перетирая хвост
                last_filled = -1
                for rr in range(n - 1, -1, -1):
                    if qc[rr] != 0 or fs[rr] != 0:
                        last_filled = rr
                        break
                if last_filled < 0:
                    last_filled = n - 1

                target_kind = "qc" if abs(250 - qc[last_filled]) <= abs(250 - fs[last_filled]) else "fs"
                main_arr = qc if target_kind == "qc" else fs
                other_arr = fs if target_kind == "qc" else qc

                last_main = max(1, main_arr[last_filled])
                last_other = max(1, other_arr[last_filled])

                add_cnt = max(1, min(3, self._choose_tail_k(last_main)))

                step = self.step_m if self.step_m is not None else 0.05
                last_depth = None
                if t.depth and last_filled < len(t.depth):
                    last_depth = _parse_depth_float(t.depth[last_filled])
                if last_depth is None:
                    base = self.depth_start if self.depth_start is not None else 0.0
                    last_depth = base + step * last_filled

                for k_i in range(1, add_cnt + 1):
                    tt = k_i / add_cnt
                    new_main = _interp_with_noise(last_main, 250, tt)
                    new_main = max(last_main, min(250, new_main))
                    if k_i == add_cnt:
                        new_main = 250

                    # второй показатель тоже слегка растёт (примерно 15–25% от прироста main)
                    inc_main = max(0, new_main - last_main)
                    inc_other = max(1, int(round(inc_main * 0.22))) if inc_main > 0 else 1
                    new_other = min(250, max(last_other, _noise_around(last_other + inc_other)))

                    # добавляем строку
                    qc.append(0); fs.append(0)
                    t.qc.append(""); t.fs.append(""); t.depth.append("")
                    n += 1

                    if target_kind == "qc":
                        qc[-1] = int(new_main); fs[-1] = int(new_other)
                        force_cells.add((n - 1, "qc")); force_cells.add((n - 1, "fs"))
                    else:
                        fs[-1] = int(new_main); qc[-1] = int(new_other)
                        force_cells.add((n - 1, "fs")); force_cells.add((n - 1, "qc"))

                    dd = last_depth + step * k_i
                    t.depth[-1] = f"{dd:.2f}"
# write back with markers (for user visibility)
            for i in range(n):
                qv = max(1, qc[i])
                fv = max(1, fs[i])
                t.qc[i] = str(qv)
                t.fs[i] = str(fv)


            # --- зелёная подсветка: что реально поменялось алгоритмом ---
            try:
                for i2 in range(n):
                    # новые значения (после записи)
                    new_q = str(t.qc[i2]).strip() if i2 < len(t.qc) else ""
                    new_f = str(t.fs[i2]).strip() if i2 < len(t.fs) else ""
                    # исходные (до корректировки)
                    old_q = str(_orig_qc[i2]).strip() if i2 < len(_orig_qc) else ""
                    old_f = str(_orig_fs[i2]).strip() if i2 < len(_orig_fs) else ""
                    if (i2, "qc") not in _prev_user_cells and new_q != old_q:
                        algo_cells.add((i2, "qc"))
                    if (i2, "fs") not in _prev_user_cells and new_f != old_f:
                        algo_cells.add((i2, "fs"))
                # если хвост добавлен — помечаем его целиком (qc/fs)
                if len(t.qc) > len(_orig_qc):
                    for i2 in range(len(_orig_qc), len(t.qc)):
                        if (i2, "qc") not in _prev_user_cells:
                            algo_cells.add((i2, "qc"))
                        if (i2, "fs") not in _prev_user_cells:
                            algo_cells.add((i2, "fs"))
            except Exception:
                pass

            self.flags[tid] = TestFlags(False, interp_cells, force_cells, _prev_user_cells, algo_cells, set())

        self._end_edit(commit=True)
        self._redraw()

        # После успешной корректировки — синяя строка в подвале
        try:
            self.footer_cmd.config(foreground="#0b5ed7")
            self.footer_cmd.config(text="Статическое зондирование откорректировано.")
        except Exception:
            pass


    # ---------------- export ----------------


    def _read_calc_params(self):
        try:
            scale_div = int(float(self.scale_var.get().replace(",", ".")))
            fmax_cone_kn = float(self.fcone_var.get().replace(",", "."))
            fmax_sleeve_kn = float(self.fsleeve_var.get().replace(",", "."))
            area_cone_cm2 = float(self.acon_var.get().replace(",", "."))
            area_sleeve_cm2 = float(self.asl_var.get().replace(",", "."))
            if scale_div <= 0:
                raise ValueError("Шкала делений должна быть > 0")
            if fmax_cone_kn <= 0 or fmax_sleeve_kn <= 0:
                raise ValueError("Диапазоны калибровки должны быть > 0")
            if area_cone_cm2 <= 0 or area_sleeve_cm2 <= 0:
                raise ValueError("Площади должны быть > 0")
            return scale_div, fmax_cone_kn, fmax_sleeve_kn, area_cone_cm2, area_sleeve_cm2
        except Exception as e:
            messagebox.showerror("Ошибка", f"Некорректные параметры пересчёта: {e}")
            return None

    def _safe_sheet_name(self, name: str) -> str:
        bad = '[]:*?/\\'
        for ch in bad:
            name = name.replace(ch, "_")
        name = name.strip()
        if not name:
            name = "Sheet"
        return name[:31]


    def _validate_export_rows(self) -> bool:
        """Блокируем экспорт, если есть строки, где заполнена только одна колонка (qc или fs).
        Подсвечиваем зондирование красным (как ошибка) и показываем предупреждение.
        """
        bad = False
        for t in self.tests:
            tid = t.tid
            fl = self.flags.get(tid) or TestFlags(False, set(), set(), set(), set(), set())
            if not getattr(t, 'export_on', True):
                self.flags[tid] = fl
                continue
            n = max(len(getattr(t, 'qc', []) or []), len(getattr(t, 'fs', []) or []))
            qc_arr = getattr(t, 'qc', []) or []
            fs_arr = getattr(t, 'fs', []) or []
            for i in range(n):
                q = (qc_arr[i].strip() if i < len(qc_arr) and qc_arr[i] is not None else '')
                f = (fs_arr[i].strip() if i < len(fs_arr) and fs_arr[i] is not None else '')
                q_filled = (q != '' and q != '0')
                f_filled = (f != '' and f != '0')
                if q_filled ^ f_filled:
                    bad = True
                    fl.invalid = True
                    break
            self.flags[tid] = fl
        if bad:
            self._redraw()
            messagebox.showwarning("Экспорт невозможен",
                                   "Есть строки, где заполнена только одна колонка (qc или fs).\n"
                                   "Заполни вторую колонку или очисти строку полностью.")
            return False
        return True


    def export_excel(self):
        if not self.tests:
            messagebox.showwarning("Нет данных", "Сначала нажми «Показать зондирования»")
            return
        if not self._validate_export_rows():
            return
        params = self._read_calc_params()
        if not params:
            return
        scale_div, fmax_cone_kn, fmax_sleeve_kn, area_cone_cm2, area_sleeve_cm2 = params
        A_cone = _cm2_to_m2(area_cone_cm2)
        A_sleeve = _cm2_to_m2(area_sleeve_cm2)

        tests_exp = [t for t in (getattr(self, 'tests', []) or []) if bool(getattr(t, 'export_on', True))]
        if not tests_exp:
            messagebox.showwarning('Нет данных', 'Нет зондирований для экспорта (все исключены).')
            return

        out = filedialog.asksaveasfilename(
            title="Куда сохранить Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if not out:
            return
        out_path = Path(out)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        used_names = set()

        for t in tests_exp:
            tid = t.tid
            base_name = self._safe_sheet_name(str(tid))
            name = base_name
            k = 1
            while name in used_names:
                suffix = f"_{k}"
                name = self._safe_sheet_name(base_name[: (31 - len(suffix))] + suffix)
                k += 1
            used_names.add(name)

            ws = wb.create_sheet(title=name)
            if getattr(self, "geo_kind", "K2") == "K4":
                ws.append(["Depth_m", "qc_del", "fs_del", "incl_raw", "qc_MPa", "fs_kPa"])
            else:
                ws.append(["Depth_m", "qc_del", "fs_del", "qc_MPa", "fs_kPa"])

            fl = self.flags.get(tid, TestFlags(False, set(), set(), set(), set(), set()))
            n = len(t.qc)

            for idx in range(n):
                depth_val = _parse_depth_float(t.depth[idx]) if idx < len(t.depth) else None
                qc_del = _parse_cell_int(t.qc[idx])
                fs_del = _parse_cell_int(t.fs[idx])

                # skip fully empty rows (and deleted rows are removed anyway)
                if depth_val is None and qc_del is None and fs_del is None:
                    continue

                qc_MPa = None
                fs_kPa = None
                if qc_del is not None and A_cone:
                    F_cone_N = (qc_del / scale_div) * (fmax_cone_kn * 1000.0)
                    qc_MPa = (F_cone_N / A_cone) / 1e6
                if fs_del is not None and A_sleeve:
                    F_sleeve_N = (fs_del / scale_div) * (fmax_sleeve_kn * 1000.0)
                    fs_kPa = (F_sleeve_N / A_sleeve) / 1e3

                ws.append([
                    depth_val,
                    qc_del, fs_del,
                    None if qc_MPa is None else round(qc_MPa, 2),
                    None if fs_kPa is None else int(round(fs_kPa, 0)),
                ])

                # color cells (B,C)
                r = ws.max_row
                if fl.invalid:
                    ws[f"B{r}"].fill = FILL_RED
                    ws[f"C{r}"].fill = FILL_RED
                else:
                    # original indices don't map after skips; keep only for non-skipped rows
                    # best-effort: if idx within flags
                    if (idx, "qc") in fl.interp_cells:
                        ws[f"B{r}"].fill = FILL_YELLOW
                    if (idx, "fs") in fl.interp_cells:
                        ws[f"C{r}"].fill = FILL_YELLOW
                    if (idx, "qc") in fl.force_cells:
                        ws[f"B{r}"].fill = FILL_BLUE
                    if (idx, "fs") in fl.force_cells:
                        ws[f"C{r}"].fill = FILL_BLUE
                    if (idx, "qc") in getattr(fl, 'user_cells', set()):
                        ws[f"B{r}"].fill = FILL_PURPLE
                    if (idx, "fs") in getattr(fl, 'user_cells', set()):
                        ws[f"C{r}"].fill = FILL_PURPLE

            ws.freeze_panes = "A2"

        ws = wb.create_sheet(title="meta")
        ws.append(["test_id", "datetime", "marker", "header_pos", "points"])
        for t in tests_exp:
            ws.append([t.tid, t.dt or "", t.marker or "", t.header_pos or "", len(t.qc)])

        wb.save(out_path)

    # ---------------- save GEO ----------------


    def _has_issues_for_fix_prompt(self):
        """Detect zeros / non-refusal (max<250) to suggest algorithmic fix before CREDO export."""
        if not getattr(self, "tests", None):
            return False, []
        issues = []
        for t in tests_exp:
            tid = getattr(t, "tid", "?")
            qc_vals = [v for v in (_parse_cell_int(x) for x in getattr(t, "qc", [])) if v is not None]
            fs_vals = [v for v in (_parse_cell_int(x) for x in getattr(t, "fs", [])) if v is not None]
            has_zero = any(v == 0 for v in qc_vals) or any(v == 0 for v in fs_vals)

            mx_qc = max(qc_vals) if qc_vals else None
            mx_fs = max(fs_vals) if fs_vals else None
            not_refusal = False
            if mx_qc is not None and mx_fs is not None:
                not_refusal = (mx_qc < 250 and mx_fs < 250)
            elif mx_qc is not None:
                not_refusal = (mx_qc < 250)
            elif mx_fs is not None:
                not_refusal = (mx_fs < 250)

            if has_zero or not_refusal:
                tag = []
                if has_zero:
                    tag.append("нули")
                if not_refusal:
                    tag.append("не доведено до 250")
                issues.append(f"{tid}: {', '.join(tag)}")
        return (len(issues) > 0), issues

    def export_credo_zip(self):
        """Export each test into two CSV (depth;qc_MPa and depth;fs_kPa) without headers, pack into ZIP.
        Naming: 'СЗ-<№> лоб.csv' and 'СЗ-<№> бок.csv'.
        If issues detected, предлагает исправить и ПОКАЗЫВАЕТ изменения; экспорт нужно нажать повторно.
        """
        if not getattr(self, "tests", None):
            messagebox.showwarning("Нет данных", "Сначала нажми «Показать зондирования»")
            return
        tests_exp = [t for t in (getattr(self, 'tests', []) or []) if bool(getattr(t, 'export_on', True))]
        if not tests_exp:
            messagebox.showwarning('Нет данных', 'Нет зондирований для экспорта (все исключены).')
            return


        params = self._read_calc_params()
        if not params:
            return
        scale_div, fmax_cone_kn, fmax_sleeve_kn, area_cone_cm2, area_sleeve_cm2 = params
        A_cone = _cm2_to_m2(area_cone_cm2)
        A_sleeve = _cm2_to_m2(area_sleeve_cm2)

        # Раньше здесь было предупреждение «Найдены проблемы» и предложение исправлять.
        # По просьбе — убрано: если пользователь нажал «Корректировка»,
        # он сам контролирует изменения. Экспортируем без повторных диалогов.

        out_zip = filedialog.asksaveasfilename(
            title="Куда сохранить ZIP для CREDO",
            defaultextension=".zip",
            filetypes=[("ZIP архив", "*.zip")]
        )
        if not out_zip:
            return

        def fmt_float_comma(x, nd=2):
            if x is None:
                return ""
            s = f"{x:.{nd}f}"
            return s.replace(".", ",")

        def fmt_depth(x):
            if x is None:
                return ""
            s = f"{x:.2f}".replace(".", ",")
            s = s.rstrip("0").rstrip(",")
            return s

        def safe_part(s):
            s = str(s)
            return re.sub(r'[<>:"/\\|?*]+', "_", s)

        tmp_dir = Path(out_zip).with_suffix("")
        if tmp_dir.exists():
            shutil.rmtree(tmp_dir)
        tmp_dir.mkdir(parents=True, exist_ok=True)

        created = []
        for t in tests_exp:
            tid = safe_part(getattr(t, "tid", ""))
            rows_lob = []
            rows_bok = []
            depth_arr = getattr(t, "depth", [])
            qc_arr = getattr(t, "qc", [])
            fs_arr = getattr(t, "fs", [])
            n = max(len(depth_arr), len(qc_arr), len(fs_arr))
            for idx in range(n):
                depth_val = _parse_depth_float(depth_arr[idx]) if idx < len(depth_arr) else None
                if depth_val is None:
                    continue  # удалённые строки не экспортируем
                qc_del = _parse_cell_int(qc_arr[idx]) if idx < len(qc_arr) else None
                fs_del = _parse_cell_int(fs_arr[idx]) if idx < len(fs_arr) else None

                qc_MPa = None
                fs_kPa = None
                if qc_del is not None and A_cone:
                    F_cone_N = (qc_del / scale_div) * (fmax_cone_kn * 1000.0)
                    qc_MPa = (F_cone_N / A_cone) / 1e6
                if fs_del is not None and A_sleeve:
                    F_sleeve_N = (fs_del / scale_div) * (fmax_sleeve_kn * 1000.0)
                    fs_kPa = (F_sleeve_N / A_sleeve) / 1e3

                d_str = fmt_depth(depth_val)
                qc_str = "" if qc_MPa is None else fmt_float_comma(round(qc_MPa, 2), nd=2)
                fs_str = "" if fs_kPa is None else str(int(round(fs_kPa, 0)))
                rows_lob.append(f"{d_str};{qc_str}")
                rows_bok.append(f"{d_str};{fs_str}")

            fn_lob = tmp_dir / f"СЗ-{tid} лоб.csv"
            fn_bok = tmp_dir / f"СЗ-{tid} бок.csv"
            fn_lob.write_text("\n".join(rows_lob) + ("\n" if rows_lob else ""), encoding="utf-8")
            fn_bok.write_text("\n".join(rows_bok) + ("\n" if rows_bok else ""), encoding="utf-8")
            created.extend([fn_lob, fn_bok])

        with zipfile.ZipFile(out_zip, "w", compression=zipfile.ZIP_DEFLATED) as z:
            for f in created:
                z.write(f, arcname=f.name)

        shutil.rmtree(tmp_dir, ignore_errors=True)
        self._credo_force_export = False





    def _center_child(self, win: tk.Toplevel):
        try:
            win.update_idletasks()
            w = win.winfo_reqwidth()
            h = win.winfo_reqheight()
            px = self.winfo_rootx()
            py = self.winfo_rooty()
            pw = self.winfo_width()
            ph = self.winfo_height()
            x = px + max(0, (pw - w) // 2)
            y = py + max(0, (ph - h) // 2)
            win.geometry(f"{w}x{h}+{x}+{y}")
        except Exception:
            pass
    def _place_calendar_near_header(self, dlg: tk.Toplevel, ti: int):
        """Позиционирует календарь у шапки выбранной зондирования (колонки) на Canvas."""
        try:
            dlg.update_idletasks()
            x0, y0, x1, y1 = self._header_bbox(max(0, int(ti)))
            # учесть прокрутку canvas (canvas coords -> screen coords)
            vx = self.canvas.canvasx(0)
            vy = self.canvas.canvasy(0)
            sx = self.canvas.winfo_rootx() + int(x0 - vx) + 10
            sy = self.canvas.winfo_rooty() + int(y0 - vy) + 10
            dlg.geometry(f"+{sx}+{sy}")
        except Exception:
            # fallback: центрируем по основному окну
            try:
                self._center_child(dlg)
            except Exception:
                pass

    def _ensure_object_code(self) -> str:
        """Раньше тут всплывало отдельное окно 'Объект', если поле пустое.
        По новой логике объект задаётся в окне 'Параметры GEO' и может быть пустым.
        """
        self.object_code = (getattr(self, "object_code", "") or "").strip()
        return self.object_code

    def _extract_file_map_text(self) -> str:
        """Возвращает текст FILE MAP из исходника (между маркерами)."""
        try:
            p = Path(__file__)
            src = p.read_text(encoding="utf-8", errors="replace").splitlines()
        except Exception as e:
            return f"Не удалось прочитать исходник для FILE MAP: {e}"

        begin = None
        end = None
        for i, line in enumerate(src):
            if "=== FILE MAP BEGIN ===" in line:
                begin = i
            if "=== FILE MAP END ===" in line:
                end = i
                break
        if begin is None or end is None or end <= begin:
            return "FILE MAP не найден в исходнике."

        # Убираем ведущие "# " для красивого отображения
        out = []
        for line in src[begin+1:end]:
            if line.startswith("# "):
                out.append(line[2:])
            elif line.startswith("#"):
                out.append(line[1:])
            else:
                out.append(line)
        return "\n".join(out).strip()

    def show_file_map(self):
        """Окно просмотра карты проекта (FILE MAP)."""
        win = tk.Toplevel(self)
        win.title("Карта проекта")
        win.geometry("860x640")
        win.transient(self)
        try:
            win.iconbitmap(self._icon_path)  # type: ignore[attr-defined]
        except Exception:
            pass

        top = ttk.Frame(win, padding=(10,10))
        top.pack(fill="both", expand=True)

        hdr = ttk.Frame(top)
        hdr.pack(fill="x")
        ttk.Label(hdr, text="FILE MAP (для разработчика)", font=("Segoe UI", 12, "bold")).pack(side="left")

        def copy_all():
            data = txt.get("1.0", "end-1c")
            self.clipboard_clear()
            self.clipboard_append(data)
            win.update_idletasks()

        ttk.Button(hdr, text="Копировать", command=copy_all).pack(side="right")

        body = ttk.Frame(top)
        body.pack(fill="both", expand=True, pady=(10,0))
        ysb = ttk.Scrollbar(body, orient="vertical")
        ysb.pack(side="right", fill="y")

        txt = tk.Text(body, wrap="none", height=1, yscrollcommand=ysb.set)
        txt.pack(side="left", fill="both", expand=True)
        ysb.config(command=txt.yview)

        content = self._extract_file_map_text()
        txt.insert("1.0", content)
        txt.config(state="disabled")

        # monospace-ish
        try:
            txt.config(font=("Cascadia Mono", 10))
        except Exception:
            pass

        ttk.Button(top, text="Закрыть", command=win.destroy).pack(anchor="e", pady=(10,0))

    def _on_close(self):
        if getattr(self, "_dirty", False):
            ans = messagebox.askyesnocancel(
                "Есть изменения",
                "Есть несохранённые изменения. Экспортировать архив перед выходом?"
            )
            if ans is None:
                return
            if ans:
                if not self.export_bundle():
                    return
        self.destroy()

    def export_bundle(self) -> bool:
        if not self.tests:
            messagebox.showwarning("Нет данных", "Сначала открой файл.")
            return False
        tests_exp = [t for t in (getattr(self, 'tests', []) or []) if bool(getattr(t, 'export_on', True))]
        if not tests_exp:
            messagebox.showwarning('Нет данных', 'Нет зондирований для экспорта (все исключены).')
            return False


        obj = self._ensure_object_code()
        ts = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"{obj}_{ts}.zip"

        out_zip = filedialog.asksaveasfilename(
            title="Сохранить архив экспорта",
            defaultextension=".zip",
            initialfile=default_name,
            filetypes=[("ZIP", "*.zip")],
        )
        if not out_zip:
            return False

        try:
            with tempfile.TemporaryDirectory() as td:
                td_path = Path(td)

                meta_path = td_path / f"{obj}_meta.txt"
                self._write_meta_txt(meta_path)

                gxl_path = td_path / f"{obj}.gxl"
                orig_tests = getattr(self, 'tests', None)
                try:
                    self.tests = list(tests_exp)
                    self.save_gxl_generated(str(gxl_path))
                finally:
                    if orig_tests is not None:
                        self.tests = orig_tests


                geo_path = None
                geo_err = None
                # В архив кладём GEO/GE0 ТОЛЬКО если исходно был открыт GEO/GE0 (есть original_bytes и путь).
                if (not getattr(self, "is_gxl", False)) and getattr(self, "original_bytes", None) and getattr(self, "geo_path", None):
                    try:
                        # Сохраняем с ИСХОДНЫМ именем файла (как просили), чтобы не путаться.
                        geo_out_name = Path(self.geo_path).name
                        geo_path = td_path / geo_out_name
                        # Подготовка данных как в save_geo
                        tests_list = tests_exp
                        prepared = []
                        for t in tests_list:
                            try:
                                d = list(getattr(t, "depth", []) or [])
                                qc = list(getattr(t, "qc", []) or [])
                                fs = list(getattr(t, "fs", []) or [])
                                rows = []
                                n = min(len(d), len(qc), len(fs))
                                for k in range(n):
                                    ds = str(d[k]).strip()
                                    if ds == "":
                                        continue
                                    rows.append((d[k], qc[k], fs[k]))
                                d2 = [r[0] for r in rows]
                                qc2 = [r[1] for r in rows]
                                fs2 = [r[2] for r in rows]
                                prepared.append(TestData(
                                    tid=int(getattr(t, "tid", 0) or 0),
                                    dt=str(getattr(t, "dt", "") or ""),
                                    depth=d2, qc=qc2, fs=fs2,
                                    marker=str(getattr(t, "marker", "") or ""),
                                    header_pos=str(getattr(t, "header_pos", "") or ""),
                                    orig_id=getattr(t, "orig_id", None),
                                    block=getattr(t, "block", None),
                                ))
                            except Exception:
                                prepared.append(t)


                        # --- GEO export safety: use ONLY tests_list (respect delete/copy/export checkbox) ---
                        try:
                            _exp_ids = [int(getattr(tt, 'tid', 0) or 0) for tt in tests_list]
                            _exp_ids = [x for x in _exp_ids if x > 0]
                            _exp_set = set(_exp_ids)
                            prepared = [pp for pp in prepared if int(getattr(pp, 'tid', 0) or 0) in _exp_set]
                            _order = {tid: i for i, tid in enumerate(_exp_ids)}
                            prepared.sort(key=lambda pp: _order.get(int(getattr(pp, 'tid', 0) or 0), 10**9))
                        except Exception:
                            pass
                        # K4: пересборка GEO из шаблона пока не поддержана (другой формат блоков).
                        # Поэтому: для K4 в архив кладём исходный GEO без пересохранения.
                        if getattr(self, 'geo_kind', 'K2') == 'K4':
                            try:
                                if getattr(self, 'original_bytes', None):
                                    geo_path.write_bytes(self.original_bytes)
                                geo_err = None
                                geo_tb = ''
                            except Exception as _e:
                                geo_err = f"{type(_e).__name__}: {_e}"
                                geo_tb = traceback.format_exc()
                                geo_path = None
                        else:
                            blocks_info = list((getattr(self, '_geo_template_blocks_info_full', None) or self._geo_template_blocks_info) or [])
                            if not blocks_info:
                                raise RuntimeError('Не удалось найти блоки опытов в исходном файле.')

                            geo_bytes = _rebuild_geo_from_template(self.original_bytes, blocks_info, prepared)
                            geo_path.write_bytes(geo_bytes)

                        # DEBUG: сверка количества/номеров блоков в собранном GEO (ловим "воскресший первый опыт")
                        try:
                            _pos = [m.start() for m in re.finditer(b'\xFF\xFF.\x1E\x0A', geo_bytes)]
                            _ids = [geo_bytes[p+2] for p in _pos]
                            _exp = [int(getattr(tt, 'tid', 0) or 0) for tt in tests_list]
                            dbg_path = td_path / f"{obj}_geo_debug.txt"
                            dbg_path.write_text(
                                "expected_ids=" + ",".join(map(str,_exp)) + "\n" +
                                "actual_ids=" + ",".join(map(str,_ids)) + "\n" +
                                f"expected_n={len(_exp)} actual_n={len(_ids)}\n" +
                                "tests_current_ids=" + ",".join(map(str, [int(getattr(tt,'tid',0) or 0) for tt in (self.tests or [])])) + "\n" +
                                "export_enabled_ids=" + ",".join(map(str, [int(getattr(tt,'tid',0) or 0) for tt in tests_list])) + "\n" +
                                "prepared_ids=" + ",".join(map(str, [int(getattr(tt,'tid',0) or 0) for tt in (prepared or [])])) + "\n" +
                                f"blocks_info_n={len(blocks_info)} using_full_template={'1' if bool(getattr(self,'_geo_template_blocks_info_full',[])) else '0'}\n",
                                encoding="utf-8"
                            )
                        except Exception:
                            dbg_path = None

                    except Exception as _e:
                        geo_err = f"{type(_e).__name__}: {_e}"
                        geo_tb = traceback.format_exc()
                        geo_path = None
                # Если GEO не собрался — сохраняем лог, чтобы понять причину.
                geo_log_path = None
                if geo_err:
                    geo_log_path = td_path / f"{obj}_geo_export_error.txt"
                    try:
                        geo_log_path.write_text(geo_err + "\n\n" + geo_tb, encoding="utf-8")
                    except Exception:
                        pass

                xlsx_path = td_path / f"{obj}.xlsx"
                self._export_excel_silent(xlsx_path)

                credo_zip_path = td_path / f"{obj}_CREDO.zip"
                self._export_credo_silent(credo_zip_path)

                with zipfile.ZipFile(out_zip, "w", compression=zipfile.ZIP_DEFLATED) as z:
                    z.write(meta_path, meta_path.name)
                    z.write(gxl_path, gxl_path.name)
                    if geo_path and geo_path.exists():
                        z.write(geo_path, geo_path.name)
                    if geo_log_path and geo_log_path.exists():
                        z.write(geo_log_path, geo_log_path.name)

                    if 'dbg_path' in locals() and dbg_path and Path(dbg_path).exists():
                        z.write(dbg_path, Path(dbg_path).name)

                    if xlsx_path.exists():
                        z.write(xlsx_path, xlsx_path.name)
                    if credo_zip_path.exists():
                        z.write(credo_zip_path, credo_zip_path.name)

            # Excel не любит открывать .xlsx прямо ИЗ zip (появляется путь вида ...zip.8a3\file.xlsx и файл недоступен).
            # Поэтому дополнительно сохраняем копию XLSX рядом с архивом, чтобы открывалась без проблем.
            try:
                side_xlsx = Path(out_zip).with_suffix('.xlsx')
                if xlsx_path.exists():
                    shutil.copy2(xlsx_path, side_xlsx)
            except Exception:
                pass


            try:
                _log_event(self.usage_logger, "EXPORT", object=self._ensure_object_code(), zip=str(out_zip))
            except Exception:
                pass

            self.status.config(text=f"Экспорт-архив сохранён: {Path(out_zip).name}")
            if geo_err:
                messagebox.showwarning(
                    "GEO не добавлен в архив",
                    "Не удалось собрать GEO из шаблона. В архив добавлен лог: "
                    f"{obj}_geo_export_error.txt",
                )
            self._dirty = False
            return True
        except Exception as e:
            messagebox.showerror("Ошибка экспорта", str(e))
            return False

    def _write_meta_txt(self, path: Path):
        scale = self.scale_var.get().strip() if hasattr(self, "scale_var") else ""
        fcone = self.fcone_var.get().strip() if hasattr(self, "fcone_var") else ""
        fsleeve = self.fsleeve_var.get().strip() if hasattr(self, "fsleeve_var") else ""
        step = getattr(self, "step_m", None)
        depth0 = getattr(self, "depth_start", None)
        src = getattr(self, "loaded_path", "")
        lines = [
            f"object={self._ensure_object_code()}",
            f"source={src}",
            f"scale={scale}",
            f"fcone_kN={fcone}",
            f"fsleeve_kN={fsleeve}",
            f"depth_start_m={depth0 if depth0 is not None else ''}",
            f"step_m={step if step is not None else ''}",
            f"tests={len([t for t in self.tests if bool(getattr(t, 'export_on', True))])}",
        ]
        path.write_text("\n".join(lines), encoding="utf-8")
    def _export_excel_silent(self, out_path: Path):
        """Тихий экспорт в Excel без диалогов (для экспорта-архива)."""
        from openpyxl import Workbook
        wb = Workbook()
        ws_meta = wb.active
        ws_meta.title = "meta"

        def _get(v):
            try:
                return v.get().strip()
            except Exception:
                return str(v).strip()

        ws_meta.append(["object", self._ensure_object_code()])
        ws_meta.append(["source", getattr(self, "loaded_path", "")])
        ws_meta.append(["scale", _get(self.scale_var) if hasattr(self, "scale_var") else ""])
        ws_meta.append(["fcone_kN", _get(self.fcone_var) if hasattr(self, "fcone_var") else ""])
        ws_meta.append(["fsleeve_kN", _get(self.fsleeve_var) if hasattr(self, "fsleeve_var") else ""])
        ws_meta.append(["depth_start_m", getattr(self, "depth_start", "")])
        ws_meta.append(["step_m", getattr(self, "step_m", "")])
        tests_exp = [t for t in (getattr(self, "tests", []) or []) if bool(getattr(t, "export_on", True))]
        ws_meta.append(["tests", len(tests_exp)])

        for t in tests_exp:
            ws = wb.create_sheet(f"Z{getattr(t, 'tid', '')}")
            if getattr(self, "geo_kind", "K2") == "K4":
                ws.append(["Depth_m", "qc_del", "fs_del", "incl_raw", "qc_MPa", "fs_kPa"])
            else:
                ws.append(["Depth_m", "qc_del", "fs_del", "qc_MPa", "fs_kPa"])
            for d, qc, fs in zip(getattr(t, "depth", []) or [], getattr(t, "qc", []) or [], getattr(t, "fs", []) or []):
                depth_val = _parse_depth_float(d)
                qv = _parse_cell_int(qc)
                fv = _parse_cell_int(fs)
                if qv is None: qv = 0
                if fv is None: fv = 0
                qc_mpa, fs_kpa = self._calc_qc_fs_from_del(int(qv), int(fv))
                ws.append([round(depth_val, 2) if depth_val is not None else None, qv, fv, round(qc_mpa, 2), int(round(fs_kpa, 0))])
        wb.save(str(out_path))

    def _export_credo_silent(self, out_zip_path: Path):
        """Тихий экспорт ZIP для CREDO (две CSV на опыт) без диалогов (для экспорта-архива)."""
        import zipfile

        tests_exp = [t for t in (getattr(self, "tests", []) or []) if bool(getattr(t, "export_on", True))]
        if not tests_exp:
            # нечего экспортировать
            with zipfile.ZipFile(out_zip_path, "w", compression=zipfile.ZIP_DEFLATED) as z:
                pass
            return

        def fmt_comma(x, nd=2):
            s = f"{x:.{nd}f}"
            return s.replace(".", ",")

        def fmt_depth(x):
            # GeoExplorer/CREDO ожидает глубину с 2 знаками после запятой
            return f"{x:.2f}".replace(".", ",")

        with zipfile.ZipFile(str(out_zip_path), "w", compression=zipfile.ZIP_DEFLATED) as z:
            for t in tests_exp:
                tid = str(getattr(t, "tid", ""))
                qc_lines = []
                fs_lines = []
                depth_arr = getattr(t, "depth", []) or []
                qc_arr = getattr(t, "qc", []) or []
                fs_arr = getattr(t, "fs", []) or []
                n = max(len(depth_arr), len(qc_arr), len(fs_arr))
                for i in range(n):
                    d = _parse_depth_float(depth_arr[i]) if i < len(depth_arr) else None
                    qv = _parse_cell_int(qc_arr[i]) if i < len(qc_arr) else None
                    fv = _parse_cell_int(fs_arr[i]) if i < len(fs_arr) else None
                    if d is None and qv is None and fv is None:
                        continue
                    if d is None:
                        continue
                    if qv is None: qv = 0
                    if fv is None: fv = 0
                    qc_mpa, fs_kpa = self._calc_qc_fs_from_del(int(qv), int(fv))
                    qc_lines.append(f"{fmt_depth(d)};{fmt_comma(qc_mpa, 2)}")
                    fs_lines.append(f"{fmt_depth(d)};{int(round(fs_kpa))}")
                z.writestr(f"СЗ-{tid} лоб.csv", "\n".join(qc_lines))
                z.writestr(f"СЗ-{tid} бок.csv", "\n".join(fs_lines))
    def save_file(self):
        """Сохранение.

        - Если открыт GXL: сохраняем только в GXL (как в исходном файле).
        - Если открыт GEO/GE0: можно сохранить обратно в GEO/GE0 ИЛИ экспортнуть в GXL.
        """
        try:
            if not getattr(self, 'tests', None):
                messagebox.showwarning("Внимание", "Нет данных для сохранения.")
                return

            if getattr(self, "is_gxl", False) or (getattr(self, "geo_path", None) and str(self.geo_path).lower().endswith(".gxl")):
                return self.save_gxl()

            from tkinter import filedialog
            import os
            base = os.path.basename(str(getattr(self, "geo_path", "data.geo") or "data.geo"))
            base_noext = os.path.splitext(base)[0]
            out_file = getattr(self, '_save_geo_path_override', None)
            if not out_file:
                out_file = filedialog.asksaveasfilename(
                title="Сохранить",
                defaultextension=".geo",
                initialfile=base_noext + ".geo",
                filetypes=[("GEO/GE0", "*.geo *.ge0 *.GEO *.GE0"), ("GXL", "*.gxl *.GXL"), ("Все файлы", "*.*")],
            )
            if not out_file:
                return
            ext = os.path.splitext(out_file)[1].lower()
            if ext == ".gxl":
                return self.save_gxl_generated(out_file)
            else:
                # save back to GEO via template
                self._save_geo_path_override = out_file
                try:
                    return self.save_geo()
                finally:
                    self._save_geo_path_override = None
        except Exception:
            import traceback
            messagebox.showerror("Ошибка", traceback.format_exc())
    def save_gxl(self):
        """Сохранение текущих данных в GXL (XML). Используется, когда открыт .gxl."""
        try:
            if not getattr(self, 'tests', None):
                messagebox.showwarning("Внимание", "Сначала загрузите файл GXL.")
                return
            if not getattr(self, 'geo_path', None) or self.geo_path.suffix.lower() != ".gxl":
                messagebox.showwarning("Внимание", "Сохранение в GXL доступно только при открытом файле .gxl.")
                return

            from tkinter import filedialog
            import os
            base = os.path.basename(str(self.geo_path))
            out_file = getattr(self, '_save_geo_path_override', None)
            if not out_file:
                out_file = filedialog.asksaveasfilename(
                title="Сохранить GXL",
                defaultextension=".gxl",
                initialfile=base,
                filetypes=[('GXL', '*.gxl *.GXL'), ('Все файлы', '*.*')]
            )
            if not out_file:
                return

            # Читаем исходный XML, чтобы сохранить «шапку» и прочие поля.
            xml_text = _decode_xml_bytes(self.geo_path.read_bytes())
            root = ET.fromstring(xml_text)
            obj = root.find(".//object")
            if obj is None:
                messagebox.showerror("Ошибка", "Не удалось найти узел <object> в GXL.")
                return

            # Индекс тестов в XML по numtest
            xml_tests = {}
            for xt in obj.findall("test"):
                num = (xt.findtext("numtest") or "").strip()
                if not num:
                    continue
                try:
                    tid = int(float(num.replace(",", ".")))
                except Exception:
                    continue
                xml_tests[tid] = xt


            # Удаляем из XML те опыты, которых больше нет в текущем списке (иначе они "воскресают" при сохранении)
            keep_tids = set()
            for t in self.tests:
                try:
                    keep_tids.add(int(getattr(t, "tid", 0) or 0))
                except Exception:
                    pass
            for tid, xt in list(xml_tests.items()):
                if tid not in keep_tids:
                    try:
                        obj.remove(xt)
                    except Exception:
                        pass
                    xml_tests.pop(tid, None)

            # Обновляем данные опытов
            for t in self.tests:
                tid = int(getattr(t, "tid", 0) or 0)
                xt = xml_tests.get(tid)
                if xt is None:
                    # Если теста нет в исходном XML — пропускаем (без добавления новых узлов, чтобы не ломать формат)
                    continue

                # deepbegin = первая глубина
                d0 = _parse_depth_float(t.depth[0]) if getattr(t, "depth", None) else None
                if d0 is None:
                    d0 = 0.0

                # stepzond = шаг по двум первым глубинам (если есть)
                step = None
                if getattr(t, "depth", None) and len(t.depth) >= 2:
                    a = _parse_depth_float(t.depth[0])
                    b = _parse_depth_float(t.depth[1])
                    if a is not None and b is not None:
                        step = b - a
                if step is None:
                    step = float(getattr(self, "step_m", 0.05) or 0.05)

                def _set_child_text(parent, tag, text):
                    node = parent.find(tag)
                    if node is None:
                        node = ET.SubElement(parent, tag)
                    node.text = str(text)

                _set_child_text(xt, "deepbegin", f"{d0:.2f}")
                _set_child_text(xt, "stepzond", f"{step:.2f}".rstrip('0').rstrip('.') if step is not None else "0.05")

                # dat: строки qc;fs
                qc = list(getattr(t, "qc", []) or [])
                fs = list(getattr(t, "fs", []) or [])
                n = min(len(qc), len(fs))
                out_lines = []
                for i in range(n):
                    qv = _parse_cell_int(qc[i])
                    fv = _parse_cell_int(fs[i])
                    qs = "" if qv is None else str(int(qv))
                    fs_s = "" if fv is None else str(int(fv))
                    out_lines.append(f"{qs};{fs_s}")
                dat_text = "\n".join(out_lines)
                _set_child_text(xt, "dat", dat_text)

            # Пишем файл
            ET.indent(root, space="  ")
            xml_out = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            with open(out_file, "wb") as f:
                f.write(xml_out)

            self._update_status_loaded(prefix=f"Сохранено GXL: {out_file} | опытов {len(self.tests)}")
        except Exception:
            import traceback
            messagebox.showerror("Ошибка сохранения GXL", traceback.format_exc())


def save_gxl_generated(self, out_file: str):
    """Сформировать GXL максимально идентичный экспорту GeoExplorer.

    См. описание внутри функции.
    """
    try:
        import math
        if not getattr(self, 'tests', None):
            from tkinter import messagebox
            messagebox.showwarning('Внимание', 'Нет данных для экспорта в GXL.')
            return

        def xml_escape(t: str) -> str:
            t = '' if t is None else str(t)
            return (t.replace('&', '&amp;')
                     .replace('<', '&lt;')
                     .replace('>', '&gt;')
                     .replace('"', '&quot;')
                     .replace("'", '&apos;'))

        def fmt_comma_num(x, ndp=2):
            try:
                x = float(x)
            except Exception:
                x = 0.0
            s = f"{x:.{ndp}f}".rstrip('0').rstrip('.')
            if s == '':
                s = '0'
            return s.replace('.', ',')

        def fmt_date(dt_s: str) -> str:
            import re
            s = (dt_s or '').strip()
            if not s:
                return ''
            p = s.split()[0]
            if re.match(r'^\d{2}\.\d{2}\.\d{4}$', p):
                return p
            if re.match(r'^\d{4}-\d{2}-\d{2}$', p):
                y, mo, d = p.split('-')
                return f"{d}.{mo}.{y}"
            return p

        # Параметры из UI (если есть)
        scale = (self.scale_var.get().strip() if getattr(self, 'scale_var', None) else '') or '250'
        fcone = (self.fcone_var.get().strip() if getattr(self, 'fcone_var', None) else '') or '30'
        fsleeve = (self.fsleeve_var.get().strip() if getattr(self, 'fsleeve_var', None) else '') or '10'

        # Числовая шкала для отсечения значений (GeoExplorer часто не принимает значения выше шкалы)
        try:
            _scale_int = int(str(scale).strip().replace(',', '.').split('.')[0])
        except Exception:
            _scale_int = 250
        if _scale_int <= 0:
            _scale_int = 250

        # object поля: фиксированные как просил
        obj_id = '60'
        obj_name = 'name'
        privazka = 'По плану...'

        # шаг
        step_cm = getattr(self, 'step_cm', None)
        step_m_default = 0.10 if step_cm == 10 else (0.05 if step_cm == 5 else 0.10)

        # Сортировка опытов по времени (как в UI)
        tests = list(self.tests)
        try:
            tests = sorted(tests, key=lambda t: (t.dt or ''))
        except Exception:
            pass

        def parse_d(x):
            try:
                if x is None:
                    return None
                s = str(x).strip().replace(',', '.')
                if s == '':
                    return None
                return float(s)
            except Exception:
                return None

        # max depth
        max_depth = 0.0
        for t in tests:
            ds = [parse_d(v) for v in (getattr(t, 'depth', []) or [])]
            ds = [d for d in ds if d is not None]
            if ds:
                max_depth = max(max_depth, max(ds))

        step_m = step_m_default
        for t in tests:
            ds = [parse_d(v) for v in (getattr(t, 'depth', []) or [])]
            ds = [d for d in ds if d is not None]
            if len(ds) >= 2:
                st = abs(ds[1] - ds[0])
                if 0.045 <= st <= 0.055:
                    step_m = 0.05
                    break
                if 0.095 <= st <= 0.105:
                    step_m = 0.10
                    break

        n_rows = int(math.floor(max_depth / step_m + 1e-9)) + 1
        grid = [round(i * step_m, 2) for i in range(max(1, n_rows))]

        footer = (
            '    <iges></iges>\r\n'
            '    <rdirectory>\r\n'
            '      <set>\r\n'
            '        <name>Уральский</name>\r\n'
            '        <psr>0,2</psr>\r\n'
            '        <pml>0,5</pml>\r\n'
            '        <ppy>0,9</ppy>\r\n'
            '        <supes>1,5</supes>\r\n'
            '        <sugl>5</sugl>\r\n'
            '        <glina>5</glina>\r\n'
            '      </set>\r\n'
            '      <set>\r\n'
            '        <name>Тюменский</name>\r\n'
            '        <psr>0,2</psr>\r\n'
            '        <pml>0,5</pml>\r\n'
            '        <ppy>0,9</ppy>\r\n'
            '        <supes>1,5</supes>\r\n'
            '        <sugl>2,8</sugl>\r\n'
            '        <glina>2,8</glina>\r\n'
            '      </set>\r\n'
            '      <set>\r\n'
            '        <name>Уральский1</name>\r\n'
            '        <psr>0,2</psr>\r\n'
            '        <pml>0,5</pml>\r\n'
            '        <ppy>0,9</ppy>\r\n'
            '        <supes>4</supes>\r\n'
            '        <sugl>5</sugl>\r\n'
            '        <glina>5</glina>\r\n'
            '      </set>\r\n'
            '    </rdirectory>\r\n'
        )

        out = []
        out.append('<?xml version="1.0"?>\r\n')
        out.append('<exportfile>\r\n')
        out.append('  <verfile>3</verfile>\r\n')
        out.append('  <verprogram>GeoExplorer v3.0.14.523</verprogram>\r\n')
        out.append('  <object>\r\n')
        out.append(f'    <id>{xml_escape(obj_id)}</id>\r\n')
        out.append(f'    <name>{xml_escape(obj_name)}</name>\r\n')
        out.append('    <h_abs_plan>0</h_abs_plan>\r\n')
        out.append('    <FullName></FullName>\r\n')
        out.append('    <NumArch></NumArch>\r\n')
        out.append('    <Cashman></Cashman>\r\n')
        out.append('    <Appendix></Appendix>\r\n')

        def make_dat(t):
            ds = [parse_d(v) for v in (getattr(t, 'depth', []) or [])]
            qc = getattr(t, 'qc', []) or []
            fs = getattr(t, 'fs', []) or []

            ds2 = [d for d in ds if d is not None]
            if not ds2:
                return ["0;0;0;0;0;"]

            # deepbegin/step: как в GeoExplorer — dat начинается с deepbegin
            deepbegin = min(ds2)
            # шаг пробуем взять из первых двух валидных глубин
            step = step_m
            if len(ds2) >= 2:
                st = abs(ds2[1] - ds2[0])
                if 0.045 <= st <= 0.055:
                    step = 0.05
                elif 0.095 <= st <= 0.105:
                    step = 0.10

            endd = max(ds2)
            n = int(round((endd - deepbegin) / step)) + 1
            if n < 1:
                n = 1

            grid_local = [round(deepbegin + i * step, 2) for i in range(n)]

            m = {}
            for i, d in enumerate(ds):
                if d is None:
                    continue
                key = round(d, 2)
                try:
                    q = int(str(qc[i]).strip() or '0') if i < len(qc) else 0
                except Exception:
                    q = 0
                try:
                    f = int(str(fs[i]).strip() or '0') if i < len(fs) else 0
                except Exception:
                    f = 0

                # отсечь по шкале и запретить отрицательные
                try:
                    q = int(q)
                except Exception:
                    q = 0
                try:
                    f = int(f)
                except Exception:
                    f = 0
                if q < 0:
                    q = 0
                if f < 0:
                    f = 0
                if q > _scale_int:
                    q = _scale_int
                if f > _scale_int:
                    f = _scale_int

                m[key] = (q, f)

            # GeoExplorer: qc;fs;0;0;0; (с завершающей ;)
            return [f"{m.get(d,(0,0))[0]};{m.get(d,(0,0))[1]};0;0;0;" for d in grid_local]

        for idx, t in enumerate(tests, start=1):
            numtest = getattr(t, 'tid', idx)
            dt_s = fmt_date(getattr(t, 'dt', '') or '')
            out.append('    <test>\r\n')
            out.append(f'      <numtest>{numtest}</numtest>\r\n')
            out.append('      <woker>Katko E</woker>\r\n')
            out.append(f'      <privazka>{xml_escape(privazka)}</privazka>\r\n')
            out.append(f'      <date>{xml_escape(dt_s)}</date>\r\n' if dt_s else '      <date></date>\r\n')
            out.append('      <sechsvai>0,3</sechsvai>\r\n')
            out.append(f'      <scaleostria>{xml_escape(fcone)}</scaleostria>\r\n')
            out.append(f'      <scalemufta>{xml_escape(fsleeve)}</scalemufta>\r\n')
            deepbegin_val = 0.0
            try:
                ds2 = [parse_d(v) for v in (getattr(t, 'depth', []) or [])]
                ds2 = [d for d in ds2 if d is not None]
                if ds2:
                    deepbegin_val = min(ds2)
            except Exception:
                deepbegin_val = 0.0
            out.append(f"      <deepbegin>{fmt_comma_num(deepbegin_val, 1)}</deepbegin>\r\n")
            step_for_test = step_m
            try:
                ds2 = [parse_d(v) for v in (getattr(t, 'depth', []) or [])]
                ds2 = [d for d in ds2 if d is not None]
                if len(ds2) >= 2:
                    st = abs(ds2[1] - ds2[0])
                    if 0.045 <= st <= 0.055:
                        step_for_test = 0.05
                    elif 0.095 <= st <= 0.105:
                        step_for_test = 0.10
            except Exception:
                pass
            out.append(f"      <stepzond>{fmt_comma_num(step_for_test, 2)}</stepzond>\r\n")
            out.append('      <sands>False</sands>\r\n')
            out.append('      <h_abs>0</h_abs>\r\n')
            out.append(f'      <scale>{xml_escape(scale)}</scale>\r\n')
            out.append('      <controllertype>1</controllertype>\r\n')
            out.append('      <id_zond>10</id_zond>\r\n')

            dat_lines = make_dat(t)
            if dat_lines:
                out.append(f'      <dat>{dat_lines[0]}\r\n')
                for ln in dat_lines[1:]:
                    out.append(f'{ln}\r\n')
                out.append('      </dat>\r\n')
            else:
                out.append('      <dat>0;0;0;0;0;\r\n')
                out.append('      </dat>\r\n')

            out.append('      <soils></soils>\r\n')
            out.append('      <rtable></rtable>\r\n')
            out.append('    </test>\r\n')

        out.append(footer)
        out.append('  </object>\r\n')
        out.append('</exportfile>\r\n')

        Path(out_file).write_bytes(''.join(out).encode('cp1251', errors='replace'))
        try:
            self.status.set(f"GXL сохранён: {out_file} | шкала={scale} Fкон={fcone} Fмуф={fsleeve}")
        except Exception:
            pass

    except Exception:
        import traceback
        from tkinter import messagebox
        messagebox.showerror('Ошибка', 'Не удалось сформировать GXL.\n\n' + traceback.format_exc())



    def save_geo(self):
        # Сохранение GEO/GE0 через шаблон исходного файла (самый надёжный вариант)
        try:
            if not getattr(self, 'tests', None):
                messagebox.showwarning("Внимание", "Сначала загрузите файл GEO/GE0.")
                return
            if not getattr(self, 'original_bytes', None):
                messagebox.showwarning("Внимание", "Экспорт в GEO/GE0 доступен только после загрузки исходного GEO/GE0.")
                return

            from tkinter import filedialog
            import os
            base = os.path.basename(getattr(self, 'loaded_path', '') or 'export.GEO')
            out_file = getattr(self, '_save_geo_path_override', None)
            if not out_file:
                out_file = filedialog.asksaveasfilename(
                title="Сохранить GEO/GE0",
                defaultextension=os.path.splitext(base)[1] or '.GEO',
                initialfile=base,
                filetypes=[('GEO/GE0', '*.GEO *.GE0'), ('Все файлы', '*.*')]
            )
            if not out_file:
                return

            tests_list = tests_exp  # файл-ориентированный порядок

            

            # Нормализация длин внутри каждого опыта перед сохранением
            tests_list = [self._normalize_test_lengths(t) for t in tests_list]
# Перед сохранением: выкидываем удалённые строки (где depth пустая) так,
            # чтобы GeoExplorer не подставлял нули на месте "пустот".
            prepared = []
            for t in tests_list:
                try:
                    d = list(getattr(t, "depth", []) or [])
                    qc = list(getattr(t, "qc", []) or [])
                    fs = list(getattr(t, "fs", []) or [])
                    rows = []
                    n = min(len(d), len(qc), len(fs))
                    for k in range(n):
                        ds = str(d[k]).strip()
                        if ds == "":
                            continue
                        rows.append((d[k], qc[k], fs[k]))
                    d2 = [r[0] for r in rows]
                    qc2 = [r[1] for r in rows]
                    fs2 = [r[2] for r in rows]
                    prepared.append(TestData(
                        tid=int(getattr(t, "tid", 0) or 0),
                        dt=str(getattr(t, "dt", "") or ""),
                        depth=d2, qc=qc2, fs=fs2,
                        marker=str(getattr(t, "marker", "") or ""),
                        header_pos=str(getattr(t, "header_pos", "") or ""),
                        orig_id=getattr(t, "orig_id", None),
                        block=getattr(t, "block", None),
                    ))
                except Exception:
                    prepared.append(t)


            # --- GEO export safety: use ONLY tests_list (respect delete/copy/export checkbox) ---
            try:
                _exp_ids = [int(getattr(tt, 'tid', 0) or 0) for tt in tests_list]
                _exp_ids = [x for x in _exp_ids if x > 0]
                _exp_set = set(_exp_ids)
                prepared = [pp for pp in prepared if int(getattr(pp, 'tid', 0) or 0) in _exp_set]
                _order = {tid: i for i, tid in enumerate(_exp_ids)}
                prepared.sort(key=lambda pp: _order.get(int(getattr(pp, 'tid', 0) or 0), 10**9))
            except Exception:
                pass
            if getattr(self, 'geo_kind', 'K2') == 'K4':
                messagebox.showwarning('K4', 'Сохранение K4 GEO пока не поддержано.\nИспользуй Экспорт-архив: там будут XLSX/GXL/CREDO и исходный GEO.')
                return
            blocks_info = list((getattr(self, '_geo_template_blocks_info_full', None) or self._geo_template_blocks_info) or [])
            if not blocks_info:
                messagebox.showerror("Ошибка", "Не удалось найти блоки опытов в исходном файле.")
                return

            out_bytes = _rebuild_geo_from_template(self.original_bytes, blocks_info, prepared)
            with open(out_file, 'wb') as f:
                f.write(out_bytes)

            # Обновляем «шаблон» и блоки после сохранения:
            # это позволяет сохранять GEO повторно (в т.ч. после добавления зондировок),
            # т.к. в новом файле появляются корректные блоки для всех опытов.
            try:
                self.original_bytes = out_bytes
                self.loaded_path = out_file
                # перепарсим только что сохранённый GEO, чтобы получить новые block-метаданные
                _tests2 = parse_geo_with_blocks(out_bytes)
                if _tests2 and len(_tests2) == len(self.tests):
                    for i in range(len(self.tests)):
                        try:
                            self.tests[i].block = getattr(_tests2[i], "block", None)
                            self.tests[i].orig_id = getattr(_tests2[i], "orig_id", getattr(self.tests[i], "orig_id", None))
                        except Exception:
                            pass
            except Exception:
                pass

            try:
                self.status.set(f"Сохранено: {out_file} | опытов: {len(tests_list)}")
            except Exception:
                pass
        except Exception:
            import traceback
            messagebox.showerror("Ошибка сохранения GEO", traceback.format_exc())


# --- bind module-level helpers as methods (fix for indentation) ---
try:
    GeoCanvasEditor.save_gxl_generated = save_gxl_generated  # type: ignore[attr-defined]
except Exception:
    pass


if __name__ == "__main__":
    # CLI helpers for installer/admin
    if "--init-license" in sys.argv:
        try:
            _write_license_file()
            print("OK: license.dat created at", _license_path())
            raise SystemExit(0)
        except SystemExit:
            raise
        except Exception as e:
            print("ERROR: cannot create license.dat:", e)
            raise SystemExit(2)

    if "--open-logs" in sys.argv:
        _open_logs_folder()
        raise SystemExit(0)

    GeoCanvasEditor().mainloop()